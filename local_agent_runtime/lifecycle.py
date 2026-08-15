from __future__ import annotations

import base64
import hashlib
import io
import json
import os
import re
import shutil
import sqlite3
import tempfile
import time
import zipfile
from pathlib import Path
from typing import Any

from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.scrypt import Scrypt
from openpyxl import Workbook, load_workbook

from .storage import AgentPaths, AgentState


def _decode_meta(raw: Any) -> dict[str, Any]:
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str) and raw:
        try:
            decoded = json.loads(raw)
        except json.JSONDecodeError:
            return {}
        return decoded if isinstance(decoded, dict) else {}
    return {}


def _safe_download_filename(value: str, fallback: str, extension: str) -> str:
    stem = Path(str(value or "")).stem
    clean = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._") or fallback
    ext = extension if str(extension).startswith(".") else f".{extension}"
    return f"{clean[:120]}{ext}"


def image_download_filename(
    *,
    output_id: str,
    aspect_ratio: str = "",
    version_metadata: Any = None,
    prompt_metadata: Any = None,
    media_type: str = "image/png",
) -> str:
    extension = {
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/webp": ".webp",
    }.get(str(media_type), ".png")
    display = str(_decode_meta(version_metadata).get("display_name") or "")
    if not display:
        stem = str(_decode_meta(prompt_metadata).get("display_stem") or "")
        if stem:
            suffix = str(aspect_ratio or "").replace(":", "_")
            display = f"{stem}_{suffix}" if suffix else stem
    return _safe_download_filename(display or output_id, output_id, extension)


def prompt_download_filename(
    *,
    prompt_id: str,
    version_metadata: Any = None,
    entry_metadata: Any = None,
) -> str:
    metadata = {
        **_decode_meta(version_metadata),
        **_decode_meta(entry_metadata),
    }
    return _safe_download_filename(
        str(metadata.get("display_stem") or prompt_id),
        prompt_id,
        ".txt",
    )


def _prompt_rows(
    state: AgentState, owner_key: str, run_id: str
) -> list[dict[str, Any]]:
    with state._connect() as conn:
        rows = conn.execute(
            """
            SELECT re.prompt_id, re.resource_id, re.resource_version,
                   re.metadata_json AS entry_metadata,
                   rv.metadata_json AS version_metadata
            FROM run_entries re
            JOIN runs run ON run.run_id = re.run_id
            JOIN resources r ON r.resource_id = re.resource_id
            JOIN resource_versions rv
              ON rv.resource_id = re.resource_id AND rv.version = re.resource_version
            WHERE re.run_id = ? AND run.owner_key = ? AND r.kind = 'prompt'
            ORDER BY re.position
            """,
            (run_id, owner_key),
        ).fetchall()
    return [dict(row) for row in rows]


def export_prompt_xlsx(state: AgentState, owner_key: str, run_id: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Prompts"
    sheet.append(["Prompt ID", "Resource ID", "Expected Version", "Content"])
    for row in _prompt_rows(state, owner_key, run_id):
        content = state.resource_path(
            str(row["resource_id"]), int(row["resource_version"])
        ).read_text(encoding="utf-8")
        sheet.append(
            [
                row["prompt_id"],
                row["resource_id"],
                int(row["resource_version"]),
                content,
            ]
        )
    output = io.BytesIO()
    workbook.save(output)
    with zipfile.ZipFile(output, "a", compression=zipfile.ZIP_DEFLATED) as archive:
        used: set[str] = set()
        for row in _prompt_rows(state, owner_key, run_id):
            name = prompt_download_filename(
                prompt_id=str(row["prompt_id"]),
                version_metadata=row.get("version_metadata"),
                entry_metadata=row.get("entry_metadata"),
            )
            while name in used:
                stem = Path(name).stem
                name = f"{stem}_dup{len(used)}{Path(name).suffix}"
            used.add(name)
            archive.writestr(
                f"prompts/{name}",
                state.resource_path(
                    str(row["resource_id"]), int(row["resource_version"])
                ).read_bytes(),
            )
    return output.getvalue()


def import_prompt_xlsx(
    state: AgentState,
    owner_key: str,
    run_id: str,
    workbook_bytes: bytes,
    *,
    operation_id: str,
) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = workbook["Prompts"] if "Prompts" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    required = ["Prompt ID", "Resource ID", "Expected Version", "Content"]
    if headers != required:
        raise ValueError("Prompt workbook columns are invalid")
    known = {str(row["prompt_id"]): row for row in _prompt_rows(state, owner_key, run_id)}
    updated = 0
    for index, values in enumerate(rows, start=1):
        prompt_id, resource_id, expected_version, content = values
        existing = known.get(str(prompt_id or ""))
        if (
            existing is None
            or str(resource_id or "") != str(existing["resource_id"])
            or not isinstance(content, str)
        ):
            raise ValueError("Prompt workbook contains an unknown or invalid prompt")
        temporary = state.paths.staging / f".xlsx-prompt-{index}-{time.time_ns()}.tmp"
        temporary.write_text(content, encoding="utf-8")
        try:
            version = state.put_resource(
                source=temporary,
                owner_key=owner_key,
                kind="prompt",
                logical_key=str(prompt_id),
                resource_id=str(resource_id),
                expected_version=int(expected_version),
                operation_id=f"{operation_id}:{index}",
                metadata={"run_id": run_id},
                media_type="text/plain; charset=utf-8",
            )
        finally:
            temporary.unlink(missing_ok=True)
        with state._connect() as conn:
            conn.execute(
                """
                UPDATE run_entries SET resource_version = ?
                WHERE run_id = ? AND prompt_id = ? AND resource_id = ?
                """,
                (version.version, run_id, str(prompt_id), str(resource_id)),
            )
        updated += 1
    return {"run_id": run_id, "updated": updated}


def build_output_zip(state: AgentState, owner_key: str, run_id: str) -> bytes:
    prompt_rows = _prompt_rows(state, owner_key, run_id)
    with state._connect() as conn:
        rows = conn.execute(
            """
            SELECT out.output_id, out.current_version, out.aspect_ratio,
                   obj.relative_path, obj.media_type,
                   rv.metadata_json AS version_metadata,
                   (
                       SELECT re.metadata_json FROM run_entries re
                       WHERE re.run_id = out.run_id
                         AND re.prompt_id = out.prompt_id
                         AND re.role = 'prompt'
                       LIMIT 1
                   ) AS prompt_metadata
            FROM outputs out
            JOIN runs run ON run.run_id = out.run_id
            JOIN output_versions ov
              ON ov.output_id = out.output_id AND ov.version = out.current_version
            JOIN resource_versions rv
              ON rv.resource_id = ov.resource_id AND rv.version = ov.resource_version
            JOIN objects obj ON obj.sha256 = rv.object_sha256
            WHERE out.run_id = ? AND run.owner_key = ? AND out.status = 'available'
            ORDER BY out.output_id
            """,
            (run_id, owner_key),
        ).fetchall()
    result = io.BytesIO()
    with zipfile.ZipFile(result, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        used_prompts: set[str] = set()
        for row in prompt_rows:
            name = prompt_download_filename(
                prompt_id=str(row["prompt_id"]),
                version_metadata=row.get("version_metadata"),
                entry_metadata=row.get("entry_metadata"),
            )
            while name in used_prompts:
                name = f"{Path(name).stem}_dup{len(used_prompts)}{Path(name).suffix}"
            used_prompts.add(name)
            archive.writestr(
                f"prompts/{name}",
                state.resource_path(
                    str(row["resource_id"]), int(row["resource_version"])
                ).read_bytes(),
            )
        used_images: set[str] = set()
        for row in rows:
            source = state.paths.root / str(row["relative_path"])
            filename = image_download_filename(
                output_id=str(row["output_id"]),
                aspect_ratio=str(row["aspect_ratio"] or ""),
                version_metadata=row["version_metadata"],
                prompt_metadata=row["prompt_metadata"],
                media_type=str(row["media_type"]),
            )
            folder = str(row["aspect_ratio"] or "4:5").replace(":", "_")
            arcname = f"{folder}/{filename}"
            while arcname in used_images:
                filename = f"{Path(filename).stem}_dup{len(used_images)}{Path(filename).suffix}"
                arcname = f"{folder}/{filename}"
            used_images.add(arcname)
            archive.write(source, arcname=arcname)
    return result.getvalue()


def backup_local_data(paths: AgentPaths, destination: Path) -> dict[str, Any]:
    paths.ensure()
    destination = destination.expanduser().resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary_dir = Path(tempfile.mkdtemp(prefix=".local-backup-", dir=paths.staging))
    snapshot = temporary_dir / "agent.sqlite3"
    manifest: dict[str, Any] = {"format": 1, "files": {}}
    try:
        with sqlite3.connect(paths.database) as source, sqlite3.connect(snapshot) as target:
            source.backup(target)
        files = [snapshot]
        for root_name in ("objects", "config"):
            root = paths.root / root_name
            if root.exists():
                files.extend(path for path in root.rglob("*") if path.is_file())
        temp_archive = destination.with_name(f".{destination.name}.{time.time_ns()}.tmp")
        with zipfile.ZipFile(temp_archive, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for source in files:
                arcname = (
                    "state/agent.sqlite3"
                    if source == snapshot
                    else source.relative_to(paths.root).as_posix()
                )
                digest = hashlib.sha256(source.read_bytes()).hexdigest()
                manifest["files"][arcname] = digest
                archive.write(source, arcname)
            archive.writestr(
                "backup-manifest.json",
                json.dumps(manifest, sort_keys=True, separators=(",", ":")),
            )
        os.replace(temp_archive, destination)
    finally:
        shutil.rmtree(temporary_dir, ignore_errors=True)
    return {
        "status": "verified",
        "file_count": len(manifest["files"]),
        "sha256": hashlib.sha256(destination.read_bytes()).hexdigest(),
    }


def restore_local_data(paths: AgentPaths, source: Path) -> dict[str, Any]:
    source = source.expanduser().resolve()
    staging = Path(tempfile.mkdtemp(prefix=".local-restore-", dir=source.parent))
    try:
        with zipfile.ZipFile(source) as archive:
            manifest = json.loads(archive.read("backup-manifest.json"))
            for name, expected_hash in manifest["files"].items():
                if name.startswith("/") or ".." in Path(name).parts:
                    raise ValueError("Backup contains an unsafe path")
                body = archive.read(name)
                if hashlib.sha256(body).hexdigest() != expected_hash:
                    raise ValueError("Backup verification failed")
                target = staging / name
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_bytes(body)
        paths.root.mkdir(parents=True, exist_ok=True)
        for name in manifest["files"]:
            source_path = staging / name
            target = paths.root / name
            target.parent.mkdir(parents=True, exist_ok=True)
            os.replace(source_path, target)
        paths.ensure()
        return {"status": "restored", "file_count": len(manifest["files"])}
    finally:
        shutil.rmtree(staging, ignore_errors=True)


def _replication_key(secret: bytes, salt: bytes) -> bytes:
    if len(secret) < 32:
        raise ValueError("Replication secret must contain at least 32 bytes")
    return Scrypt(salt=salt, length=32, n=2**14, r=8, p=1).derive(secret)


def export_shared_config(
    state: AgentState,
    *,
    owner_key: str,
    logical_key: str,
    authority_device_id: str,
    approved_device_id: str,
    replication_secret: bytes,
) -> bytes:
    with state._connect() as conn:
        row = conn.execute(
            """
            SELECT resource_id, current_version FROM resources
            WHERE owner_key = ? AND kind = 'config_file' AND logical_key = ?
              AND deleted_at IS NULL
            """,
            (owner_key, logical_key),
        ).fetchone()
    if row is None:
        raise ValueError("Shared config not found")
    metadata = {
        "format": 1,
        "owner_key": owner_key,
        "logical_key": logical_key,
        "authority_device_id": authority_device_id,
        "approved_device_id": approved_device_id,
        "source_resource_id": str(row["resource_id"]),
        "source_version": int(row["current_version"]),
    }
    aad = json.dumps(metadata, sort_keys=True, separators=(",", ":")).encode()
    salt, nonce = os.urandom(16), os.urandom(12)
    content = state.resource_path(
        str(row["resource_id"]), int(row["current_version"])
    ).read_bytes()
    encrypted = AESGCM(_replication_key(replication_secret, salt)).encrypt(nonce, content, aad)
    package = {
        "metadata": metadata,
        "salt": base64.b64encode(salt).decode("ascii"),
        "nonce": base64.b64encode(nonce).decode("ascii"),
        "ciphertext": base64.b64encode(encrypted).decode("ascii"),
    }
    return json.dumps(package, sort_keys=True, separators=(",", ":")).encode()


def import_shared_config(
    state: AgentState,
    package_bytes: bytes,
    *,
    importing_device_id: str,
    replication_secret: bytes,
    operation_id: str,
) -> dict[str, Any]:
    package = json.loads(package_bytes)
    metadata = package["metadata"]
    if metadata.get("approved_device_id") != importing_device_id:
        raise ValueError("This device is not approved for the shared config replica")
    aad = json.dumps(metadata, sort_keys=True, separators=(",", ":")).encode()
    salt = base64.b64decode(package["salt"], validate=True)
    nonce = base64.b64decode(package["nonce"], validate=True)
    ciphertext = base64.b64decode(package["ciphertext"], validate=True)
    content = AESGCM(_replication_key(replication_secret, salt)).decrypt(
        nonce, ciphertext, aad
    )
    temporary = state.paths.staging / f".shared-config-{time.time_ns()}.tmp"
    temporary.write_bytes(content)
    try:
        resource = state.put_resource(
            source=temporary,
            owner_key=str(metadata["owner_key"]),
            kind="config_file",
            logical_key=str(metadata["logical_key"]),
            operation_id=operation_id,
            metadata={
                "authority_device_id": metadata["authority_device_id"],
                "verified_replica_device_id": importing_device_id,
                "source_version": int(metadata["source_version"]),
            },
            media_type="application/json",
        )
    finally:
        temporary.unlink(missing_ok=True)
    return {
        "resource_id": resource.resource_id,
        "version": resource.version,
        "authority_device_id": metadata["authority_device_id"],
        "verified_replica_device_id": importing_device_id,
        "status": "verified",
    }
