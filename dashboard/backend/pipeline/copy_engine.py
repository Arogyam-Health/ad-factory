from __future__ import annotations

import asyncio
import base64
import copy
import hashlib
import importlib.util
import json
import os
import random
import re
import shutil
import subprocess
import sys
import threading
import time
import traceback
import uuid
import urllib.request
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

import httpx
import psutil
from fastapi import Body, FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.responses import FileResponse, JSONResponse, PlainTextResponse, StreamingResponse

from dashboard.backend.pipeline.paths import (
    CONVERT_916_TEMPLATE_PATH,
    COPY_ARCH_PATH,
    COPY_PROMPTS_PATH,
    DEFAULT_GOOGLE_API_URL,
    DEFAULT_GOOGLE_MODEL,
    DEFAULT_IMAGE_SOURCES_FILE,
    DEFAULT_PRODUCT_MASTER,
    ENV_PATH,
    FORMATS,
    GENERATED_IMAGES_ROOT,
    INPUT_IMAGES_DIR,
    INPUT_ROOT,
    LEGACY_ACTIVE_IMAGES_FILE,
    LLM_TRACES_DIR,
    OPENCODE_AD_TIMEOUT_SECONDS,
    OPENCODE_ADS_PER_SESSION_SCHEDULE,
    OPENCODE_MAX_CONCURRENT,
    OPENCODE_QUEUE_DIR,
    PERSONA_SEEDS_PATH,
    ROOT,
    RUNS_ROOT,
    RUNTIME_ROOT,
    STARTING_PROMPT_PATH,
    STORAGE_ROOT,
)
from dashboard.backend.pipeline.user_overrides import (
    clear_user_config_overrides as _clear_user_config_overrides,
    resolve_user_config as _resolve_user_config,
    set_user_config_overrides as _set_user_config_overrides,
)
from dashboard.backend.pipeline.clock import ensure_dirs, load_env_file, make_run_id, now_iso
from dashboard.backend.pipeline.run_control import (
    _cancel_current_run,
    _cancel_events,
    _close_active_httpx_clients,
    _kill_tracked_subprocesses,
    _register_httpx_client,
    _register_subprocess,
    _unregister_httpx_client,
    _unregister_subprocess,
    cancel_event_for_run,
    signal_cancel_current_run,
    signal_cancel_run,
)
from dashboard.backend.pipeline.run_owners import (
    _extract_run_id_from_generated_path,
    _extract_run_id_from_output_path,
    _get_run_owner,
    _manifest_fields_for_db,
    _parse_prompt_meta,
    _persist_run_manifest_db,
    _record_run_owner,
    _resolve_file_owner,
    _resolve_run_owner_scope,
    _store_output_mapping,
    _update_run_status_db,
)
from dashboard.backend.pipeline.copy_text import (
    COPY_ARCH,
    COPY_PROMPTS,
    HYPOTHESIS_VARIABLES,
    PERSONA_SEED_INPUTS,
    _PERSONA_SEED_MAPPING,
    _TESTIMONIAL_GUIDANCE,
    _build_hypothesis_variables,
    _build_persona_payload_field,
    _compact_creative_entry,
    _compact_product_truth,
    _entry_direction,
    _framework_item,
    _headline_architecture_group,
    _hypothesis_guidance,
    _hypothesis_variant_label,
    _invalidate_config_cache,
    _load_copy_architecture,
    _load_copy_prompts,
    _load_persona_seeds,
    _normalize_how_kit_solves,
    _persona_name_to_slug,
    _persona_theme,
    _resolve_copy_architecture,
    _resolve_copy_prompts,
    _resolve_persona_seeds,
    _select_headline_architecture,
    assembler_language_mode,
    build_ad_copy_system_prompt,
    build_ad_prompt_tail,
    build_copy_requirements,
    build_generation_payload_for_llm,
    build_persona_payload,
    build_response_skeleton,
    build_strict_schema_note,
    classify_hook_structure,
    compact_format_rules_for_copy,
    copy_text_for_candidate,
    cta_for_candidate,
    detect_template_leakage,
    extract_generated_ad_candidate,
    filter_valid_ads,
    headline_for_candidate,
    hook_structure_mismatch,
    hydrate_generated_ad_candidate,
    hypothesis_mismatch,
    load_format_visual_archetypes,
    parse_persona_library,
    persona_number_from_slug,
    persona_slug,
    resolve_language_mode,
    validate_generated_copy_payload,
    validate_single_ad,
)
from dashboard.backend.pipeline.input_assets import (
    default_image_sources_file,
    default_product_doc_info,
    list_input_images,
    read_active_images,
    store_uploaded_input_images,
)
from dashboard.backend.pipeline.browser_env import (
    dashboard_subprocess_env,
    debugger_endpoint_reachable,
    detect_wsl_user,
    detect_wsl_windows_host_ip,
    extension_browser_required_for_chatgpt,
    render_chatgpt_uses_local_agent,
    start_extension_cdp_proxy_for_user,
    wsl_chrome_cdp_url,
)
from dashboard.backend.pipeline.text_scrub import (
    PROOF_NOTE_MARKERS,
    choose_text,
    enforce_unique_ctas,
    scrub_on_image_copy,
    shorten_copy_line,
    strip_ansi,
    strip_ba_panel_label,
    strip_internal_marker,
    strip_internal_markers_from_payload,
    strip_price_tokens,
)
from dashboard.backend.services.opencode_catalog import (
    DEFAULT_OPENCODE_API_URL,
    build_opencode_catalog,
    choose_openai_gpt52,
    list_opencode_models,
    sanitize_dashboard_model,
)
from dashboard.backend.db.settings import settings as app_settings

def parse_uniqueness_collisions(error_text: str) -> list[dict[str, Any]]:
    collisions: list[dict[str, Any]] = []
    for raw_line in error_text.splitlines():
        line = raw_line.strip()
        match = re.search(r"ads\[(\d+)\]\.copy\.(EN|HI)\.([a-z_]+)", line)
        if not match:
            continue
        collisions.append(
            {
                "ad_index": int(match.group(1)),
                "language": match.group(2),
                "field": match.group(3),
                "line": line,
            }
        )
    return collisions

def parse_json_object_from_text(content: str) -> dict[str, Any] | None:
    text = (content or "").strip()
    if not text:
        return None

    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    fence = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, flags=re.DOTALL | re.IGNORECASE)
    if fence:
        try:
            parsed = json.loads(fence.group(1))
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass

    decoder = json.JSONDecoder()
    best: dict[str, Any] | None = None
    best_span = -1
    for match in re.finditer(r"\{", text):
        start = match.start()
        try:
            parsed, end = decoder.raw_decode(text[start:])
        except json.JSONDecodeError:
            continue
        if not isinstance(parsed, dict):
            continue
        span = end
        if span > best_span:
            best = parsed
            best_span = span
    return best

def parse_opencode_json_output(stdout: str) -> dict[str, Any] | None:
    text_chunks: list[str] = []

    def collect_text(value: Any) -> None:
        if isinstance(value, str) and value.strip():
            text_chunks.append(value.strip())
            return
        if isinstance(value, list):
            for item in value:
                collect_text(item)
            return
        if not isinstance(value, dict):
            return

        for key in ("text", "content", "message", "output"):
            candidate = value.get(key)
            if isinstance(candidate, str) and candidate.strip():
                text_chunks.append(candidate.strip())
        part = value.get("part")
        if isinstance(part, dict):
            collect_text(part)
        elif isinstance(part, str) and part.strip():
            text_chunks.append(part.strip())

    for raw_line in (stdout or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            event = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(event, str):
            text_chunks.append(event.strip())
            continue
        if isinstance(event, list):
            collect_text(event)
            continue
        if not isinstance(event, dict):
            continue
        event_type = str(event.get("type") or "").lower()
        if event_type and event_type not in {"text", "message", "content", "output"}:
            part = event.get("part")
            if not isinstance(part, (dict, str)):
                continue
        collect_text(event)

    if text_chunks:
        parsed = parse_json_object_from_text("\n".join(text_chunks).strip())
        if parsed is not None:
            return parsed

    return parse_json_object_from_text((stdout or "").strip())

def _find_session_id(value: Any, session_scoped: bool = False) -> str | None:
    if isinstance(value, dict):
        event_type = str(value.get("type") or "").lower()
        scoped = session_scoped or "session" in event_type
        for key in ("sessionID", "sessionId", "session_id"):
            candidate = value.get(key)
            if isinstance(candidate, str) and candidate.strip():
                return candidate.strip()
        if scoped:
            candidate = value.get("id")
            if isinstance(candidate, str) and candidate.strip():
                return candidate.strip()
        for key, nested in value.items():
            nested_scoped = scoped or "session" in str(key).lower()
            found = _find_session_id(nested, nested_scoped)
            if found:
                return found
    elif isinstance(value, list):
        for item in value:
            found = _find_session_id(item, session_scoped)
            if found:
                return found
    return None

def parse_opencode_session_id(stdout: str) -> str | None:
    for raw_line in (stdout or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            event = json.loads(line)
        except json.JSONDecodeError:
            continue
        found = _find_session_id(event)
        if found:
            return found

    match = re.search(r'"session(?:ID|Id|_id)"\s*:\s*"([^"]+)"', stdout or "")
    if match:
        return match.group(1).strip()
    return None

def build_product_doc_bootstrap_prompt() -> str:
    return _resolve_copy_prompts().get("product_doc_bootstrap_prompt", "Read the attached product master doc completely. Return only valid JSON: {\"status\":\"product_doc_loaded\"}.")

def _clean_str(value: Any) -> str:
    return value.strip() if isinstance(value, str) else ""

def _clean_bullets(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    out: list[str] = []
    for item in value:
        if isinstance(item, str) and item.strip():
            out.append(item.strip())
    return out

def concept_ids_from_requirements(copy_req: dict[str, Any]) -> dict[str, str]:
    concept = copy_req.get("concept_variation") if isinstance(copy_req.get("concept_variation"), dict) else {}

    def nested_id(key: str, fallback: str) -> str:
        item = concept.get(key) if isinstance(concept.get(key), dict) else {}
        value = item.get("id") if isinstance(item, dict) else ""
        return _clean_str(value) or fallback

    return {
        "concept_angle": nested_id("concept_angle", "desired_outcome"),
    }

def ensure_testimonial_headline(headline: str, lang: str, persona: dict[str, Any]) -> str:
    clean = shorten_copy_line(headline)
    guidance = _TESTIMONIAL_GUIDANCE
    cfg = guidance.get(lang, guidance.get("EN", {}))
    first_pat = cfg.get("first_person_pattern", "")
    weight_pat = cfg.get("weight_pattern", "")
    suffix = cfg.get("suffix", "")
    desire_template = cfg.get("desire_template", "")
    fallback_text = cfg.get("fallback", "")
    desire_field = cfg.get("desire_field", "")

    if lang == "EN":
        if first_pat and re.search(first_pat, clean, flags=re.IGNORECASE):
            if weight_pat and re.search(weight_pat, clean, flags=re.IGNORECASE):
                return clean
            return shorten_copy_line(f'{clean.rstrip(".")}. {suffix}')
        desire = _clean_str(persona.get(desire_field)).rstrip(".")
        if desire:
            desire_phrase = desire[:1].lower() + desire[1:] if len(desire) > 1 else desire.lower()
            return shorten_copy_line(desire_template.format(desire_phrase=desire_phrase))
        return fallback_text

    if first_pat and re.search(first_pat, clean):
        if weight_pat and re.search(weight_pat, clean):
            return clean
        return shorten_copy_line(f'{clean.rstrip("।")}। {suffix}')
    desire = _clean_str(persona.get(desire_field)).rstrip("।")
    if desire:
        return shorten_copy_line(desire_template.format(desire_phrase=desire))
    return fallback_text

def _persona_number_from_candidate(candidate: dict[str, Any]) -> int | None:
    persona = candidate.get("persona") if isinstance(candidate, dict) else None
    if not isinstance(persona, dict):
        return None
    val = persona.get("number")
    if isinstance(val, int):
        return val
    val = persona.get("persona_number")
    if isinstance(val, int):
        return val
    if isinstance(val, str) and val.strip().isdigit():
        return int(val.strip())
    return None

def _persona_name_from_candidate(candidate: dict[str, Any]) -> str:
    persona = candidate.get("persona") if isinstance(candidate, dict) else None
    if not isinstance(persona, dict):
        return ""
    return _clean_str(persona.get("name") or persona.get("persona_name") or "")

def _build_copy_skeleton(context: dict[str, Any], run_id: str) -> dict[str, Any]:
    ads: list[dict[str, Any]] = []
    token = run_id[-4:]
    for idx, item in enumerate(context.get("ads") or [], start=1):
        persona = item["persona"]
        fmt = item["format"]
        persona_num = int(persona["persona_number"])
        persona_name = persona["persona_name"]
        copy_req = item.get("copy_requirements") if isinstance(item.get("copy_requirements"), dict) else {}
        concept_ids = concept_ids_from_requirements(copy_req)

        pain_en = choose_text(persona.get("pain_points", []), f"Daily routine feels heavy and hard to sustain for persona {persona_num}.")
        desire_en = choose_text(persona.get("core_message", []), "A practical routine that feels easy to follow.")
        friction_en = choose_text(persona.get("objections", []), "Past plans felt too strict and difficult to maintain.")
        proof_en = choose_text(persona.get("trust_anchors", []), "Needs proof through clear structure and believable support.")
        tone_en = "Practical, empathetic, and confidence-building"
        pain_hi = "रोज की वजन-घटाने की दिनचर्या टूटना आसान है।"
        desire_hi = "ऐसा आसान सिस्टम चाहिए जो रोज निभ सके।"
        friction_hi = "पहले के प्लान बहुत सख्त और मुश्किल थे।"
        proof_hi = "साफ कदम, भरोसेमंद सपोर्ट और व्यावहारिक प्रमाण चाहिए।"
        tone_hi = "सरल, भरोसेमंद, और व्यावहारिक"

        if fmt in {"HERO", "UGC"}:
            copy_en: dict[str, Any] = {"headline": "", "support_line": "", "cta": ""}
            copy_hi: dict[str, Any] = {"headline": "", "support_line": "", "cta": ""}
            copy_hing: dict[str, Any] = {"headline": "", "support_line": "", "cta": ""}
        elif fmt in {"BA", "FEAT"}:
            copy_en = {"headline": "", "bullets": [], "cta": ""}
            copy_hi = {"headline": "", "bullets": [], "cta": ""}
            copy_hing = {"headline": "", "bullets": [], "cta": ""}
        else:
            copy_en = {"headline": "", "trust_line": "", "cta": ""}
            copy_hi = {"headline": "", "trust_line": "", "cta": ""}
            copy_hing = {"headline": "", "trust_line": "", "cta": ""}

        ad_payload = {
            "format": fmt,
            "headline_angle": "",
            "concept_angle": concept_ids["concept_angle"],
            "persona": {
                "number": persona_num,
                "name": persona_name,
                "pain_en": pain_en,
                "desire_en": desire_en,
                "friction_en": friction_en,
                "proof_needed_en": proof_en,
                "tone_cue_en": tone_en,
                "pain_hi": pain_hi,
                "desire_hi": desire_hi,
                "friction_hi": friction_hi,
                "proof_needed_hi": proof_hi,
                "tone_cue_hi": tone_hi,
            },
            "copy": {"EN": copy_en, "HI": copy_hi, "HINGLISH": copy_hing},
        }
        hypothesis = item.get("hypothesis") if isinstance(item.get("hypothesis"), dict) else None
        if hypothesis:
            ad_payload["hypothesis"] = hypothesis
        for key in [
            "visual_archetype",
            "visual_pattern_reused_from_run_id",
            "visual_pattern_reuse_key",
            "creative_index",
            "creative_total",
            "background_group_key",
        ]:
            if key in item:
                ad_payload[key] = item[key]
        ads.append(ad_payload)

    return {"default_aspect_ratio": "4:5", "ads": ads}

def normalize_generated_copy(
    generated: dict[str, Any] | None,
    context: dict[str, Any],
    run_id: str,
) -> dict[str, Any]:
    base = _build_copy_skeleton(context, run_id)
    generated = generated or {}
    ads_generated = generated.get("ads") if isinstance(generated.get("ads"), list) else None
    # Normalise flat responses: single ad wrapped in "ad" key, direct flat object, or array
    if ads_generated is None:
        single = generated.get("ad") if isinstance(generated.get("ad"), dict) else None
        if single:
            ads_generated = [single]
        elif isinstance(generated.get("format"), str) and generated.get("format").strip():
            ads_generated = [generated]
        elif isinstance(generated, list):
            ads_generated = generated
        else:
            ads_generated = []
    candidates = ads_generated if isinstance(ads_generated, list) else []
    for cand in candidates:
        if isinstance(cand, dict):
            for lang_copy in (cand.get("copy") or {}).values():
                if isinstance(lang_copy, dict) and "subheadline" in lang_copy:
                    lang_copy["support_line"] = lang_copy.pop("subheadline")

    used_indices: set[int] = set()

    def pick_candidate(fmt: str, persona_num: int, persona_name: str) -> dict[str, Any] | None:
        for idx, cand in enumerate(candidates):
            if idx in used_indices or not isinstance(cand, dict):
                continue
            cand_fmt = _clean_str(cand.get("format")).upper()
            if cand_fmt != fmt:
                continue
            cand_num = _persona_number_from_candidate(cand)
            cand_name = _persona_name_from_candidate(cand)
            if cand_num == persona_num or (cand_name and cand_name.lower() == persona_name.lower()):
                used_indices.add(idx)
                return cand

        for idx, cand in enumerate(candidates):
            if idx in used_indices or not isinstance(cand, dict):
                continue
            cand_fmt = _clean_str(cand.get("format")).upper()
            if cand_fmt == fmt:
                used_indices.add(idx)
                return cand
        return None

    for ad in base.get("ads", []):
        fmt = _clean_str(ad.get("format")).upper()
        persona = ad.get("persona") or {}
        persona_num = int(persona.get("number"))
        persona_name = _clean_str(persona.get("name"))
        candidate = pick_candidate(fmt, persona_num, persona_name)
        if not candidate:
            continue

        hypothesis = candidate.get("hypothesis") if isinstance(candidate.get("hypothesis"), dict) else None
        if hypothesis:
            ad["hypothesis"] = hypothesis

        angle = _clean_str(candidate.get("headline_angle"))
        if angle:
            ad["headline_angle"] = angle

        for key in ["concept_angle"]:
            value = _clean_str(candidate.get(key) or candidate.get("image_description"))
            if value:
                ad[key] = value

        cand_copy = candidate.get("copy") if isinstance(candidate.get("copy"), dict) else {}
        # If the candidate has flat fields at top level (no "copy" key), wrap them into a copy block
        copy_level_keys = {"headline", "subheadline", "support_line", "cta", "body", "bullets", "trust_line", "call_to_action", "image_description"}
        if not cand_copy and any(k in candidate for k in copy_level_keys):
            cand_flat = {k: candidate[k] for k in copy_level_keys if k in candidate}
            for k in list(copy_level_keys):
                candidate.pop(k, None)
            cand_copy = cand_flat
        # If the LLM returned flat copy (no EN/HI/HINGLISH wrapper), treat it as EN
        # and propagate to HI/HINGLISH as fallback so validation doesn't reject
        flat_keys = {"headline", "subheadline", "support_line", "cta", "body", "bullets", "trust_line"}
        if cand_copy and not any(k in cand_copy for k in {"EN", "HI", "HINGLISH"}):
            flat_en = dict(cand_copy)
            cand_copy = {"EN": flat_en, "HI": dict(flat_en), "HINGLISH": dict(flat_en)}
        for lang in ["EN", "HI", "HINGLISH"]:
            if lang not in ad["copy"]:
                continue
            base_lang = ad["copy"][lang]
            src_lang = cand_copy.get(lang) if isinstance(cand_copy.get(lang), dict) else {}

            headline = _clean_str(src_lang.get("headline"))
            cta = _clean_str(src_lang.get("cta") or src_lang.get("call_to_action"))
            if headline:
                base_lang["headline"] = shorten_copy_line(headline)
            if cta:
                base_lang["cta"] = cta

            if fmt == "TEST":
                if not _clean_str(src_lang.get("headline")):
                    base_lang["headline"] = ensure_testimonial_headline(base_lang.get("headline", ""), lang, persona)

            if fmt in {"HERO", "UGC"}:
                support = _clean_str(src_lang.get("support_line")) or _clean_str(src_lang.get("subheadline")) or _clean_str(src_lang.get("body"))
                if support:
                    base_lang["support_line"] = shorten_copy_line(support)
            elif fmt in {"BA", "FEAT"}:
                bullets = _clean_bullets(src_lang.get("bullets"))
                min_bullets = 4 if fmt == "BA" else 2
                if len(bullets) >= min_bullets:
                    if fmt == "BA":
                        bullets = [strip_ba_panel_label(b) for b in bullets]
                    base_lang["bullets"] = [shorten_copy_line(b) for b in bullets]
            else:
                trust = _clean_str(src_lang.get("trust_line") or src_lang.get("body"))
                if trust:
                    base_lang["trust_line"] = shorten_copy_line(trust)

    return base

def _prompt_copy_records_from_mongo(user_id: str, run_id: str) -> list[dict[str, Any]]:
    from dashboard.backend.pipeline.images import extract_on_image_copy_lines, parse_prompt_filename
    from dashboard.backend.pipeline.personas import parse_persona_number_from_prompt
    records: list[dict[str, Any]] = []
    if not user_id:
        return records
    try:
        from dashboard.backend.db.client import get_sync_db
        from dashboard.backend.db.collections import COLL_PROMPTS
        docs = list(
            get_sync_db()[COLL_PROMPTS]
            .find({"user_id": user_id, "run_id": run_id})
            .sort("created_at", 1)
        )
    except Exception:
        return records
    for doc in docs:
        rel_path = str(doc.get("file_path") or "")
        text = str(doc.get("content") or "")
        parsed_name = parse_prompt_filename(rel_path)
        persona_number = parsed_name[2] if parsed_name else None
        if persona_number is None:
            persona_number = parse_persona_number_from_prompt(text)
        records.append({
            "prompt_file": rel_path,
            "format": parsed_name[0] if parsed_name else doc.get("format", ""),
            "language": parsed_name[1] if parsed_name else doc.get("language", ""),
            "persona_number": persona_number,
            "review_url": f"/api/files/download/prompt/{doc.get('prompt_id')}",
            "copy_lines": extract_on_image_copy_lines(text),
        })
    return records

def api_run_prompt_copies(run_id: str, user_id: str = "") -> dict[str, Any]:
    from dashboard.backend.pipeline.images import extract_on_image_copy_lines, parse_prompt_filename
    from dashboard.backend.pipeline.personas import parse_persona_number_from_prompt
    from dashboard.backend.pipeline.runs_db import _check_ownership, api_run, load_manifest_for_run
    if user_id:
        _check_ownership(run_id, user_id)
        manifest = api_run(run_id, user_id=user_id)
        return {"run_id": run_id, "batch": manifest.get("batch"), "prompts": _prompt_copy_records_from_mongo(user_id, run_id)}

    _run_dir, manifest, _has_storage_manifest = load_manifest_for_run(run_id)
    prompt_files_all = manifest.get("prompt_files") or []
    prompt_files = [path for path in prompt_files_all if "/45/" in str(path)] or prompt_files_all
    records: list[dict[str, Any]] = []
    for rel_path in prompt_files:
        prompt_path = ROOT / rel_path
        if not prompt_path.exists() or not prompt_path.is_file():
            continue
        text = prompt_path.read_text(encoding="utf-8", errors="ignore")
        parsed_name = parse_prompt_filename(rel_path)
        persona_number = parsed_name[2] if parsed_name else None
        if persona_number is None:
            persona_number = parse_persona_number_from_prompt(text)
        records.append(
            {
                "prompt_file": rel_path,
                "format": parsed_name[0] if parsed_name else "",
                "language": parsed_name[1] if parsed_name else "",
                "persona_number": persona_number,
                "review_url": "/output/" + rel_path.replace("output/", ""),
                "copy_lines": extract_on_image_copy_lines(text),
            }
        )

    if not records:
        user_id = _get_run_owner(run_id) or ""
        if user_id:
            records.extend(_prompt_copy_records_from_mongo(user_id, run_id))

    return {"run_id": run_id, "batch": manifest.get("batch"), "prompts": records}

def api_run_update_prompt_copies(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    from dashboard.backend.pipeline.generation import generate_916_for_run, rerender_prompts_for_run
    from dashboard.backend.pipeline.images import collect_run_result, parse_prompt_filename
    from dashboard.backend.pipeline.runs_db import load_run_language_mode, merge_manifest
    run_dir = RUNS_ROOT / run_id
    manifest_path = run_dir / "manifest.json"
    copy_path = run_dir / "context" / "copy_batch.json"

    if not manifest_path.exists():
        raise HTTPException(status_code=404, detail="Run not found")
    if not copy_path.exists():
        raise HTTPException(status_code=404, detail="copy_batch.json not found for run")

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    batch = (manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    edits = payload.get("edits")
    if not isinstance(edits, list) or not edits:
        raise HTTPException(status_code=400, detail="edits must be a non-empty array")

    copy_json = json.loads(copy_path.read_text(encoding="utf-8"))
    ads = copy_json.get("ads")
    if not isinstance(ads, list) or not ads:
        raise HTTPException(status_code=400, detail="Invalid copy batch payload")

    updated_count = 0
    for entry in edits:
        if not isinstance(entry, dict):
            continue
        prompt_file = str(entry.get("prompt_file") or "").strip()
        if not prompt_file:
            continue
        parsed_name = parse_prompt_filename(prompt_file)
        if not parsed_name:
            continue
        fmt, lang, parsed_persona = parsed_name
        persona_number = entry.get("persona_number")
        if not isinstance(persona_number, int):
            persona_number = parsed_persona
        line_items = entry.get("copy_lines")
        if not isinstance(line_items, list) or not line_items:
            continue

        target_ad = None
        for ad in ads:
            if not isinstance(ad, dict):
                continue
            if str(ad.get("format") or "").strip().upper() != fmt:
                continue
            if isinstance(persona_number, int):
                persona = ad.get("persona")
                ad_persona_no = None
                if isinstance(persona, dict) and isinstance(persona.get("number"), int):
                    ad_persona_no = int(persona.get("number"))
                if ad_persona_no != persona_number:
                    continue
            target_ad = ad
            break
        if not isinstance(target_ad, dict):
            continue

        ad_copy = target_ad.setdefault("copy", {})
        if not isinstance(ad_copy, dict):
            continue
        lang_copy = ad_copy.setdefault(lang, {})
        if not isinstance(lang_copy, dict):
            continue

        for line_item in line_items:
            if not isinstance(line_item, dict):
                continue
            label = str(line_item.get("label") or "").strip()
            value = str(line_item.get("value") or "").strip()
            if not label:
                continue
            key = label.lower()

            if key == "headline":
                lang_copy["headline"] = value
            elif key == "subheadline":
                lang_copy["subheadline"] = value
            elif key == "support line":
                lang_copy["support_line"] = value
            elif key == "context line":
                lang_copy["context_line"] = value
            elif key == "cta":
                lang_copy["cta"] = value
            elif key == "attribution":
                lang_copy["attribution"] = value
            elif key == "trust line":
                lang_copy["trust_line"] = value
            elif key.startswith("bullet "):
                match = re.match(r"^bullet\s+(\d+)$", key)
                if not match:
                    continue
                index = int(match.group(1)) - 1
                if index < 0:
                    continue
                bullets = lang_copy.get("bullets")
                if not isinstance(bullets, list):
                    bullets = []
                while len(bullets) <= index:
                    bullets.append("")
                bullets[index] = value
                lang_copy["bullets"] = bullets
            elif key.startswith("left situation ") or key.startswith("right shift "):
                match = re.match(r"^(left situation|right shift)\s+(\d+)$", key)
                if not match:
                    continue
                side = match.group(1)
                ordinal = int(match.group(2))
                if ordinal <= 0:
                    continue
                if side == "left situation":
                    index = ordinal - 1
                else:
                    index = ordinal + 1
                bullets = lang_copy.get("bullets")
                if not isinstance(bullets, list):
                    bullets = []
                while len(bullets) <= index:
                    bullets.append("")
                bullets[index] = value
                lang_copy["bullets"] = bullets

        updated_count += 1

    if updated_count == 0:
        raise HTTPException(status_code=400, detail="No valid prompt edits were provided")

    copy_path.write_text(json.dumps(copy_json, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    rerender_prompts_for_run(run_dir, batch, copy_path, load_run_language_mode(run_dir))

    _append_audit_log(
        run_dir,
        "prompt_updates",
        {
            "run_id": run_id,
            "batch": batch,
            "updated_count": updated_count,
        },
    )

    has_916 = any("/96/" in str(path) for path in (manifest.get("prompt_files") or []))
    if has_916:
        manifest = generate_916_for_run(run_dir, manifest)

    refreshed = collect_run_result(run_dir, batch, bool(manifest.get("image_generated", False)))
    refreshed["copy_edits_applied"] = updated_count
    merged = merge_manifest(run_dir, manifest, refreshed)
    return merged

EXACT_COPY_SHEET_COLUMNS = [
    "prompt_id",
    "vn",
    "format",
    "persona_name",
    "persona_pain",
    "persona_desire",
    "persona_friction",
    "persona_proof",
    "persona_tone",
    "concept_angle",
    "concept_angle_definition",
    "hypothesis_type",
    "hypothesis_variant",
    "headline_copy",
    "exact_on_image_copy_block",
    "created_at",
]

def _extract_vn_from_prompt_rel_path(prompt_rel_path: str) -> str:
    # Expected pattern: output/v{N}/...
    # Keep backward compatible: if not found, return empty string.
    m = re.search(r"/output/(v\d+)(/|$)", prompt_rel_path.replace("\\", "/"))
    return m.group(1) if m else ""

def _parse_exact_block_headline_value(block_text: str) -> str | None:
    """
    Preserve EXACT headline value text as written in the exact block.

    We intentionally do NOT trim/normalize:
    - keep any spaces immediately after the colon
    - keep punctuation/case/capitalization
    """
    for raw_line in (block_text or "").splitlines():
        # Allow optional whitespace before "-" and around "-", but preserve everything after "Headline:"
        m = re.match(r"^\s*-\s*Headline:(.*)$", raw_line)
        if not m:
            continue
        return m.group(1)
    return None

def _replace_exact_copy_block(prompt_text: str, new_block_text: str) -> str | None:
    from dashboard.backend.pipeline.images import EXACT_COPY_BLOCK_RE
    m = EXACT_COPY_BLOCK_RE.search(prompt_text or "")
    if not m:
        return None
    start_idx = m.start("block")
    end_idx = m.end("block")
    return (prompt_text[:start_idx] + new_block_text + prompt_text[end_idx:])

def _load_run_prompt_files(run_id: str, aspect_ratios: list[str] | None = None) -> list[str]:
    from dashboard.backend.pipeline.runs_db import load_manifest_for_run
    _run_dir, manifest, _has_storage_manifest = load_manifest_for_run(run_id)
    prompt_files_all = manifest.get("prompt_files") or []
    if not aspect_ratios:
        return prompt_files_all
    result: list[str] = []
    for p in prompt_files_all:
        for ar in aspect_ratios:
            if f"/{ar}/" in str(p):
                result.append(p)
                break
    return result or prompt_files_all

def _get_architecture_definition(arch: dict[str, Any], group: str, variant: str) -> str:
    """Get the intent summary for a concept_variation field from copy_architecture.json."""
    if not arch or not group or not variant:
        return ""
    headline_archs = arch.get("headline_architectures") or {}
    group_data = headline_archs.get(group) or {}
    variant_data = group_data.get(variant) or {}
    return str(variant_data.get("meaning") or variant_data.get("intent") or variant_data.get("direction") or variant_data.get("template") or "").strip()

def _extract_prompt_row_metadata(run_id: str, copy_batch: dict[str, Any], prompt_rel_path: str, batch_vn: str = "") -> dict[str, Any]:
    from dashboard.backend.pipeline.images import _extract_created_at_iso_from_file, extract_exact_on_image_copy_block, parse_prompt_creative_index, parse_prompt_filename
    from dashboard.backend.pipeline.personas import parse_persona_number_from_prompt
    prompt_path = ROOT / prompt_rel_path
    text = prompt_path.read_text(encoding="utf-8", errors="ignore")

    block = extract_exact_on_image_copy_block(text)
    headline_copy = None
    exact_block = ""
    if block is not None:
        headline_copy = _parse_exact_block_headline_value(block)
        exact_block = block.strip()
    if headline_copy is None:
        headline_copy = ""

    vn = _extract_vn_from_prompt_rel_path(prompt_rel_path)
    if not vn and batch_vn:
        vn = batch_vn
    created_at = _extract_created_at_iso_from_file(prompt_path)

    parsed = parse_prompt_filename(prompt_rel_path)
    fmt = parsed[0] if parsed else ""
    persona_number = parsed[2] if parsed else None
    creative_index = parse_prompt_creative_index(prompt_rel_path)
    if persona_number is None:
        persona_number = parse_persona_number_from_prompt(text)

    persona_name = ""
    persona_pain = ""
    persona_desire = ""
    persona_friction = ""
    persona_proof = ""
    persona_tone = ""
    concept_angle = ""
    hypothesis_type = ""
    hypothesis_variant = ""

    if isinstance(persona_number, int):
        # Pull full persona data from persona_seeds.json
        seed = _resolve_persona_seeds().get(persona_number) or {}
        persona_name = str(seed.get("persona_name") or f"Persona {persona_number}")
        persona_pain = str(seed.get("core_pattern", ""))
        persona_desire = str(seed.get("common_indian_moments", ""))
        persona_friction = str(seed.get("objections_raw", ""))
        persona_proof = str(seed.get("common_indian_moments", ""))
        persona_tone = str(seed.get("guardrail", ""))

        for ad in copy_batch.get("ads") or []:
            if not isinstance(ad, dict):
                continue
            if str(ad.get("format") or "").strip().upper() != fmt:
                continue
            persona_obj = ad.get("persona")
            if isinstance(persona_obj, dict):
                if isinstance(persona_obj.get("number"), int) and int(persona_obj.get("number")) == persona_number:
                    if int(ad.get("creative_index") or 1) != creative_index:
                        continue
                    # Override persona_name from copy_batch if available
                    pn = str(persona_obj.get("persona_name") or persona_obj.get("name") or "")
                    if pn:
                        persona_name = pn
                    concept_angle = str(ad.get("concept_angle") or ad.get("headline_angle") or "")
                    hyp = ad.get("hypothesis") if isinstance(ad.get("hypothesis"), dict) else {}
                    if hyp:
                        hypothesis_type = str(hyp.get("type") or hyp.get("variable_label") or "")
                        hypothesis_variant = str(hyp.get("variant") or "")
                    break

    arch = _resolve_copy_architecture()
    concept_angle_def = _get_architecture_definition(arch, "concept_angle", concept_angle)

    return {
        "prompt_id": prompt_rel_path,
        "vn": vn,
        "format": fmt,
        "persona_name": persona_name,
        "persona_pain": persona_pain,
        "persona_desire": persona_desire,
        "persona_friction": persona_friction,
        "persona_proof": persona_proof,
        "persona_tone": persona_tone,
        "concept_angle": concept_angle,
        "concept_angle_definition": concept_angle_def,
        "hypothesis_type": hypothesis_type,
        "hypothesis_variant": hypothesis_variant,
        "headline_copy": headline_copy,
        "exact_on_image_copy_block": exact_block,
        "created_at": created_at,
    }

def _append_audit_log(run_dir: Path, event_type: str, payload: dict[str, Any]) -> None:
    pass

def api_export_on_image_copy(run_id: str) -> StreamingResponse:
    from dashboard.backend.pipeline.runs_db import load_manifest_for_run
    from openpyxl import Workbook
    import io

    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)

    copy_batch: dict[str, Any] = {"ads": []}
    if has_storage_manifest and run_dir is not None:
        copy_path = run_dir / "context" / "copy_batch.json"
        if copy_path.exists():
            copy_batch = json.loads(copy_path.read_text(encoding="utf-8"))

    prompt_files = _load_run_prompt_files(run_id)

    unique_vns = set()
    for rel in prompt_files:
        prompt_rel_path = str(rel).replace("\\", "/")
        vn = _extract_vn_from_prompt_rel_path(prompt_rel_path)
        if vn:
            unique_vns.add(vn)

    batch = manifest.get("batch", "")

    if not unique_vns:
        if batch and batch.startswith("v"):
            unique_vns.add(batch)

    if unique_vns:
        vn_suffix = "-".join(sorted(unique_vns))
    else:
        vn_suffix = None

    wb = Workbook()
    ws = wb.active
    ws.title = "on-image-copy"

    ws.append(EXACT_COPY_SHEET_COLUMNS)
    for rel in prompt_files:
        prompt_rel_path = str(rel).replace("\\", "/")
        if not (ROOT / prompt_rel_path).exists():
            continue
        row = _extract_prompt_row_metadata(run_id, copy_batch, prompt_rel_path, batch)
        ws.append([row.get(col, "") for col in EXACT_COPY_SHEET_COLUMNS])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if run_dir is not None:
        _append_audit_log(run_dir, "export_on_image_copy", {"run_id": run_id, "prompt_rows": len(prompt_files)})

    if vn_suffix:
        filename = f"on-image-copy-{vn_suffix}.xlsx"
    else:
        filename = f"on-image-copy-{run_id}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

async def api_import_on_image_copy(
    run_id: str,
    file: UploadFile = File(...),
    confirm: bool = Form(False),
) -> dict[str, Any]:
    from dashboard.backend.pipeline.images import collect_run_result, extract_exact_on_image_copy_block
    from dashboard.backend.pipeline.runs_db import collect_backfill_result, load_manifest_for_run, merge_manifest
    from openpyxl import load_workbook

    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)

    # Parse xlsx (no prompt regeneration; only exact-block replacement)
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing upload filename")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty upload")

    import_root = (run_dir / "imports") if run_dir is not None else (RUNTIME_ROOT / "imports")
    tmp_path = import_root / f"upload-{int(time.time())}-{file.filename}"
    tmp_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path.write_bytes(content)

    wb = load_workbook(tmp_path)
    ws = wb.active

    # Build column index
    header = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(header) if name}

    # Required columns for import (core copy fields)
    REQUIRED_IMPORT_COLUMNS = ["prompt_id", "headline_copy", "exact_on_image_copy_block"]
    missing_required = [c for c in REQUIRED_IMPORT_COLUMNS if c not in col_idx]
    if missing_required:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing_required}")

    # Optional context columns (from export, used for validation/reference)
    OPTIONAL_CONTEXT_COLUMNS = [
        "format", "persona_name", "persona_pain", "persona_desire", "persona_friction",
        "persona_proof", "persona_tone",
        "concept_angle", "concept_angle_definition",
        "hypothesis_type", "hypothesis_variant",
        "vn", "created_at",
    ]

    seen_prompt_ids: set[str] = set()
    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for excel_row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in excel_row]
        prompt_id = str(values[col_idx["prompt_id"]] or "").replace("\\", "/").strip()
        if not prompt_id:
            continue
        if prompt_id in seen_prompt_ids:
            errors.append({"prompt_id": prompt_id, "error": "duplicate_prompt_id"})
            continue
        seen_prompt_ids.add(prompt_id)

        headline_copy = str(values[col_idx["headline_copy"]] or "").strip()
        full_block = str(values[col_idx.get("exact_on_image_copy_block", -1)] or "").strip() if "exact_on_image_copy_block" in col_idx else ""

        # Read optional context columns if present
        ctx: dict[str, str] = {}
        for opt_col in OPTIONAL_CONTEXT_COLUMNS:
            if opt_col in col_idx:
                ctx[opt_col] = str(values[col_idx[opt_col]] or "").strip()

        if not headline_copy.strip() and not full_block:
            errors.append({"prompt_id": prompt_id, "error": "empty_headline_copy_and_block"})
            continue

        rows.append(
            {
                "prompt_id": prompt_id,
                "headline_copy": headline_copy,
                "full_block": full_block,
                "context": ctx,
            }
        )

    # Validate prompt_id exists
    for r in rows:
        p = ROOT / r["prompt_id"]
        if not p.exists() or not p.is_file():
            errors.append({"prompt_id": r["prompt_id"], "error": "prompt_id_not_found"})
    if errors:
        if run_dir is not None:
            _append_audit_log(run_dir, "import_on_image_copy_validation_failed", {"run_id": run_id, "errors": errors, "confirm": confirm})
        raise HTTPException(status_code=400, detail={"validation_errors": errors})

    # Preview diffs
    preview_items: list[dict[str, Any]] = []
    applied_count = 0
    skipped_count = 0

    for r in rows:
        prompt_rel_path = r["prompt_id"]
        prompt_path = ROOT / prompt_rel_path
        old_text = prompt_path.read_text(encoding="utf-8", errors="ignore")

        old_block = extract_exact_on_image_copy_block(old_text, warn_log_path=None)
        if old_block is None:
            skipped_count += 1
            preview_items.append({"prompt_id": prompt_rel_path, "status": "skipped_missing_exact_block"})
            continue

        full_block = r.get("full_block", "")
        new_block = None

        if full_block:
            new_block = full_block
            old_copy = old_block.strip()
            new_copy = new_block.strip()
        else:
            headline_copy = r.get("headline_copy", "")
            new_lines: list[str] = []
            headline_replaced = False
            for line in old_block.splitlines():
                m = re.match(r"^(\s*-\s*Headline:)(.*)$", line)
                if m:
                    new_lines.append(m.group(1) + headline_copy)
                    headline_replaced = True
                else:
                    new_lines.append(line)

            if not headline_replaced:
                skipped_count += 1
                preview_items.append({"prompt_id": prompt_rel_path, "status": "skipped_headline_line_not_found"})
                continue

            new_block = "\n".join(new_lines)
            old_copy = _parse_exact_block_headline_value(old_block) or ""
            new_copy = _parse_exact_block_headline_value(new_block) or ""

        preview_items.append(
            {
                "prompt_id": prompt_rel_path,
                "status": "ready_to_apply" if confirm else "preview",
                "old_copy": old_copy[:100] + "..." if len(old_copy) > 100 else old_copy,
                "new_copy": new_copy[:100] + "..." if len(new_copy) > 100 else new_copy,
            }
        )

        if confirm:
            updated_text = _replace_exact_copy_block(old_text, new_block)
            if updated_text is None:
                skipped_count += 1
                preview_items[-1]["status"] = "skipped_replace_failed"
                continue
            prompt_path.write_text(updated_text, encoding="utf-8")
            applied_count += 1

    if run_dir is not None:
        _append_audit_log(
            run_dir,
            "import_on_image_copy",
            {"run_id": run_id, "confirm": confirm, "rows": len(rows), "applied": applied_count, "skipped": skipped_count},
        )

    if not confirm:
        return {
            "run_id": run_id,
            "preview": True,
            "changed_rows_count": applied_count,
            "skipped_rows": skipped_count,
            "failed_rows": len(errors),
            "items": preview_items,
        }

    # Re-assemble prompts side-effects: since we directly edited prompt text,
    # we do not mutate copy_batch.json metadata (per requirements).
    # However, manifest/prompt_files state should be refreshed.
    merged: dict[str, Any] | None = None
    if run_dir is not None and has_storage_manifest:
        manifest = json.loads((run_dir / "manifest.json").read_text(encoding="utf-8"))
        batch = (manifest.get("batch") or "").strip()
        refreshed = collect_run_result(run_dir, batch, bool(manifest.get("image_generated", False)))
        refreshed["on_image_copy_import_applied"] = applied_count
        merged = merge_manifest(run_dir, manifest, refreshed)
    else:
        batch = str(manifest.get("batch") or "")
        merged = collect_backfill_result(run_id, batch) if batch else manifest
        merged["on_image_copy_import_applied"] = applied_count

    return {
        "run_id": run_id,
        "preview": False,
        "changed_rows_count": applied_count,
        "skipped_rows": skipped_count,
        "failed_rows": len(errors),
        "items": preview_items,
        "manifest": merged,
    }

def extract_persona_input_block(prompt_text: str) -> str:
    markers = ["EXACT ON-IMAGE COPY", "PERSONA INPUT", "PERSONA:", "INPUT:"]
    for marker in markers:
        if marker in prompt_text.upper():
            start = prompt_text.upper().find(marker)
            if start != -1:
                return prompt_text[start:].strip()
    if len(prompt_text) > 50:
        return prompt_text.strip()
    return ""

def api_file_content(path: str, max_lines: int = 400) -> dict[str, Any]:
    from dashboard.backend.pipeline.files import resolve_safe_path
    file_path = resolve_safe_path(path)
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    lines = file_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    clipped = lines[:max_lines]
    return {
        "path": str(file_path.relative_to(ROOT)),
        "total_lines": len(lines),
        "shown_lines": len(clipped),
        "content": "\n".join(clipped),
    }

def api_edit_prompt(run_id: str, payload: dict[str, Any] = Body(...), user_id: str = "") -> dict[str, Any]:
    """Edit a prompt file in-place, replacing only the EXACT ON-IMAGE COPY block."""
    from dashboard.backend.pipeline.runs_db import _check_ownership
    prompt_path = payload.get("prompt_file", "")
    new_text = payload.get("text", "")
    if not prompt_path or not new_text:
        raise HTTPException(status_code=400, detail="prompt_file and text are required")

    if user_id:
        _check_ownership(run_id, user_id)
        try:
            from dashboard.backend.db.client import get_sync_db
            from dashboard.backend.db.collections import COLL_PROMPTS
            db = get_sync_db()
            doc = db[COLL_PROMPTS].find_one({"user_id": user_id, "run_id": run_id, "file_path": prompt_path})
            if not doc:
                raise HTTPException(status_code=404, detail="Prompt not found")
            old_text = str(doc.get("content") or "")
            updated_text = _replace_exact_copy_block(old_text, new_text)
            if updated_text is None:
                raise HTTPException(status_code=400, detail="No EXACT ON-IMAGE COPY block found in prompt file")
            db[COLL_PROMPTS].update_one(
                {"_id": doc["_id"]},
                {"$set": {"content": updated_text, "updated_at": time.time()}},
            )
            full_path = ROOT / prompt_path
            if full_path.exists() and full_path.is_file():
                full_path.write_text(updated_text, encoding="utf-8")
            return {"status": "saved", "prompt_file": prompt_path}
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Could not edit prompt: {exc}") from exc

    full_path = ROOT / prompt_path
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="Prompt file not found")

    old_text = full_path.read_text(encoding="utf-8")
    updated_text = _replace_exact_copy_block(old_text, new_text)
    if updated_text is None:
        raise HTTPException(status_code=400, detail="No EXACT ON-IMAGE COPY block found in prompt file")
    full_path.write_text(updated_text, encoding="utf-8")
    return {"status": "saved", "prompt_file": prompt_path}

def _build_persona_name_map(run_dir: Path) -> dict[str, str]:
    """Map persona number (P01) to persona name from run's copy_batch.json."""
    copy_path = run_dir / "context" / "copy_batch.json"
    if not copy_path.exists():
        return {}
    try:
        data = json.loads(copy_path.read_text(encoding="utf-8"))
        ads = data.get("ads") if isinstance(data, dict) else []
        if not isinstance(ads, list):
            return {}
        mapping: dict[str, str] = {}
        for ad in ads:
            p = ad.get("persona") if isinstance(ad, dict) else {}
            if not isinstance(p, dict):
                continue
            num = p.get("number")
            name = p.get("name") or p.get("persona_name") or ""
            if isinstance(num, int) and name:
                mapping[f"P{num:02d}"] = str(name)
        return mapping
    except Exception:
        return {}
