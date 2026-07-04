#!/usr/bin/env python3

from __future__ import annotations

import copy
import json
import os
import random
import re
import shutil
import subprocess
import threading
import sys
import time
import urllib.request
import hashlib
import importlib.util
import mimetypes
import traceback
import uuid
import asyncio
import httpx
import psutil

if sys.platform == "win32":
    import msvcrt
else:
    import fcntl
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

from fastapi import Body, FastAPI, File, Form, HTTPException, UploadFile, Request, Response
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles


ROOT = Path(__file__).resolve().parents[2]
STORAGE_ROOT = ROOT / "dashboard_storage"
RUNS_ROOT = STORAGE_ROOT / "runs"
RUNTIME_ROOT = ROOT / "runtime"
ENV_PATH = ROOT / ".env.dashboard"

DEFAULT_PRODUCT_MASTER = ROOT / "input" / "docs" / "product master doc.txt"
DEFAULT_IMAGE_SOURCES_FILE = ROOT / "input" / "image_sources.txt"
LEGACY_ACTIVE_IMAGES_FILE = ROOT / "input" / "activeimages.txt"
INPUT_IMAGES_DIR = ROOT / "input" / "images"
GENERATED_IMAGES_ROOT = ROOT / "generated_images"
CONVERT_916_TEMPLATE_PATH = ROOT / "input" / "prompt_916_from_45.txt"
PERSONA_SEEDS_PATH = ROOT / "persona_seeds.json"
COPY_ARCH_PATH = ROOT / "dashboard" / "backend" / "copy_architecture.json"
COPY_PROMPTS_PATH = ROOT / "dashboard" / "backend" / "copy_prompt_templates.json"
STARTING_PROMPT_PATH = ROOT / "input" / "startingprompt.txt"

FORMATS = ["HERO", "BA", "TEST", "FEAT", "UGC"]
DEFAULT_OPENCODE_API_URL = os.getenv("OPENCODE_API_URL", "http://127.0.0.1:4090")
DEFAULT_GOOGLE_API_URL = "https://generativelanguage.googleapis.com/v1beta"
DEFAULT_GOOGLE_MODEL = "gemini-2.0-flash"
OPENCODE_ADS_PER_SESSION_SCHEDULE = [25, 15, 10, 5, 2, 1]
OPENCODE_AD_TIMEOUT_SECONDS = 600
OPENCODE_MAX_CONCURRENT = 2
OPENCODE_QUEUE_DIR = RUNTIME_ROOT / "opencode_queue"
LLM_TRACES_DIR = RUNTIME_ROOT / "llm_traces"

_PERSONA_SEED_MAPPING: dict[str, Any] = {
    "seed_to_payload": {
        "core_pattern": {"field": "pain_points", "wrap_list": True, "prefix": None},
        "common_indian_moments": {"field": "trust_anchors", "wrap_list": True, "prefix": None},
        "objections_raw": {"field": "objections", "wrap_list": True, "prefix": None},
        "how_kit_solves": {"field": "how_kit_solves", "wrap_list": False, "prefix": None},
        "guardrail": {"field": "guardrails", "wrap_list": True, "prefix": "Guardrail: "},
    },
    "persona_fallbacks": {
        "core_pattern": "Daily routine feels heavy and hard to sustain.",
        "common_indian_moments": "Everyday situations make it harder to stay consistent.",
        "objections_raw": "Past plans felt too strict and difficult to maintain.",
        "how_kit_solves": {},
        "guardrail": "",
    },
    "static_fields": {
        "trigger_scenarios": [],
        "language_bank": [],
        "grounded_mechanism_map": [],
    },
    "hindi_ready_default": "टोन संकेत: सरल, भरोसेमंड, व्यावहारिक",
}

_TESTIMONIAL_GUIDANCE: dict[str, Any] = {
    "EN": {
        "first_person_pattern": "\\b(i|i'm|i've|i'd|my|me)\\b",
        "weight_pattern": "\\b(weight|obesity|excess\\s*weight|kg|kilo)\\b",
        "suffix": "It finally fit my weight-loss routine.",
        "desire_template": "\"I finally found {desire_phrase} for my weight-loss goal.\"",
        "fallback": "\"I finally found a routine I can follow for weight loss every day.\"",
        "desire_field": "desire_en",
    },
    "HI": {
        "first_person_pattern": "(मैं|मेरी|मेरा|मुझे|मैंने)",
        "weight_pattern": "(वजन|मोटापा|किलो|kg)",
        "suffix": "यह मेरे वजन घटाने के लिए काम आया।",
        "desire_template": "\"मुझे आखिर {desire_phrase} वाला रूटीन मिला जो वजन घटाने में मदद करता है।\"",
        "fallback": "\"मुझे आखिर ऐसा रूटीन मिला जिसे मैं रोज निभा सकूं और वजन घटा सकूं।\"",
        "desire_field": "desire_hi",
    },
}

# Pipeline cancellation signal, keyed by run_id.
_cancel_events: dict[str, threading.Event] = {}
_cancel_current_run: threading.Event = threading.Event()
# Registry of subprocesses spawned by run_cmd() so they can be killed on cancel.
# Each entry: (run_id_or_None, subprocess.Popen)
_tracked_subprocesses: list[tuple[str | None, "subprocess.Popen[str]"]] = []
_tracked_subprocesses_lock = threading.Lock()
# Registry of active httpx clients so cancel can interrupt in-flight HTTP calls.
_active_httpx_clients: dict[int, httpx.Client] = {}
_active_httpx_clients_lock = threading.Lock()

# Run ownership registry: maps run_id -> user_id
# Populated when runs are created; used for ownership checks and LLM traces
_run_owner_registry: dict[str, str] = {}
_run_owner_registry_lock = threading.Lock()


def _record_run_owner(run_id: str, user_id: str, config: dict[str, Any] | None = None) -> None:
    """Record that a user owns a run, both in-memory and in MongoDB."""
    with _run_owner_registry_lock:
        _run_owner_registry[run_id] = user_id
    try:
        from dashboard.backend.services.run_storage import create_run, get_run, update_run
        doc = get_run(user_id, run_id)
        if doc:
            update_run(user_id, run_id, {"status": "created"})
        else:
            create_run(user_id, run_id, {"status": "created", "config": config or {}})
    except Exception:
        pass
    try:
        owner_path = RUNS_ROOT / run_id / ".owner"
        owner_path.parent.mkdir(parents=True, exist_ok=True)
        owner_path.write_text(user_id, encoding="utf-8")
    except Exception:
        pass


def _update_run_status_db(run_id: str, status: str, user_id: str | None = None, extra: dict[str, Any] | None = None) -> None:
    """Update a run's status in MongoDB (best-effort, no-op on failure)."""
    if not user_id:
        user_id = _get_run_owner(run_id) or ""
    if not user_id:
        return
    try:
        from dashboard.backend.services.run_storage import update_run
        updates: dict[str, Any] = {"status": status}
        if extra:
            updates.update(extra)
        update_run(user_id, run_id, updates)
    except Exception:
        pass


def _get_run_owner(run_id: str) -> str | None:
    """Resolve the owner user_id for a run.

    Checks: in-memory registry -> MongoDB -> .owner file.
    Returns None if unknown.
    """
    with _run_owner_registry_lock:
        uid = _run_owner_registry.get(run_id)
        if uid is not None:
            return uid
    try:
        from dashboard.backend.db.client import get_sync_db
        from dashboard.backend.db.collections import COLL_RUNS
        doc = get_sync_db()[COLL_RUNS].find_one({"run_id": run_id}, {"user_id": 1})
        if doc and doc.get("user_id"):
            uid = doc["user_id"]
            with _run_owner_registry_lock:
                _run_owner_registry[run_id] = uid
            return uid
    except Exception:
        pass
    owner_file = RUNS_ROOT / run_id / ".owner"
    if owner_file.exists():
        uid = owner_file.read_text(encoding="utf-8").strip()
        if uid:
            with _run_owner_registry_lock:
                _run_owner_registry[run_id] = uid
            return uid
    return None


def _parse_prompt_meta(file_path: str) -> dict[str, str]:
    """Parse format, language, persona_slug, concept_angle from a prompt filename.

    Expected pattern: {FMT}_{slug}_{LANG}_{angle}[_A{NN}].txt
    Returns dict with format, language, persona_slug, concept_angle keys (may be empty).
    """
    import re
    stem = Path(file_path).stem
    m = re.match(
        r"^(HERO|BA|TEST|FEAT|UGC)_(.+?)_(EN|HI|HINGLISH)_(.+?)(?:_A\d{2,})?$",
        stem,
        re.IGNORECASE,
    )
    if m:
        return {
            "format": m.group(1).upper(),
            "persona_slug": m.group(2),
            "language": m.group(3).upper(),
            "concept_angle": m.group(4),
        }
    return {"format": "", "persona_slug": "", "language": "", "concept_angle": ""}


def _store_output_mapping(run_id: str, user_id: str, batch: str, manifest: dict[str, Any]) -> None:
    """Scan output and generated_images dirs for a batch and store file->run mapping in MongoDB.

    Called after a pipeline completes so that download endpoints can resolve
    file paths to run owners via MongoDB instead of parsing filesystem paths.
    Also upserts prompt and image documents into COLL_PROMPTS and COLL_IMAGES
    so the DB-backed download endpoints (/image/{image_id}, /prompt/{prompt_id})
    have usable records.
    Best-effort: silently ignores if MongoDB is unavailable.
    """
    try:
        import time
        import hashlib
        from dashboard.backend.db.client import get_sync_db
        from dashboard.backend.db.collections import COLL_FILE_MAP, COLL_PROMPTS, COLL_IMAGES

        db = get_sync_db()
        now = time.time()

        # Remove any stale mapping for this run
        db[COLL_FILE_MAP].delete_many({"run_id": run_id})

        file_entries: list[dict[str, Any]] = []

        # Map the batch's output directory
        output_batch_dir = ROOT / "output" / batch
        if output_batch_dir.is_dir():
            file_entries.append({
                "file_path": f"output/{batch}/",
                "file_type": "output_dir",
                "run_id": run_id,
                "user_id": user_id,
                "batch": batch,
                "created_at": now,
            })
            for f in sorted(output_batch_dir.rglob("*")):
                if f.is_file():
                    rel = f.relative_to(ROOT).as_posix()
                    ext = f.suffix.lower()
                    file_type = "output_json" if ext == ".json" else "prompt"
                    file_entries.append({
                        "file_path": rel,
                        "file_type": file_type,
                        "run_id": run_id,
                        "user_id": user_id,
                        "batch": batch,
                        "created_at": now,
                    })
                    if ext == ".txt":
                        prompt_id = hashlib.sha256(rel.encode()).hexdigest()[:16]
                        meta = _parse_prompt_meta(rel)
                        content = f.read_text(encoding="utf-8", errors="ignore")
                        db[COLL_PROMPTS].update_one(
                            {"prompt_id": prompt_id},
                            {"$set": {
                                "user_id": user_id,
                                "run_id": run_id,
                                "prompt_id": prompt_id,
                                "batch": batch,
                                "file_path": rel,
                                "format": meta.get("format", ""),
                                "persona_slug": meta.get("persona_slug", ""),
                                "language": meta.get("language", ""),
                                "concept_angle": meta.get("concept_angle", ""),
                                "filename": f.name,
                                "content": content,
                                "status": "completed",
                                "storage_provider": "local",
                                "created_at": now,
                                "updated_at": now,
                            }},
                            upsert=True,
                        )

        # Map the batch's generated_images directory
        img_batch_dir = GENERATED_IMAGES_ROOT / batch
        if img_batch_dir.is_dir():
            file_entries.append({
                "file_path": f"generated_images/{batch}/",
                "file_type": "image_dir",
                "run_id": run_id,
                "user_id": user_id,
                "batch": batch,
                "created_at": now,
            })
            for f in sorted(img_batch_dir.rglob("*")):
                if f.is_file():
                    rel = f.relative_to(ROOT).as_posix()
                    ext = f.suffix.lower()
                    file_type = "image_metadata" if ext == ".json" else "image"
                    file_entries.append({
                        "file_path": rel,
                        "file_type": file_type,
                        "run_id": run_id,
                        "user_id": user_id,
                        "batch": batch,
                        "created_at": now,
                    })
                    if ext in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
                        image_id = hashlib.sha256(rel.encode()).hexdigest()[:16]
                        meta: dict[str, Any] = {}
                        sidecar = f.with_suffix(f"{f.suffix}.json")
                        if not sidecar.exists():
                            sidecar = f.with_suffix(".json")
                        if sidecar.exists():
                            try:
                                meta = json.loads(sidecar.read_text(encoding="utf-8", errors="ignore"))
                            except Exception:
                                pass
                        # Upload via active storage backend (local or cloudinary)
                        try:
                            from dashboard.backend.services.storage.service import image_metadata_for_db
                            img_doc = image_metadata_for_db(
                                f, run_id=run_id, user_id=user_id, batch=batch,
                                file_path=rel,
                                width=meta.get("width", 0), height=meta.get("height", 0),
                            )
                        except Exception:
                            img_doc = {
                                "user_id": user_id,
                                "run_id": run_id,
                                "image_id": image_id,
                                "batch": batch,
                                "file_path": rel,
                                "local_path": rel,
                                "filename": f.name,
                                "format": meta.get("format", ""),
                                "status": "completed",
                                "storage_provider": "local",
                                "metadata": meta,
                                "created_at": now,
                                "updated_at": now,
                            }
                        db[COLL_IMAGES].update_one(
                            {"image_id": image_id},
                            {"$set": img_doc},
                            upsert=True,
                        )

        # Also map all known prompt files from manifest
        for pf in manifest.get("prompt_files") or []:
            pf_str = str(pf).replace("\\", "/")
            if not any(e["file_path"] == pf_str for e in file_entries):
                file_entries.append({
                    "file_path": pf_str,
                    "file_type": "prompt",
                    "run_id": run_id,
                    "user_id": user_id,
                    "batch": batch,
                    "created_at": now,
                })

        # Also map all known image files from manifest
        for imgf in manifest.get("image_files") or []:
            imgf_str = str(imgf).replace("\\", "/")
            if not any(e["file_path"] == imgf_str for e in file_entries):
                file_entries.append({
                    "file_path": imgf_str,
                    "file_type": "image",
                    "run_id": run_id,
                    "user_id": user_id,
                    "batch": batch,
                    "created_at": now,
                })

        if file_entries:
            db[COLL_FILE_MAP].insert_many(file_entries)
    except Exception:
        pass


def _resolve_file_owner(file_path: str) -> dict[str, str] | None:
    """Look up the owner of a file path from the MongoDB file map.

    Returns dict with "run_id" and "user_id" if found, else None.
    """
    try:
        from dashboard.backend.db.client import get_sync_db
        from dashboard.backend.db.collections import COLL_FILE_MAP
        doc = get_sync_db()[COLL_FILE_MAP].find_one(
            {"file_path": file_path},
            {"run_id": 1, "user_id": 1},
        )
        if doc and doc.get("run_id") and doc.get("user_id"):
            return {"run_id": doc["run_id"], "user_id": doc["user_id"]}
    except Exception:
        pass
    return None


def _extract_run_id_from_output_path(path: str) -> str | None:
    """Resolve output/{batch}/... path to run_id via MongoDB file map."""
    clean = path.replace("\\", "/").lstrip("/")
    owner = _resolve_file_owner(clean)
    if owner:
        return owner["run_id"]
    # Fallback: extract batch from path, look up in runs table
    parts = clean.split("/")
    for i, p in enumerate(parts):
        if p.startswith("v") and p[1:].isdigit():
            batch = p
            try:
                from dashboard.backend.db.client import get_sync_db
                from dashboard.backend.db.collections import COLL_RUNS
                doc = get_sync_db()[COLL_RUNS].find_one(
                    {"batch": batch},
                    {"run_id": 1, "user_id": 1},
                    sort=[("created_at", -1)],
                )
                if doc and doc.get("run_id"):
                    return doc["run_id"]
            except Exception:
                pass
            break
    return None


def _extract_run_id_from_generated_path(path: str) -> str | None:
    """Resolve generated_images/{batch}/... path to run_id via MongoDB file map."""
    clean = path.replace("\\", "/").lstrip("/")
    owner = _resolve_file_owner(clean)
    if owner:
        return owner["run_id"]
    # Legacy fallback: check if path contains run_X / batch_v prefix
    parts = clean.split("/")
    for part in parts:
        if part.startswith("run_"):
            return part
    # Fallback to batch lookup
    for i, p in enumerate(parts):
        if p.startswith("v") and p[1:].isdigit():
            batch = p
            try:
                from dashboard.backend.db.client import get_sync_db
                from dashboard.backend.db.collections import COLL_RUNS
                doc = get_sync_db()[COLL_RUNS].find_one(
                    {"batch": batch},
                    {"run_id": 1},
                    sort=[("created_at", -1)],
                )
                if doc and doc.get("run_id"):
                    return doc["run_id"]
            except Exception:
                pass
            break
    return None


def _register_subprocess(run_id: str | None, proc: "subprocess.Popen[str]") -> None:
    with _tracked_subprocesses_lock:
        _tracked_subprocesses.append((run_id, proc))


def _unregister_subprocess(proc: "subprocess.Popen[str]") -> None:
    with _tracked_subprocesses_lock:
        _tracked_subprocesses[:] = [(rid, p) for (rid, p) in _tracked_subprocesses if p is not proc]


def _kill_tracked_subprocesses(run_id: str | None) -> list[str]:
    """Kill all tracked subprocesses (optionally filtered by run_id). Returns list of killed commands."""
    killed: list[str] = []
    with _tracked_subprocesses_lock:
        snapshot = list(_tracked_subprocesses)
    for rid, proc in snapshot:
        if run_id is not None and rid is not None and rid != run_id:
            continue
        if proc.poll() is not None:
            continue
        try:
            cmd_str = " ".join(proc.args) if isinstance(proc.args, list) else str(proc.args)
        except Exception:
            cmd_str = "<unknown>"
        try:
            proc.kill()
            killed.append(cmd_str)
        except Exception:
            pass
    return killed


def _register_httpx_client(client: httpx.Client) -> None:
    with _active_httpx_clients_lock:
        _active_httpx_clients[threading.get_ident()] = client

def _unregister_httpx_client() -> None:
    with _active_httpx_clients_lock:
        _active_httpx_clients.pop(threading.get_ident(), None)

def _close_active_httpx_clients() -> None:
    with _active_httpx_clients_lock:
        for tid, client in list(_active_httpx_clients.items()):
            try:
                client.close()
            except Exception:
                pass
        _active_httpx_clients.clear()

def signal_cancel_run(run_id: str) -> None:
    ev = _cancel_events.get(run_id)
    if ev:
        ev.set()
    _kill_tracked_subprocesses(run_id)


def signal_cancel_current_run() -> None:
    _cancel_current_run.set()
    _close_active_httpx_clients()
    _kill_tracked_subprocesses(None)


def cancel_event_for_run(run_id: str) -> threading.Event:
    if run_id not in _cancel_events:
        _cancel_events[run_id] = threading.Event()
    return _cancel_events[run_id]


def load_format_visual_archetypes() -> dict[str, list[dict[str, str]]]:
    raw = COPY_PROMPTS.get("visual_archetypes") or {}
    out: dict[str, list[dict[str, str]]] = {}
    for fmt in FORMATS:
        items = raw.get(fmt) or []
        out[fmt] = [
            {"id": str(item.get("id") or ""), "label": str(item.get("label") or item.get("id") or "")}
            for item in items
            if isinstance(item, dict) and str(item.get("id") or "").strip()
        ]
    return out



def classify_hook_structure(headline: str) -> str:
    """Classify a headline opening pattern for hypothesis sanity checks.

    The classifier only verifies whether a pattern is present.
    It does not prescribe which pattern should be used — that's the LLM's job.
    A headline is question-led if it opens as a question or contains an
    early question mark. Everything else is classified by visible pattern cues.
    """
    text = (headline or "").strip().lower().replace("’", "'")
    if not text:
        return "proof_led"
    if text.startswith(("why ", "what ", "how ", "when ", "can ", "could ", "will ", "want ", "need ", "tired ")) or "?" in text[:50]:
        return "question_led"
    if text.startswith(("stop", "start", "try", "see", "check", "take")):
        return "command_led"
    contrast_terms = ["before", "after", "without", "instead", " but ", " yet ", " still ", "doesn't have to", "even with", " not ", ", not ", " vs ", " versus ", "rather than"]
    if any(term in f" {text} " for term in contrast_terms):
        return "contrast_loop"
    if text.startswith(("i ", "my ")) or "felt" in text or "struggled" in text:
        return "confession_led"
    if text.startswith(("finally", "trusted", "proven")) or "70,000" in text or "doctor" in text:
        return "proof_led"
    return "proof_led"


def headline_for_candidate(candidate: dict[str, Any], lang: str = "EN") -> str:
    copy = candidate.get("copy") if isinstance(candidate.get("copy"), dict) else {}
    block = copy.get(lang) if isinstance(copy.get(lang), dict) else {}
    return str(block.get("headline") or "").strip()


def copy_text_for_candidate(candidate: dict[str, Any], lang: str = "EN") -> str:
    copy = candidate.get("copy") if isinstance(candidate.get("copy"), dict) else {}
    block = copy.get(lang) if isinstance(copy.get(lang), dict) else {}
    parts: list[str] = []
    for key in ["headline", "subheadline", "support_line", "trust_line", "attribution", "cta"]:
        value = block.get(key)
        if isinstance(value, str) and value.strip():
            parts.append(value.strip())
    bullets = block.get("bullets")
    if isinstance(bullets, list):
        parts.extend(str(item).strip() for item in bullets if str(item).strip())
    return " ".join(parts)


def cta_for_candidate(candidate: dict[str, Any], lang: str = "EN") -> str:
    copy = candidate.get("copy") if isinstance(candidate.get("copy"), dict) else {}
    block = copy.get(lang) if isinstance(copy.get(lang), dict) else {}
    return str(block.get("cta") or "").strip()


def hook_structure_mismatch(candidate: dict[str, Any], planned_ad: dict[str, Any]) -> str | None:
    hypothesis = planned_ad.get("hypothesis") if isinstance(planned_ad.get("hypothesis"), dict) else {}
    if hypothesis.get("type") != "hook_structure":
        return None
    expected = str(hypothesis.get("variant") or "").strip()
    if not expected:
        return None
    headline = headline_for_candidate(candidate, "EN")
    actual = classify_hook_structure(headline)
    if actual != expected:
        return f"Expected hook_structure {expected}, but EN headline classified as {actual}: {headline!r}"
    return None


def hypothesis_mismatch(candidate: dict[str, Any], planned_ad: dict[str, Any]) -> str | None:
    hypothesis = planned_ad.get("hypothesis") if isinstance(planned_ad.get("hypothesis"), dict) else {}
    hyp_type = hypothesis.get("type")
    expected = str(hypothesis.get("variant") or "").strip()
    if not hyp_type or hyp_type == "none" or not expected:
        return None
    if hyp_type == "hook_structure":
        return hook_structure_mismatch(candidate, planned_ad)
    if hyp_type == "concept_angle":
        actual = str(candidate.get("concept_angle") or "").strip()
        if actual != expected:
            return f"Expected concept_angle {expected}, but candidate returned {actual or 'blank'}"
    return None


HYPOTHESIS_VARIABLES: dict[str, dict[str, Any]] = {}




def resolve_language_mode(config: dict[str, Any]) -> str:
    mode = str(config.get("language_mode") or "ALL").strip().upper()
    if mode in {"EN", "HI", "HINGLISH", "ALL"}:
        return mode
    return "ALL"


def assembler_language_mode(config: dict[str, Any]) -> str:
    mode = resolve_language_mode(config)
    if mode == "EN":
        return "EN"
    if mode == "HI":
        return "HI"
    if mode == "HINGLISH":
        return "HINGLISH"
    return "BOTH"


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def make_run_id() -> str:
    return f"run_{int(time.time())}_{random.randint(1000, 9999)}"


def ensure_dirs() -> None:
    RUNS_ROOT.mkdir(parents=True, exist_ok=True)
    RUNTIME_ROOT.mkdir(parents=True, exist_ok=True)
    INPUT_IMAGES_DIR.mkdir(parents=True, exist_ok=True)


def load_env_file(path: Path) -> None:
    if not path.exists() or not path.is_file():
        return
    for raw_line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def parse_persona_library() -> list[dict[str, Any]]:
    path = PERSONA_SEEDS_PATH
    if not path.exists():
        return []
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return [{"number": int(e["persona_number"]), "name": str(e["persona_name"])} for e in data]


def _persona_name_to_slug(name: str) -> str:
    """Convert a persona name to a filename-safe slug.

    Examples:
        "Always Hungry" -> "always_hungry"
        "35+ Slow Progress Dieter" -> "35_slow_progress_dieter"
        "Ayurveda-First Buyer" -> "ayurveda_first_buyer"
    """
    s = name.strip().lower()
    s = re.sub(r"[+\-]+", " ", s)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


_PERSONA_SLUG_TO_NUMBER: dict[str, int] = {}


def _build_persona_slug_index() -> dict[str, int]:
    global _PERSONA_SLUG_TO_NUMBER
    if _PERSONA_SLUG_TO_NUMBER:
        return _PERSONA_SLUG_TO_NUMBER
    for entry in parse_persona_library():
        slug = _persona_name_to_slug(entry["name"])
        _PERSONA_SLUG_TO_NUMBER[slug] = int(entry["number"])
    return _PERSONA_SLUG_TO_NUMBER


def persona_slug(persona_number: int) -> str:
    """Return the slug for a persona number. Falls back to P{NN:02d} if unknown."""
    for entry in parse_persona_library():
        if int(entry["number"]) == int(persona_number):
            return _persona_name_to_slug(entry["name"])
    return f"P{int(persona_number):02d}"


def persona_number_from_slug(slug: str) -> int | None:
    """Return the persona number for a slug. Returns None if unknown."""
    index = _build_persona_slug_index()
    return index.get(slug)


def _load_persona_seeds() -> dict[int, dict[str, str]]:
    path = PERSONA_SEEDS_PATH
    if not path.exists():
        print(f"WARNING: {path} not found. Using empty persona seeds.", file=sys.stderr)
        return {}
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    seeds: dict[int, dict[str, str]] = {}
    for entry in data:
        pn = int(entry.get("persona_number", 0))
        if pn < 1:
            continue
        raw_attempts = str(entry.get("failed_attempts", "")).strip()
        raw_why = str(entry.get("why_it_failed", "")).strip()
        objections_raw = (raw_attempts + " " + raw_why).strip()
        seeds[pn] = {
            "core_pattern": str(entry.get("core_pattern", "")),
            "common_indian_moments": str(entry.get("common_indian_moments", "")),
            "objections_raw": objections_raw,
            "how_kit_solves": str(entry.get("relevant_ok_kit_role", "")),
            "guardrail": str(entry.get("guardrail", "")),
        }
    return seeds


PERSONA_SEED_INPUTS = _load_persona_seeds()


def _load_copy_architecture() -> dict[str, Any]:
    path = COPY_ARCH_PATH
    if not path.exists():
        print(f"WARNING: {path} not found. Copy architecture templates disabled.", file=sys.stderr)
        return {"headline_architectures": {}}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"WARNING: Failed to load {path}: {exc}", file=sys.stderr)
        return {"headline_architectures": {}}


COPY_ARCH = _load_copy_architecture()


def _load_copy_prompts() -> dict[str, Any]:
    path = COPY_PROMPTS_PATH
    if not path.exists():
        print(f"WARNING: {path} not found. Prompt templates disabled.", file=sys.stderr)
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"WARNING: Failed to load {path}: {exc}", file=sys.stderr)
        return {}


def _hypothesis_variant_label(variant_id: str) -> str:
    acronyms = {"pas", "bab", "fab"}
    if variant_id in acronyms:
        return variant_id.upper()
    return variant_id.replace("_", " ").title()


def _build_hypothesis_variables() -> dict[str, dict[str, Any]]:
    hv: dict[str, dict[str, Any]] = {
        "none": {
            "label": "No hypothesis test",
            "description": "Generate ads normally without controlled A/B testing.",
            "options": [],
        }
    }

    arch_types = {
        "hook_structure": {
            "label": "Hook Structure (H1)",
            "description": "Test which headline opening pattern performs best: question vs. proof vs. contrast vs. confession vs. command.",
        },
        "concept_angle": {
            "label": "Concept Angle (H2)",
            "description": "Test which messaging angle drives better results: pain vs. outcome vs. proof vs. authority vs. curiosity vs. comparison vs. offer vs. story.",
        },
    }
    arch = COPY_ARCH.get("headline_architectures", {})
    for hyp_type, meta in arch_types.items():
        options = [{"id": vid, "label": _hypothesis_variant_label(vid)} for vid in arch.get(hyp_type, {})]
        hv[hyp_type] = {**meta, "options": options}

    return hv


COPY_PROMPTS = _load_copy_prompts()

HYPOTHESIS_VARIABLES = _build_hypothesis_variables()


def _invalidate_config_cache(full_path: Path) -> None:
    """Reload the in-memory global for a config file that was just saved."""
    global PERSONA_SEED_INPUTS, COPY_ARCH, HYPOTHESIS_VARIABLES, COPY_PROMPTS
    if full_path == PERSONA_SEEDS_PATH:
        PERSONA_SEED_INPUTS = _load_persona_seeds()
    elif full_path == COPY_ARCH_PATH:
        COPY_ARCH = _load_copy_architecture()
        HYPOTHESIS_VARIABLES = _build_hypothesis_variables()
    elif full_path == COPY_PROMPTS_PATH:
        COPY_PROMPTS = _load_copy_prompts()
    # prompt_assembler_templates.json and background_variant.json are used by
    # subprocess scripts (generate_ads.py, chatgpt_web_sutomation.py),
    # not cached in this process — no reload needed.



def _headline_architecture_group(group: str) -> str:
    return group


def _entry_direction(entry: dict[str, Any]) -> str:
    return str(entry.get("meaning") or entry.get("intent") or entry.get("direction") or entry.get("template") or "")


def _compact_creative_entry(entry_id: str, entry: dict[str, Any]) -> dict[str, Any]:
    allowed_keys = ["meaning", "intent", "headline_role", "support_role", "avoid_skeletons", "avoid"]
    out: dict[str, Any] = {"id": entry_id}
    for key in allowed_keys:
        value = entry.get(key)
        if value not in (None, "", []):
            out[key] = value
    return out


def _framework_item(group: str, item_id: str) -> dict[str, str]:
    arch_group = _headline_architecture_group(group)
    items = COPY_ARCH.get("headline_architectures", {}).get(arch_group, {})
    if not isinstance(items, dict) or not items:
        return {"id": item_id, "direction": ""}
    entry = items.get(item_id)
    if not isinstance(entry, dict):
        first_id, first_entry = next(iter(items.items()))
        return {"id": first_id, "direction": _entry_direction(first_entry)}
    return {"id": item_id, "direction": _entry_direction(entry)}


def _hypothesis_guidance(hyp_type: str, variant: str) -> str:
    headline_group = _headline_architecture_group(hyp_type)
    headline_entry = COPY_ARCH.get("headline_architectures", {}).get(headline_group, {}).get(variant)
    if isinstance(headline_entry, dict):
        return _entry_direction(headline_entry)
    aux_entry = COPY_ARCH.get("non_headline_hypotheses", {}).get(hyp_type, {}).get(variant)
    if isinstance(aux_entry, dict):
        return _entry_direction(aux_entry)
    return ""


def _select_headline_architecture(persona_number: int, fmt: str) -> dict[str, Any]:
    arch = COPY_ARCH.get("headline_architectures", {})
    hook_arch = arch.get("hook_structure", {})
    hook_keys = list(hook_arch.keys())
    if hook_keys:
        idx = (persona_number + sum(ord(c) for c in fmt)) % len(hook_keys)
        hook_id = hook_keys[idx]
        return {"source": "hook_structure", "variant": hook_id, **hook_arch[hook_id]}
    return {"source": "four_us", "variant": "four_us", "template": "", "examples": []}


def _persona_theme(persona_seed: dict[str, Any]) -> str:
    text = " ".join(str(persona_seed.get(key) or "").lower() for key in ["pain", "desire", "friction", "proof", "tone", "persona_name"])
    if any(word in text for word in ["craving", "hunger", "snack", "food noise", "willpower"]):
        return "cravings"
    if any(word in text for word in ["digestion", "gut", "bloat", "stomach", "acidity"]):
        return "digestion"
    if any(word in text for word in ["event", "wedding", "outfit", "photo", "deadline"]):
        return "event_deadline"
    if any(word in text for word in ["busy", "professional", "work", "travel", "schedule", "office"]):
        return "busy_life"
    return ""


def _compact_product_truth() -> dict[str, Any]:
    return {
        "source": "attached_product_master_doc",
        "instruction": "Use the attached product master doc as the only source for product claims. Do not treat the request JSON as a product-claim source.",
        "hard_bans": [
            "no fat burner",
            "no metabolism boost",
            "no cure claims",
            "no guaranteed results",
            "no disease treatment claims",
            "no price in on-image copy",
            "no product component names in headline",
            "no protocol mechanics in headline",
        ],
    }


def build_copy_requirements(persona_number: int, fmt: str, format_sequence_index: int, variation_seed: str = "") -> dict[str, Any]:
    persona_seed = PERSONA_SEED_INPUTS.get(persona_number, {})

    return {
        "selection_mode": "llm_choose",
    }


def compact_format_rules_for_copy(fmt: str, format_rules: dict[str, Any]) -> dict[str, Any]:
    return {"format": fmt, "rules": []}

def build_ad_copy_system_prompt(fmt: str, formats: list[str] | None = None) -> str:
    fmt = fmt.strip().upper()
    prompts = COPY_PROMPTS
    base_rules = prompts.get("system_prompt_base_rules", [])
    parts = list(base_rules)
    if formats:
        fr_map = prompts.get("system_prompt_format_rules", {})
        for f in sorted(formats):
            rules = fr_map.get(f)
            if rules:
                parts.append("")
                parts.append(f"Ad format rules for {f}:")
                parts.extend(rules)
    return "\n".join(parts)


def build_strict_schema_note(fmt: str, languages: list[str] | None = None) -> str:
    fmt = fmt.strip().upper()
    prompts = COPY_PROMPTS.get("strict_schema_note", {})
    field_map = prompts.get("field_map", {})
    if fmt == "ALL":
        copy_fields = "the fields required by each ad's format (see Ad format rules above)"
    else:
        copy_fields = field_map.get(fmt, prompts.get("default_fields", "headline, cta"))
    parts = [
        prompts.get("intro", ""),
        prompts.get("persona_fields_en", ""),
    ]
    if languages and any("HINGLISH" in l.upper() for l in languages):
        parts.append(prompts.get("language_extension", ""))
    parts.append(prompts.get("format_closure_template", "").format(fmt=fmt or "this", copy_fields=copy_fields))
    return " ".join(parts)


def build_ad_prompt_tail(fmt: str, formats: list[str] | None = None) -> str:
    fmt = fmt.strip().upper()
    tail = COPY_PROMPTS.get("prompt_tail", {})
    support_map = tail.get("support_target_map", {})
    support_target = support_map.get(fmt, tail.get("default_support_target", "subheadline"))
    display_fmt = fmt if fmt != "ALL" else "ad"
    lines = [line.format(fmt=display_fmt, support_target=support_target) for line in tail.get("lines", [])]
    skeleton = build_response_skeleton(fmt, formats=formats)
    if skeleton:
        lines.append(f"\nReturn your response using exactly this JSON skeleton (replace placeholder values with your actual copy):\n{skeleton}")
    return "\n".join(lines)


def build_response_skeleton(fmt: str, formats: list[str] | None = None) -> str:
    fmt = fmt.strip().upper()
    skeletons = COPY_PROMPTS.get("response_skeleton", {})
    base = copy.deepcopy(skeletons.get("default", {}))

    if fmt != "ALL" or not formats:
        # Single-format skeleton
        fmt_override = skeletons.get(fmt, {})
        if fmt_override and "copy" in fmt_override:
            if "copy" in base.get("ads", [{}])[0]:
                base["ads"][0]["copy"] = copy.deepcopy(fmt_override["copy"])
        placeholder_fmt = fmt if fmt != "ALL" else "<FORMAT>"
        raw = json.dumps(base, ensure_ascii=False, indent=2)
        return raw.replace("{format}", placeholder_fmt)

    # Multi-format skeleton: one example ad per unique format in the batch
    seen: set[str] = set()
    example_ads: list[dict] = []
    for f in formats:
        f = f.strip().upper()
        if f in seen:
            continue
        seen.add(f)
        ad = copy.deepcopy(base["ads"][0])
        ad["format"] = f
        fmt_override = skeletons.get(f, {})
        if fmt_override and "copy" in fmt_override:
            if "copy" in ad:
                ad["copy"] = copy.deepcopy(fmt_override["copy"])
        example_ads.append(ad)
    base["ads"] = example_ads
    return json.dumps(base, ensure_ascii=False, indent=2)


def build_generation_payload_for_llm(context: dict[str, Any]) -> dict[str, Any]:
    seen: dict[tuple[str, int], dict[str, Any]] = {}
    total = 0
    for item in context.get("ads") or []:
        if not isinstance(item, dict):
            continue
        fmt = str(item.get("format") or "").strip().upper()
        persona = item.get("persona") if isinstance(item.get("persona"), dict) else {}
        pn = persona.get("persona_number")
        if not isinstance(pn, int):
            continue
        key = (fmt, pn)
        total += 1
        if key not in seen:
            format_rules = item.get("format_rules") if isinstance(item.get("format_rules"), dict) else {}
            copy_requirements = item.get("copy_requirements") if isinstance(item.get("copy_requirements"), dict) else {}
            seen[key] = {
                "format": fmt,
                "persona": persona,
                "format_rules": compact_format_rules_for_copy(fmt, format_rules),
                "copy_requirements": copy_requirements,
                "count": 1,
            }
        else:
            seen[key]["count"] += 1

    compact_ads = list(seen.values())

    return {
        "generated_at": context.get("generated_at"),
        "run_id": context.get("run_id"),
        "language_mode": context.get("language_mode"),
        "context_source": context.get("context_source"),
        "product_doc": {
            "attached_in_session": True,
            "source_file": context.get("product_file_path"),
            "instruction": "Read and use the attached product master doc as source of truth for all product claims.",
        },
        "product_truth": _compact_product_truth(),
        "ads": compact_ads,
    }


TARGET_LANGS_MAP = {"EN": ["EN"], "HI": ["HI"], "HINGLISH": ["HINGLISH"], "ALL": ["EN", "HI", "HINGLISH"]}

def validate_generated_copy_payload(copy_json: dict[str, Any], planned_ads: list[dict[str, Any]], language_mode: str = "ALL") -> str | None:
    target_langs = TARGET_LANGS_MAP.get(language_mode, ["EN", "HI"])
    ads = copy_json.get("ads") if isinstance(copy_json, dict) else None
    if not isinstance(ads, list):
        return "Generated payload did not include an ads array"
    if len(ads) < len(planned_ads):
        return f"Generated ads count {len(ads)} is lower than planned count {len(planned_ads)}"

    planned_keys = {
        (str(item.get("format") or "").strip().upper(), int((item.get("persona") or {}).get("persona_number") or 0))
        for item in planned_ads
        if isinstance(item, dict) and isinstance(item.get("persona"), dict)
    }
    seen_keys: set[tuple[str, int]] = set()
    for ad in ads:
        if not isinstance(ad, dict):
            return "Generated ads payload contains a non-object item"
        fmt = str(ad.get("format") or "").strip().upper()
        persona = ad.get("persona") if isinstance(ad.get("persona"), dict) else {}
        persona_number = persona.get("number")
        if not isinstance(persona_number, int):
            persona_number = persona.get("persona_number")
        if not isinstance(persona_number, int):
            return f"Generated ad for format {fmt or '?'} is missing persona number"
        seen_keys.add((fmt, persona_number))
        copy = ad.get("copy") if isinstance(ad.get("copy"), dict) else {}
        for lang in target_langs:
            block = copy.get(lang) if isinstance(copy.get(lang), dict) else {}
            if not str(block.get("headline") or "").strip():
                return f"Generated ad {fmt}/P{persona_number} is missing {lang} headline"
            if fmt in {"HERO", "UGC"} and not str(block.get("subheadline") or block.get("support_line") or "").strip():
                return f"Generated ad {fmt}/P{persona_number} is missing {lang} subheadline"
            if fmt in {"BA", "FEAT"}:
                bullets = block.get("bullets") if isinstance(block.get("bullets"), list) else []
                min_bullets = 4 if fmt == "BA" else 2
                if len([item for item in bullets if isinstance(item, str) and item.strip()]) < min_bullets:
                    return f"Generated ad {fmt}/P{persona_number} has insufficient {lang} bullets"

    missing = sorted(planned_keys - seen_keys)
    if missing:
        return "Generated payload is missing planned ads: " + ", ".join(f"{fmt}/P{persona}" for fmt, persona in missing)
    return None


def extract_generated_ad_candidate(payload: dict[str, Any]) -> dict[str, Any] | None:
    if not isinstance(payload, dict):
        return None

    def normalize_candidate(candidate: dict[str, Any]) -> dict[str, Any]:
        cloned = json.loads(json.dumps(candidate, ensure_ascii=False))
        persona = cloned.get("persona") if isinstance(cloned.get("persona"), dict) else {}
        if not isinstance(persona, dict):
            persona = {}
        if not isinstance(persona.get("persona_number"), int) and isinstance(cloned.get("persona_number"), int):
            persona["persona_number"] = cloned.get("persona_number")
        if not isinstance(persona.get("number"), int) and isinstance(cloned.get("persona_number"), int):
            persona["number"] = cloned.get("persona_number")
        if not isinstance(persona.get("persona_name"), str) and isinstance(cloned.get("persona_name"), str):
            persona["persona_name"] = cloned.get("persona_name")
        if not isinstance(persona.get("name"), str) and isinstance(cloned.get("persona_name"), str):
            persona["name"] = cloned.get("persona_name")
        if persona:
            cloned["persona"] = persona
        return cloned

    ads = payload.get("ads")
    if isinstance(ads, list):
        for item in ads:
            if isinstance(item, dict) and item.get("copy"):
                return normalize_candidate(item)
    if payload.get("format") and payload.get("copy"):
        return normalize_candidate(payload)
    return None


def detect_template_leakage(candidate: dict[str, Any] | None) -> str | None:
    if not isinstance(candidate, dict):
        return None
    copy_raw = candidate.get("copy") if isinstance(candidate.get("copy"), dict) else {}
    texts: list[str] = []
    for lang in ("EN", "HI"):
        block = copy_raw.get(lang) if isinstance(copy_raw.get(lang), dict) else copy_raw if lang == "EN" else {}
        for key in ("headline", "subheadline", "support_line", "trust_line"):
            val = block.get(key) if isinstance(block, dict) else None
            if isinstance(val, str) and val.strip():
                texts.append(val.strip())
        bullets = block.get("bullets") if isinstance(block, dict) and isinstance(block.get("bullets"), list) else []
        for b in bullets:
            if isinstance(b, str) and b.strip():
                texts.append(b.strip())

    for text in texts:
        low = text.lower()
        if any(label in low for label in ("structured_system", "cravings_down", "desired_outcome")):
            return f"Template label leaked into copy: {text!r}"
        if re.search(r"\ba\s+clear\s+.{1,30}\s+system\s+for\b", low):
            return f"Template sentence pattern detected: {text!r}"
        if "simple steps rooted in" in low:
            return f"Template sentence pattern detected: {text!r}"
        if re.match(r"^a\s+doctor-formulated\s+ayurvedic\s+kit:", low):
            return f"Colon-led feature list support line: {text!r}"
    if len(texts) >= 2:
        sup = texts[1] if len(texts) > 1 else ""
        if sup.count(",") >= 3 and sum(1 for kw in ("appetite", "digestion", "coach", "tracker", "15-day", "morning", "night") if kw in sup.lower()) >= 3:
            return f"Support line has stacked feature list: {sup!r}"
    return None


def hydrate_generated_ad_candidate(candidate: dict[str, Any], planned_ad: dict[str, Any]) -> dict[str, Any]:
    hydrated = json.loads(json.dumps(candidate, ensure_ascii=False))

    planned_format = str(planned_ad.get("format") or "").strip().upper()
    candidate_format = str(hydrated.get("format") or "").strip().upper()
    hydrated["format"] = candidate_format or planned_format

    planned_persona = planned_ad.get("persona") if isinstance(planned_ad.get("persona"), dict) else {}
    candidate_persona = hydrated.get("persona") if isinstance(hydrated.get("persona"), dict) else {}

    persona_number = candidate_persona.get("number")
    if not isinstance(persona_number, int):
        persona_number = candidate_persona.get("persona_number")
    if not isinstance(persona_number, int):
        planned_number = planned_persona.get("persona_number")
        if isinstance(planned_number, int):
            persona_number = planned_number

    persona_name = candidate_persona.get("name")
    if not isinstance(persona_name, str) or not persona_name.strip():
        persona_name = candidate_persona.get("persona_name")
    if not isinstance(persona_name, str) or not persona_name.strip():
        planned_name = planned_persona.get("persona_name")
        if isinstance(planned_name, str):
            persona_name = planned_name

    merged_persona = dict(candidate_persona)
    if isinstance(persona_number, int):
        merged_persona["number"] = persona_number
        merged_persona["persona_number"] = persona_number
    if isinstance(persona_name, str) and persona_name.strip():
        clean_name = persona_name.strip()
        merged_persona["name"] = clean_name
        merged_persona["persona_name"] = clean_name
    for k in ["pain_en", "desire_en", "friction_en", "proof_needed_en", "tone_cue_en",
              "pain_hi", "desire_hi", "friction_hi", "proof_needed_hi", "tone_cue_hi"]:
        if k not in merged_persona and k in planned_persona:
            merged_persona[k] = planned_persona[k]
    if merged_persona:
        hydrated["persona"] = merged_persona

    copy_payload = hydrated.get("copy") if isinstance(hydrated.get("copy"), dict) else {}
    normalized_copy: dict[str, Any] = {}
    if copy_payload:
        has_lang_keys = any(k in copy_payload for k in ["EN", "HI"])
        if has_lang_keys:
            for lang in ["EN", "HI"]:
                lang_block = copy_payload.get(lang)
                normalized_copy[lang] = lang_block if isinstance(lang_block, dict) else {}
        else:
            normalized_copy["EN"] = copy_payload
            normalized_copy["HI"] = {}
    else:
        normalized_copy = {"EN": {}, "HI": {}}
    hydrated["copy"] = normalized_copy

    return hydrated


def _normalize_how_kit_solves(value: Any) -> dict[str, str]:
    if isinstance(value, str):
        trimmed = value.strip()
        return {"kit_lever": trimmed} if trimmed else {}
    if isinstance(value, dict):
        if isinstance(value.get("how_kit_solves"), dict):
            value = value["how_kit_solves"]
        allowed = ["failure_point", "kit_lever", "causal_bridge", "support_line_instruction"]
        return {k: str(value.get(k, "")).strip() for k in allowed if str(value.get(k, "")).strip()}
    return {}


def _build_persona_payload_field(seed_field_value: Any, config: dict[str, Any]) -> Any:
    wrap = config.get("wrap_list", False)
    prefix = config.get("prefix")
    if wrap:
        val = str(seed_field_value) if seed_field_value else ""
        return [f"{prefix}{val}" if prefix and val else val]
    return seed_field_value

def build_persona_payload(persona_number: int, personas: list[dict[str, Any]]) -> dict[str, Any]:
    persona_name = f"Persona {persona_number}"
    for item in personas:
        if int(item.get("number") or 0) == persona_number:
            name = str(item.get("name") or "").strip()
            if name:
                persona_name = name
            break
    seed = PERSONA_SEED_INPUTS.get(persona_number, {})
    mapping = _PERSONA_SEED_MAPPING
    seed_to_payload = mapping.get("seed_to_payload", {})
    fallbacks = mapping.get("persona_fallbacks", {})

    payload: dict[str, Any] = {
        "persona_number": persona_number,
        "persona_name": persona_name,
    }

    for seed_key, field_cfg in seed_to_payload.items():
        raw = seed.get(seed_key, fallbacks.get(seed_key, ""))
        if seed_key == "how_kit_solves":
            raw = _normalize_how_kit_solves(raw)
        payload[field_cfg["field"]] = _build_persona_payload_field(raw, field_cfg)

    static = mapping.get("static_fields", {})
    for key, val in static.items():
        payload[key] = val

    hindi_default = mapping.get("hindi_ready_default", "")
    if hindi_default and payload.get("hindi_ready") in (None, []):
        payload["hindi_ready"] = [hindi_default]

    return payload


def read_active_images(path: Path) -> list[str]:
    if not path.exists():
        return []
    lines = [line.strip() for line in path.read_text(encoding="utf-8").splitlines()]
    return [line for line in lines if line and not line.startswith("#")]


def default_image_sources_file() -> Path:
    if DEFAULT_IMAGE_SOURCES_FILE.exists():
        return DEFAULT_IMAGE_SOURCES_FILE
    return LEGACY_ACTIVE_IMAGES_FILE


def list_input_images() -> list[str]:
    if not INPUT_IMAGES_DIR.exists():
        return []
    allowed = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}
    items = [
        p for p in sorted(INPUT_IMAGES_DIR.iterdir())
        if p.is_file() and p.suffix.lower() in allowed
    ]
    return [str(p.relative_to(ROOT)).replace("\\", "/") for p in items]


def default_product_doc_info() -> dict[str, Any]:
    return {
        "path": str(DEFAULT_PRODUCT_MASTER.relative_to(ROOT)).replace("\\", "/"),
        "name": DEFAULT_PRODUCT_MASTER.name,
        "exists": DEFAULT_PRODUCT_MASTER.exists(),
        "size_bytes": DEFAULT_PRODUCT_MASTER.stat().st_size if DEFAULT_PRODUCT_MASTER.exists() else 0,
    }


def store_uploaded_input_images(files: list[UploadFile], clear_existing: bool) -> list[str]:
    ensure_dirs()
    if clear_existing and INPUT_IMAGES_DIR.exists():
        for existing in INPUT_IMAGES_DIR.iterdir():
            if existing.is_file():
                existing.unlink(missing_ok=True)

    allowed = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}
    saved: list[str] = []
    for upload in files:
        filename = Path(upload.filename or "").name
        if not filename:
            continue
        ext = Path(filename).suffix.lower()
        if ext not in allowed:
            continue
        target = INPUT_IMAGES_DIR / filename
        counter = 1
        while target.exists():
            target = INPUT_IMAGES_DIR / f"{Path(filename).stem}_{counter}{ext}"
            counter += 1
        data = upload.file.read()
        target.write_bytes(data)
        saved.append(str(target.relative_to(ROOT)).replace("\\", "/"))
    return saved


def api_delete_input_image(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    rel_path = str(payload.get("path") or "").strip().replace("\\", "/")
    if not rel_path.startswith("input/images/"):
        raise HTTPException(status_code=400, detail="path must be under input/images")
    target = (ROOT / rel_path).resolve()
    images_root = INPUT_IMAGES_DIR.resolve()
    if images_root not in target.parents:
        raise HTTPException(status_code=400, detail="Invalid image path")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="Input image not found")
    target.unlink()
    return {"status": "deleted", "path": rel_path}


def api_upload_input_images(
    files: list[UploadFile] = File(...),
    clear_existing: bool = Form(False),
) -> dict[str, Any]:
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")
    saved = store_uploaded_input_images(files, clear_existing)
    return {
        "status": "ok",
        "saved": saved,
        "input_images": list_input_images(),
    }


def api_product_doc() -> dict[str, Any]:
    info = default_product_doc_info()
    content = DEFAULT_PRODUCT_MASTER.read_text(encoding="utf-8", errors="ignore") if DEFAULT_PRODUCT_MASTER.exists() else ""
    return {**info, "content": content}


def api_save_product_doc(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    content = str(payload.get("content") or "")
    DEFAULT_PRODUCT_MASTER.parent.mkdir(parents=True, exist_ok=True)
    DEFAULT_PRODUCT_MASTER.write_text(content, encoding="utf-8")
    return {"status": "saved", **default_product_doc_info()}


def api_prompt_file_content(prompt_path: str = "") -> dict[str, Any]:
    """Return the full text of a prompt file."""
    if not prompt_path:
        raise HTTPException(status_code=400, detail="prompt_path is required")
    full_path = ROOT / prompt_path
    if not full_path.exists() or not full_path.is_file():
        raise HTTPException(status_code=404, detail=f"Prompt file not found: {prompt_path}")
    content = full_path.read_text(encoding="utf-8", errors="ignore")
    return {"content": content, "path": prompt_path}


def api_save_prompt_file_content(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Save full content of a prompt file."""
    prompt_path = str(payload.get("prompt_path") or "").strip()
    content = str(payload.get("content") or "")
    if not prompt_path:
        raise HTTPException(status_code=400, detail="prompt_path is required")
    full_path = ROOT / prompt_path
    if not full_path.exists():
        raise HTTPException(status_code=404, detail=f"Prompt file not found: {prompt_path}")
    full_path.write_text(content, encoding="utf-8")
    _invalidate_config_cache(full_path)
    return {"status": "saved", "path": prompt_path}


def api_input_prompt(prompt_type: str = "916_conversion") -> dict[str, Any]:
    """Return the content of an input prompt file."""
    path_map = {
        "916_conversion": CONVERT_916_TEMPLATE_PATH,
        "starting_prompt": STARTING_PROMPT_PATH,
    }
    p = path_map.get(prompt_type)
    if not p or not p.exists():
        raise HTTPException(status_code=404, detail=f"Input prompt not found: {prompt_type}")
    return {"content": p.read_text(encoding="utf-8"), "path": str(p.relative_to(ROOT))}


def api_save_input_prompt(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Save an input prompt file."""
    prompt_type = str(payload.get("prompt_type") or "").strip()
    content = str(payload.get("content") or "")
    path_map = {
        "916_conversion": CONVERT_916_TEMPLATE_PATH,
        "starting_prompt": STARTING_PROMPT_PATH,
    }
    p = path_map.get(prompt_type)
    if not p:
        raise HTTPException(status_code=400, detail="prompt_type must be '916_conversion' or 'starting_prompt'")
    p.write_text(content, encoding="utf-8")
    return {"status": "saved", "path": str(p.relative_to(ROOT))}



def _append_opencode_queue_log(message: str) -> None:
    pass


def dashboard_subprocess_env() -> dict[str, str]:
    env = dict(os.environ)
    if sys.platform == "win32":
        sep = ";"
        venv_lib = Path(sys.executable).parent.parent / "Lib" / "site-packages"
    else:
        sep = ":"
        venv_lib = Path(sys.executable).parent.parent / "lib" / f"python{sys.version_info.major}.{sys.version_info.minor}" / "site-packages"
    playwright_path = str(venv_lib)
    if playwright_path not in env.get("PYTHONPATH", ""):
        env["PYTHONPATH"] = playwright_path + (sep + env["PYTHONPATH"] if env.get("PYTHONPATH") else "")
    return env


@contextmanager
def _opencode_queue_slot(label: str) -> Iterator[None]:
    OPENCODE_QUEUE_DIR.mkdir(parents=True, exist_ok=True)
    queued_at = time.time()
    logged_wait = False
    while True:
        for slot in range(OPENCODE_MAX_CONCURRENT):
            lock_path = OPENCODE_QUEUE_DIR / f"slot_{slot}.lock"
            lock_handle = lock_path.open("a+")
            acquired = False
            try:
                if sys.platform == "win32":
                    try:
                        msvcrt.locking(lock_handle.fileno(), msvcrt.LK_NBLCK, 1)
                        acquired = True
                    except OSError:
                        pass
                else:
                    try:
                        fcntl.flock(lock_handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                        acquired = True
                    except BlockingIOError:
                        pass
            except Exception:
                pass
            if not acquired:
                lock_handle.close()
                continue
            wait_seconds = time.time() - queued_at
            if wait_seconds >= 0.25:
                _append_opencode_queue_log(f"{label} started slot={slot} wait_seconds={wait_seconds:.1f}")
            try:
                yield
                return
            finally:
                try:
                    if sys.platform == "win32":
                        lock_handle.seek(0)
                        msvcrt.locking(lock_handle.fileno(), msvcrt.LK_UNLCK, 1)
                    else:
                        fcntl.flock(lock_handle.fileno(), fcntl.LOCK_UN)
                finally:
                    lock_handle.close()
        if not logged_wait:
            _append_opencode_queue_log(f"{label} queued max_concurrent={OPENCODE_MAX_CONCURRENT}")
            logged_wait = True
        time.sleep(0.25)


def run_cmd(
    cmd: list[str],
    cwd: Path,
    *,
    run_id: str | None = None,
    poll_cancel: bool = True,
) -> subprocess.CompletedProcess[str]:
    env = dashboard_subprocess_env()
    proc = subprocess.Popen(
        cmd,
        cwd=str(cwd),
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=env,
        encoding="utf-8",
        errors="replace",
    )
    _register_subprocess(run_id, proc)
    cancel_event = cancel_event_for_run(run_id) if run_id else _cancel_current_run
    poll_interval = 0.1
    try:
        while True:
            if poll_cancel and cancel_event.is_set():
                proc.kill()
                stdout, stderr = proc.communicate()
                return subprocess.CompletedProcess(
                    proc.args,
                    proc.returncode if proc.returncode is not None else -1,
                    stdout or "",
                    (stderr or "") + "\n[killed: cancel signaled]",
                )
            try:
                stdout, stderr = proc.communicate(timeout=poll_interval)
                return subprocess.CompletedProcess(proc.args, proc.returncode, stdout, stderr)
            except subprocess.TimeoutExpired:
                continue
    finally:
        _unregister_subprocess(proc)


def generated_image_roots() -> list[Path]:
    return [GENERATED_IMAGES_ROOT]


def ensure_916_conversion_template() -> Path:
    CONVERT_916_TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    return CONVERT_916_TEMPLATE_PATH


def build_916_conversion_prompt_job(fmt: str, persona_num: int, lang: str, index: int, source_stem: str = "") -> str:
    """Build the prompt filename for a 9:16 conversion job.

    If ``source_stem`` is given (the stem of the source 4:5 image, e.g.
    ``BA_always_hungry_EN_pain_point``), the 9:16 prompt reuses it so the
    generated 9:16 image lands in ``9_16/`` with the same stem as its 4:5
    source. Otherwise falls back to a synthesized name with an index suffix.
    """
    if source_stem:
        clean = source_stem.strip()
        if clean:
            return f"{clean}.txt"
    fmt_clean = (fmt or "HERO").strip().upper() or "HERO"
    lang_clean = (lang or "EN").strip().upper() or "EN"
    persona_safe = max(0, int(persona_num or 0))
    persona_part = persona_slug(persona_safe) if persona_safe > 0 else f"persona_{int(index):02d}"
    if persona_safe > 0:
        return f"{fmt_clean}_{persona_part}_{lang_clean}_A{max(1, int(index)):02d}.txt"
    return f"{fmt_clean}_{persona_part}_{lang_clean}.txt"


def collect_45_reference_jobs_for_batch(batch: str) -> list[dict[str, Any]]:
    summary = load_batch_image_summary(batch)
    jobs: list[dict[str, Any]] = []
    seen_refs: set[str] = set()

    for entry in summary:
        prompt_file = str(entry.get("prompt_file") or "").strip().replace("\\", "/")
        saved_files = entry.get("saved_files") if isinstance(entry.get("saved_files"), list) else []
        if not prompt_file or not saved_files:
            continue

        parsed = parse_prompt_filename(prompt_file)
        if not parsed:
            continue
        fmt, lang, persona_num = parsed
        if persona_num is None:
            continue

        for candidate in saved_files:
            c = str(candidate or "").strip().replace("\\", "/")
            if not c:
                continue
            if c in seen_refs:
                continue
            image_abs = (ROOT / c).resolve()
            if not image_abs.exists() or not image_abs.is_file():
                continue

            seen_refs.add(c)
            jobs.append(
                {
                    "format": fmt.upper(),
                    "persona_number": int(persona_num),
                    "language": lang.upper(),
                    "image_rel": c,
                    "image_abs": str(image_abs),
                }
            )

    if jobs:
        return jobs

    # Fallback: derive from prompt files + filesystem scan under 4_5
    prompt_files = scan_prompt_files_for_batch(batch)
    for prompt_file in prompt_files:
        if "/45/" not in str(prompt_file):
            continue
        parsed = parse_prompt_filename(prompt_file)
        if not parsed:
            continue
        fmt, lang, persona_num = parsed
        if persona_num is None:
            continue
        # New filename format uses the persona slug (e.g. "always_hungry"), not "p01".
        # Match on the slug OR the legacy p{NN} pattern for older batches.
        slug = persona_slug(persona_num)
        patterns = [f"*{slug}*", f"*p{persona_num:02d}*"]
        for img_root in generated_image_roots():
            ref_dir = img_root / batch / "4_5"
            if not ref_dir.exists():
                continue
            for ext in ("png", "jpg", "jpeg", "webp"):
                seen_in_pattern: set[str] = set()
                for pattern in patterns:
                    for f in sorted(ref_dir.glob(f"**/{pattern}.{ext}")):
                        rel = str(f.relative_to(ROOT)).replace("\\", "/")
                        if rel in seen_in_pattern or rel in seen_refs:
                            continue
                        seen_in_pattern.add(rel)
                        image_abs = (ROOT / rel).resolve()
                        if not image_abs.exists() or not image_abs.is_file():
                            continue
                        seen_refs.add(rel)
                        jobs.append(
                            {
                                "format": fmt.upper(),
                                "persona_number": int(persona_num),
                                "language": lang.upper(),
                                "image_rel": rel,
                                "image_abs": str(image_abs),
                            }
                        )

    return jobs


def image_static_route_for_path(path: str) -> str:
    normalized = path.replace("\\", "/")
    if normalized.startswith("generated_images/"):
        return f"/generated_images/{normalized.removeprefix('generated_images/')}"
    return f"/generated_images/{normalized}"


def gemini_debugger_args() -> list[str]:
    address = resolve_gemini_debugger_address()
    return ["--attach-debugger-address", address]


def debugger_endpoint_reachable(address: str) -> bool:
    if not address:
        return False
    url = f"http://{address}/json/version"
    try:
        with urllib.request.urlopen(url, timeout=1.5) as resp:
            return resp.status == 200
    except Exception:
        return False


def detect_wsl_windows_host_ip() -> str:
    """Return the Windows host IP from WSL's perspective (e.g., 172.18.160.1).
    Returns '127.0.0.1' if not in WSL or detection fails."""
    if not Path("/mnt/c").exists():
        return "127.0.0.1"
    try:
        ip_route = subprocess.run(
            ["ip", "route"], capture_output=True, text=True, timeout=5
        )
        for line in (ip_route.stdout or "").splitlines():
            if "default" in line:
                parts = line.split()
                if len(parts) >= 3:
                    return parts[2]
    except Exception:
        pass
    return "127.0.0.1"


def detect_wsl_user() -> str:
    """Return the current WSL user (matches /mnt/c/Users/<name> for that user's home).
    Empty string if detection fails."""
    user = os.getenv("USER") or os.getenv("USERNAME")
    if user:
        return user
    home = str(Path.home())
    if home.startswith("/home/"):
        parts = home.split("/")
        if len(parts) >= 3:
            return parts[2]
    return ""


def wsl_chrome_cdp_url() -> str:
    """Return the CDP URL for Windows Chrome from WSL.
    Uses the portproxy on 9223 → Windows 9222. Set up via scripts/setup_cdp_proxy.ps1."""
    return f"http://{detect_wsl_windows_host_ip()}:9223"


def resolve_gemini_debugger_address() -> str:
    configured = str(os.getenv("GEMINI_DEBUGGER_ADDRESS") or "").strip()
    candidates = [configured] if configured else []
    candidates.extend(["127.0.0.1:9222", "localhost:9222", "127.0.0.1:9223", "localhost:9223"])
    for candidate in candidates:
        if candidate and debugger_endpoint_reachable(candidate):
            return candidate
    # No reachable endpoint now: return preferred default so automation script
    # can auto-launch a debuggable Chrome session and continue.
    return configured or "127.0.0.1:9222"


def run_gemini_generation(
    *,
    batch: str,
    prompt_files: list[str],
    aspect_ratio: str,
    image_sources_file: str | None,
    prompt_reference_map: Path | None = None,
    headless: bool = False,
    run_dir: Path | None = None,
    prepend_starting_prompt: bool = True,
    first_tab_mode: str = "reuse-blank",
) -> subprocess.CompletedProcess[str]:
    aspect_folder = "9_16" if aspect_ratio == "9:16" else "4_5"
    prompt_work_dir = RUNTIME_ROOT / "gemini_selected_prompts" / f"{batch}_{aspect_folder}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
    prompt_work_dir.mkdir(parents=True, exist_ok=True)

    starting_prompt = ""
    if prepend_starting_prompt:
        starting_prompt_path = ROOT / "input" / "startingprompt.txt"
        starting_prompt = starting_prompt_path.read_text(encoding="utf-8").strip() if starting_prompt_path.exists() else ""
    for prompt_file in prompt_files:
        source = Path(prompt_file)
        if not source.is_absolute():
            source = ROOT / source
        source = source.resolve()
        if not source.exists():
            raise RuntimeError(f"Prompt file not found: {source}")
        prompt_text = source.read_text(encoding="utf-8")
        combined = f"{starting_prompt}\n\n{prompt_text.strip()}\n" if starting_prompt else prompt_text
        (prompt_work_dir / source.name).write_text(combined, encoding="utf-8")
        sidecar = source.with_suffix(".json")
        if sidecar.exists():
            (prompt_work_dir / sidecar.name).write_text(sidecar.read_text(encoding="utf-8"), encoding="utf-8")

    out_dir = GENERATED_IMAGES_ROOT / batch / aspect_folder
    image_source_arg = image_sources_file
    if prompt_reference_map is not None:
        try:
            reference_payload = json.loads(prompt_reference_map.read_text(encoding="utf-8"))
        except Exception as exc:
            raise RuntimeError(f"Could not read prompt reference map: {exc}") from exc
        flattened_sources: list[str] = []
        if isinstance(reference_payload, dict):
            for value in reference_payload.values():
                if isinstance(value, list):
                    for item in value:
                        if isinstance(item, str) and item.strip() and item.strip() not in flattened_sources:
                            flattened_sources.append(item.strip())
        if flattened_sources:
            source_file = prompt_work_dir / "image_sources.txt"
            source_file.write_text("\n".join(flattened_sources) + "\n", encoding="utf-8")
            image_source_arg = str(source_file)

    cmd = [
        sys.executable,
        "scripts/gemini_web_automation.py",
        "--prompt-dir",
        str(prompt_work_dir),
        "--prompt-glob",
        "*.txt",
        "--out-dir",
        str(out_dir),
        "--aspect-ratio",
        aspect_ratio,
        "--timeout",
        str(int(os.getenv("GEMINI_GENERATION_TIMEOUT_SECONDS") or "420")),
        "--manual-login-timeout",
        str(int(os.getenv("GEMINI_MANUAL_LOGIN_TIMEOUT_SECONDS") or "180")),
        "--upload-dir",
        str(INPUT_IMAGES_DIR),
    ]
    if headless:
        cmd.append("--headless")
    if first_tab_mode and first_tab_mode != "reuse-blank":
        cmd.extend(["--first-tab-mode", first_tab_mode])
    if image_source_arg:
        cmd.extend(["--image-source-file", image_source_arg])
    if run_dir is not None:
        hyp_path = run_dir / "context" / "hypothesis_config.json"
        if hyp_path.exists():
            cmd.extend(["--hypothesis-config", str(hyp_path)])

    log_dir = RUNTIME_ROOT / "generation_logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"gen_{batch}_{aspect_folder}.log"

    env = dashboard_subprocess_env()

    with open(log_path, "w") as log_file:
        result = subprocess.run(cmd, cwd=str(ROOT), text=True, stdout=log_file, stderr=subprocess.STDOUT, check=False, env=env)

    full_output = log_path.read_text() if log_path.exists() else ""
    result.stdout = full_output
    result.stderr = ""
    return result


def run_chatgpt_generation(
    *,
    batch: str,
    prompt_files: list[str],
    aspect_ratio: str,
    image_sources_file: str | None,
    headless: bool = False,
    run_dir: Path | None = None,
    prepend_starting_prompt: bool = True,
    first_tab_mode: str = "reuse-blank",
) -> subprocess.CompletedProcess[str]:
    aspect_folder = "9_16" if aspect_ratio == "9:16" else "4_5"
    prompt_work_dir = RUNTIME_ROOT / "chatgpt_selected_prompts" / f"{batch}_{aspect_folder}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
    prompt_work_dir.mkdir(parents=True, exist_ok=True)

    starting_prompt = ""
    if prepend_starting_prompt:
        starting_prompt_path = ROOT / "input" / "startingprompt.txt"
        starting_prompt = starting_prompt_path.read_text(encoding="utf-8").strip() if starting_prompt_path.exists() else ""
    for prompt_file in prompt_files:
        source = Path(prompt_file)
        if not source.is_absolute():
            source = ROOT / prompt_file
        source = source.resolve()
        if not source.exists():
            raise RuntimeError(f"Prompt file not found: {source}")
        prompt_text = source.read_text(encoding="utf-8")
        combined = f"{starting_prompt}\n\n{prompt_text.strip()}\n" if starting_prompt else prompt_text
        (prompt_work_dir / source.name).write_text(combined, encoding="utf-8")
        sidecar = source.with_suffix(".json")
        if sidecar.exists():
            (prompt_work_dir / sidecar.name).write_text(sidecar.read_text(encoding="utf-8"), encoding="utf-8")

    out_dir = GENERATED_IMAGES_ROOT / batch / aspect_folder
    out_dir.mkdir(parents=True, exist_ok=True)

    cmd = [
        sys.executable,
        "scripts/chatgpt_web_sutomation.py",
        "--prompt-dir",
        str(prompt_work_dir),
        "--prompt-glob",
        "*.txt",
        "--out-dir",
        str(out_dir),
        "--timeout",
        str(int(os.getenv("CHATGPT_GENERATION_TIMEOUT_SECONDS") or "420")),
        "--download-timeout",
        str(int(os.getenv("CHATGPT_DOWNLOAD_TIMEOUT_SECONDS") or "90")),
        "--manual-login-timeout",
        str(int(os.getenv("CHATGPT_MANUAL_LOGIN_TIMEOUT_SECONDS") or "180")),
        "--upload-dir",
        str(INPUT_IMAGES_DIR),
    ]
    if headless:
        cmd.append("--headless")
    if first_tab_mode and first_tab_mode != "reuse-blank":
        cmd.extend(["--first-tab-mode", first_tab_mode])
    if image_sources_file:
        cmd.extend(["--image-source-file", image_sources_file])
    cmd.extend(["--aspect-ratio", aspect_ratio])

    if Path("/mnt/c").exists():
        cdp_url = wsl_chrome_cdp_url()
        cmd.extend(["--cdp-url", cdp_url])

    env = dashboard_subprocess_env()

    result = subprocess.run(cmd, cwd=str(ROOT), text=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=False, env=env)

    full_output = ""
    result.stdout = full_output
    result.stderr = ""
    return result


def build_multipart_form(fields: dict[str, str], file_field: str, file_path: Path) -> tuple[bytes, str]:
    boundary = f"----dashboard{uuid.uuid4().hex}"
    lines: list[bytes] = []

    for key, value in fields.items():
        lines.append(f"--{boundary}\r\n".encode("utf-8"))
        lines.append(f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode("utf-8"))
        lines.append(f"{value}\r\n".encode("utf-8"))

    mime_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
    lines.append(f"--{boundary}\r\n".encode("utf-8"))
    lines.append(
        (
            f'Content-Disposition: form-data; name="{file_field}"; filename="{file_path.name}"\r\n'
            f"Content-Type: {mime_type}\r\n\r\n"
        ).encode("utf-8")
    )
    lines.append(file_path.read_bytes())
    lines.append(b"\r\n")
    lines.append(f"--{boundary}--\r\n".encode("utf-8"))

    body = b"".join(lines)
    content_type = f"multipart/form-data; boundary={boundary}"
    return body, content_type


def upload_image_to_cloudinary(image_path: Path, cloud_name: str, api_key: str, api_secret: str) -> str:
    if not image_path.exists() or not image_path.is_file():
        raise RuntimeError(f"Image not found for upload: {image_path}")

    timestamp = str(int(time.time()))
    signature_base = f"timestamp={timestamp}{api_secret}"
    signature = hashlib.sha1(signature_base.encode("utf-8")).hexdigest()

    fields = {
        "api_key": api_key,
        "timestamp": timestamp,
        "signature": signature,
    }
    body, content_type = build_multipart_form(fields, "file", image_path)
    upload_url = f"https://api.cloudinary.com/v1_1/{cloud_name}/image/upload"
    req = urllib.request.Request(
        url=upload_url,
        data=body,
        method="POST",
        headers={"Content-Type": content_type},
    )
    with urllib.request.urlopen(req, timeout=180) as response:
        raw = response.read().decode("utf-8")
    payload = json.loads(raw)
    secure_url = str(payload.get("secure_url") or "").strip()
    if not secure_url:
        raise RuntimeError(f"Cloudinary upload did not return secure_url: {payload}")
    return secure_url


def load_batch_image_summary(batch: str) -> list[dict[str, Any]]:
    summary_path = GENERATED_IMAGES_ROOT / batch / "batch_run_summary.json"
    if not summary_path.exists():
        jobs_by_prompt: dict[str, dict[str, Any]] = {}
        for generated_root in generated_image_roots():
            generated_batch_dir = generated_root / batch
            if not generated_batch_dir.exists():
                continue
            for meta_file in sorted(generated_batch_dir.glob("**/*.json")):
                try:
                    payload = json.loads(meta_file.read_text(encoding="utf-8"))
                except Exception:
                    continue
                if not isinstance(payload, dict):
                    continue
                rec_type = str(payload.get("type") or payload.get("record_type") or "").strip()
                if rec_type not in ("ad_image", "generated_image", "gemini_ad_image", "chatgpt_ad_image"):
                    continue

                prompt_file = str(payload.get("prompt_file_relative") or payload.get("prompt_file") or "").strip().replace("\\", "/")
                saved_file = str(payload.get("saved_file") or "").strip().replace("\\", "/")
                if not prompt_file or not saved_file:
                    continue

                existing = jobs_by_prompt.get(prompt_file)
                if not existing:
                    fmt = payload.get("format") or payload.get("format_id") or ""
                    lang = payload.get("language") or payload.get("lang_id") or ""
                    existing = {
                        "prompt_file": prompt_file,
                        "saved_files": [],
                        "format": fmt,
                        "language": lang,
                        "variation": payload.get("variation"),
                        "task_id": payload.get("task_id"),
                        "prompt_metadata": payload.get("prompt_metadata") or {},
                    }
                    jobs_by_prompt[prompt_file] = existing
                saved_files = existing.get("saved_files")
                if not isinstance(saved_files, list):
                    saved_files = []
                    existing["saved_files"] = saved_files
                if saved_file not in saved_files:
                    saved_files.append(saved_file)

        return list(jobs_by_prompt.values())
    try:
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    jobs = summary.get("jobs")
    if isinstance(jobs, list):
        return [job for job in jobs if isinstance(job, dict)]
    return []


def strip_ansi(text: str | None) -> str:
    if not text:
        return ""
    return re.sub(r"\x1B\[[0-?]*[ -/]*[@-~]", "", text)










def list_opencode_models(api_url: str | None = None) -> list[str]:
    url = (api_url or "").strip() or os.getenv("OPENCODE_API_URL", "").strip() or DEFAULT_OPENCODE_API_URL
    api_key = os.getenv("OPENCODE_API_KEY", "").strip()
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    try:
        with httpx.Client(timeout=httpx.Timeout(15, connect=10)) as client:
            resp = client.get(f"{url}/models", headers=headers)
        if resp.status_code != 200:
            return []
        try:
            data = resp.json()
        except json.JSONDecodeError:
            return []
    except (httpx.HTTPError, OSError, json.JSONDecodeError):
        return []
    if isinstance(data, list):
        model_ids = [m["id"] if isinstance(m, dict) else str(m) for m in data]
    elif isinstance(data, dict):
        models_data = data.get("data") or data.get("models") or []
        model_ids = [m["id"] if isinstance(m, dict) else str(m) for m in models_data]
    else:
        return []
    return [m for m in model_ids if "/" in m] or [f"opencode/{m}" for m in model_ids]








def choose_openai_gpt52(models: list[str]) -> str:
    if not models:
        return ""
    preferred = "openai/gpt-5.2"
    if preferred in models:
        return preferred
    for model in models:
        lower = model.lower()
        if lower.startswith("openai/") and "gpt-5.2" in lower:
            return model
    for model in models:
        if model.lower().startswith("openai/"):
            return model
    non_copilot = [m for m in models if not m.lower().startswith("github-copilot/")]
    if non_copilot:
        return non_copilot[0]
    return models[0]


def sanitize_dashboard_model(selected: str, models: list[str]) -> str:
    chosen = (selected or "").strip()
    if chosen and (not models or chosen in models):
        return chosen
    return choose_openai_gpt52(models)


def build_opencode_catalog() -> dict[str, Any]:
    models = list_opencode_models()
    grouped: dict[str, list[str]] = {}
    for model in models:
        provider = model.split("/", 1)[0]
        grouped.setdefault(provider, []).append(model)
    for provider in grouped:
        grouped[provider] = sorted(grouped[provider])
    providers = sorted(grouped.keys())
    default_model = ""
    if models:
        default_model = choose_openai_gpt52(models)
    return {
        "api_url": DEFAULT_OPENCODE_API_URL,
        "providers": providers,
        "models_by_provider": grouped,
        "default_model": default_model,
    }


def parse_json_stdout(result: subprocess.CompletedProcess[str], context: str) -> Any:
    if result.returncode != 0:
        raise RuntimeError(f"{context} failed: {result.stderr.strip() or result.stdout.strip()}")
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"{context} returned invalid JSON") from exc


def save_upload(target: Path, upload: UploadFile | None) -> Path | None:
    if upload is None or not upload.filename:
        return None
    target.parent.mkdir(parents=True, exist_ok=True)
    data = upload.file.read()
    target.write_bytes(data)
    return target


def coalesce_path(uploaded: Path | None, default_path: Path) -> Path:
    return uploaded if uploaded and uploaded.exists() else default_path


def resolve_safe_path(relative_path: str) -> Path:
    candidate = (ROOT / relative_path).resolve()
    if str(candidate).startswith(str(ROOT.resolve())):
        return candidate
    raise HTTPException(status_code=400, detail="Invalid path")


def choose_text(items: list[str], fallback: str) -> str:
    for item in items:
        clean = item.strip()
        if clean:
            return clean
    return fallback


def shorten_copy_line(text: str) -> str:
    return " ".join((text or "").split()).strip()


def strip_internal_marker(text: str) -> str:
    if not isinstance(text, str):
        return ""
    cleaned = re.sub(r"\s*\b\d{4}-\d{2}-(hero|ba|test|feat|ugc)\.?\b", "", text, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*\(\s*\d+[_-]\d+\s*\)", "", cleaned)
    cleaned = re.sub(r"\s{2,}", " ", cleaned).strip()
    return cleaned


def strip_price_tokens(text: str) -> str:
    if not isinstance(text, str):
        return ""
    cleaned = text
    cleaned = re.sub(r"\bINR\b\s*\d+[\d,]*(?:\.\d+)?", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"[₹$]\s*\d+[\d,]*(?:\.\d+)?", "", cleaned)
    cleaned = re.sub(r"\b\d+[\d,]*(?:\.\d+)?\s*(?:INR|Rs\.?|rupees?)\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b(?:price|only|discount|off|mrp)\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s{2,}", " ", cleaned).strip(" ,;:-")
    return cleaned


def strip_ba_panel_label(text: str) -> str:
    if not isinstance(text, str):
        return ""
    cleaned = text.strip()
    cleaned = re.sub(r"^\s*(?:before|after)\s*[:\-]\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"^\s*(?:पहले|बाद|पहले\s*में|बाद\s*में)\s*[:\-]\s*", "", cleaned)
    cleaned = re.sub(r"\s{2,}", " ", cleaned).strip()
    return cleaned


def strip_internal_markers_from_payload(payload: dict[str, Any]) -> dict[str, Any]:
    ads = payload.get("ads")
    if not isinstance(ads, list):
        return payload

    for ad in ads:
        if not isinstance(ad, dict):
            continue
        copy = ad.get("copy")
        if not isinstance(copy, dict):
            continue
        for lang in ["EN", "HI"]:
            block = copy.get(lang)
            if not isinstance(block, dict):
                continue
            for key in ["headline", "subheadline", "support_line", "cta", "trust_line", "attribution"]:
                if key in block and isinstance(block.get(key), str):
                    value = strip_internal_marker(block[key])
                    block[key] = strip_price_tokens(value)
            if isinstance(block.get("context_line"), str):
                context_line = strip_price_tokens(strip_internal_marker(block["context_line"]))
                if re.search(r"\bneeds\b|proof needed|tone cue|persona", context_line, flags=re.IGNORECASE):
                    block["context_line"] = ""
                else:
                    block["context_line"] = context_line
            if isinstance(block.get("bullets"), list):
                cleaned_bullets = []
                for item in block["bullets"]:
                    if not isinstance(item, str):
                        continue
                    value = strip_price_tokens(strip_internal_marker(item))
                    if value:
                        cleaned_bullets.append(value)
                block["bullets"] = cleaned_bullets
    return payload


def enforce_unique_ctas(payload: dict[str, Any], context: dict[str, Any]) -> dict[str, Any]:
    return payload


PROOF_NOTE_MARKERS = [
    "needs",
    "proof needed",
    "tone cue",
    "persona",
    "non-cure",
    "compliant",
    "weight-support framing",
]


def scrub_on_image_copy(payload: dict[str, Any]) -> dict[str, Any]:
    ads = payload.get("ads") if isinstance(payload.get("ads"), list) else []
    for ad in ads:
        if not isinstance(ad, dict):
            continue
        copy = ad.get("copy") if isinstance(ad.get("copy"), dict) else {}
        for lang in ["EN", "HI"]:
            block = copy.get(lang)
            if not isinstance(block, dict):
                continue
            ctx = block.get("context_line")
            if isinstance(ctx, str) and ctx.strip():
                lowered = ctx.lower()
                if any(marker in lowered for marker in PROOF_NOTE_MARKERS):
                    block.pop("context_line", None)
    return payload


def parse_uniqueness_collisions(error_text: str) -> list[dict[str, Any]]:
    collisions: list[dict[str, Any]] = []
    for raw_line in error_text.splitlines():
        line = raw_line.strip()
        match = re.search(r"ads\[(\d+)\]\.copy\.(EN|HI)\.([a-z_]+)", line)
        if not match:
            continue
        collisions.append(
            {
                "ad_index": int(match.group(1)),
                "language": match.group(2),
                "field": match.group(3),
                "line": line,
            }
        )
    return collisions


def parse_json_object_from_text(content: str) -> dict[str, Any] | None:
    text = (content or "").strip()
    if not text:
        return None

    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    fence = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, flags=re.DOTALL | re.IGNORECASE)
    if fence:
        try:
            parsed = json.loads(fence.group(1))
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass

    decoder = json.JSONDecoder()
    best: dict[str, Any] | None = None
    best_span = -1
    for match in re.finditer(r"\{", text):
        start = match.start()
        try:
            parsed, end = decoder.raw_decode(text[start:])
        except json.JSONDecodeError:
            continue
        if not isinstance(parsed, dict):
            continue
        span = end
        if span > best_span:
            best = parsed
            best_span = span
    return best


def parse_opencode_json_output(stdout: str) -> dict[str, Any] | None:
    text_chunks: list[str] = []
    for raw_line in (stdout or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            event = json.loads(line)
        except json.JSONDecodeError:
            continue
        if event.get("type") != "text":
            continue
        part = event.get("part") or {}
        text = part.get("text")
        if isinstance(text, str) and text.strip():
            text_chunks.append(text.strip())

    if text_chunks:
        parsed = parse_json_object_from_text("\n".join(text_chunks).strip())
        if parsed is not None:
            return parsed

    return parse_json_object_from_text((stdout or "").strip())


def _find_session_id(value: Any, session_scoped: bool = False) -> str | None:
    if isinstance(value, dict):
        event_type = str(value.get("type") or "").lower()
        scoped = session_scoped or "session" in event_type
        for key in ("sessionID", "sessionId", "session_id"):
            candidate = value.get(key)
            if isinstance(candidate, str) and candidate.strip():
                return candidate.strip()
        if scoped:
            candidate = value.get("id")
            if isinstance(candidate, str) and candidate.strip():
                return candidate.strip()
        for key, nested in value.items():
            nested_scoped = scoped or "session" in str(key).lower()
            found = _find_session_id(nested, nested_scoped)
            if found:
                return found
    elif isinstance(value, list):
        for item in value:
            found = _find_session_id(item, session_scoped)
            if found:
                return found
    return None


def parse_opencode_session_id(stdout: str) -> str | None:
    for raw_line in (stdout or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            event = json.loads(line)
        except json.JSONDecodeError:
            continue
        found = _find_session_id(event)
        if found:
            return found

    match = re.search(r'"session(?:ID|Id|_id)"\s*:\s*"([^"]+)"', stdout or "")
    if match:
        return match.group(1).strip()
    return None


def build_product_doc_bootstrap_prompt() -> str:
    return COPY_PROMPTS.get("product_doc_bootstrap_prompt", "Read the attached product master doc completely. Return only valid JSON: {\"status\":\"product_doc_loaded\"}.")


def append_run_log(run_dir: Path, filename: str, message: str) -> None:
    log_path = run_dir / "logs" / filename
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write(message.rstrip() + "\n")


def _log_llm_trace(run_id: str, label: str, model: str, request_body: dict, response_body: Any, status_code: int, duration_s: float, error: str | None = None) -> None:
    try:
        user_id = _get_run_owner(run_id) or "unknown"
        from dashboard.backend.services.run_storage import save_llm_trace
        save_llm_trace(user_id, {
            "run_id": run_id,
            "batch": label,
            "provider": "opencode" if "opencode" in label.lower() else "google",
            "model": model,
            "prompt": json.dumps(request_body, ensure_ascii=False)[:5000],
            "response": json.dumps(response_body, ensure_ascii=False)[:10000] if response_body else "",
            "duration_ms": int(duration_s * 1000),
            "status": "error" if error else "completed",
        })
    except Exception:
        pass


def call_opencode_repair_copy(
    config: dict[str, Any],
    context: dict[str, Any],
    current_copy: dict[str, Any],
    collisions: list[dict[str, Any]],
    run_dir: Path,
) -> dict[str, Any] | None:
    api_key = (config.get("opencode_api_key") or "").strip() or os.getenv("OPENCODE_API_KEY", "").strip() or os.getenv("OPENCODE_SERVER_PASSWORD", "").strip()
    api_url = (config.get("opencode_api_url") or "").strip() or os.getenv("OPENCODE_API_URL", "").strip() or DEFAULT_OPENCODE_API_URL
    model = sanitize_dashboard_model((config.get("opencode_model") or "").strip(), list_opencode_models())
    api_model = model.split("/", 1)[-1] if "/" in model else model
    if not api_url:
        return None

    payload = {
        "task": "Repair uniqueness collisions only",
        "rules": [
            "Return valid JSON only",
            "Keep existing structure and fields",
            "Only change collided fields",
            "Write creative, fresh support lines that feel specific to this ad, not generic or templated",
            "Do not add internal tags or IDs",
        ],
        "collisions": collisions,
        "current_copy": current_copy,
        "context": build_generation_payload_for_llm(context),
    }
    prompt = (
        "You are fixing ad copy JSON after uniqueness collisions. "
        "Return only corrected JSON object with keys default_aspect_ratio and ads.\n\n"
        + json.dumps(payload, ensure_ascii=False)
    )

    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    body = {
        "model": api_model,
        "messages": [
            {"role": "user", "content": prompt},
        ],
        "response_format": {"type": "json_object"},
        "temperature": 0.3,
        "max_tokens": 4096,
    }

    run_id = run_dir.name if run_dir else "repair"
    t0 = time.time()
    client = httpx.Client(timeout=httpx.Timeout(120, connect=10))
    _register_httpx_client(client)
    try:
        resp = client.post(f"{api_url}/chat/completions", headers=headers, json=body)
    except (httpx.HTTPError, OSError) as exc:
        _log_llm_trace(run_id, "repair", model, body, None, -1, time.time() - t0, error=str(exc))
        (run_dir / "logs" / "opencode_repair_error.txt").write_text(f"Repair HTTP call failed: {exc}", encoding="utf-8")
        return None
    finally:
        _unregister_httpx_client()
        client.close()

    elapsed = time.time() - t0

    if resp.status_code != 200:
        _log_llm_trace(run_id, "repair", model, body, None, resp.status_code, elapsed, error=resp.text[:2000])
        (run_dir / "logs" / "opencode_repair_error.txt").write_text(f"Repair API error {resp.status_code}\n{resp.text}", encoding="utf-8")
        return None

    try:
        data = resp.json()
        content = data["choices"][0]["message"]["content"]
    except (json.JSONDecodeError, KeyError, IndexError) as exc:
        _log_llm_trace(run_id, "repair", model, body, None, resp.status_code, elapsed, error=str(exc))
        (run_dir / "logs" / "opencode_repair_error.txt").write_text(f"Repair parse error: {exc}\n{resp.text}", encoding="utf-8")
        return None

    _log_llm_trace(run_id, "repair", model, body, data, resp.status_code, elapsed)
    return parse_opencode_json_output(content)


def _clean_str(value: Any) -> str:
    return value.strip() if isinstance(value, str) else ""


def _clean_bullets(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    out: list[str] = []
    for item in value:
        if isinstance(item, str) and item.strip():
            out.append(item.strip())
    return out


def concept_ids_from_requirements(copy_req: dict[str, Any]) -> dict[str, str]:
    concept = copy_req.get("concept_variation") if isinstance(copy_req.get("concept_variation"), dict) else {}

    def nested_id(key: str, fallback: str) -> str:
        item = concept.get(key) if isinstance(concept.get(key), dict) else {}
        value = item.get("id") if isinstance(item, dict) else ""
        return _clean_str(value) or fallback

    return {
        "concept_angle": nested_id("concept_angle", "desired_outcome"),
    }


def ensure_testimonial_headline(headline: str, lang: str, persona: dict[str, Any]) -> str:
    clean = shorten_copy_line(headline)
    guidance = _TESTIMONIAL_GUIDANCE
    cfg = guidance.get(lang, guidance.get("EN", {}))
    first_pat = cfg.get("first_person_pattern", "")
    weight_pat = cfg.get("weight_pattern", "")
    suffix = cfg.get("suffix", "")
    desire_template = cfg.get("desire_template", "")
    fallback_text = cfg.get("fallback", "")
    desire_field = cfg.get("desire_field", "")

    if lang == "EN":
        if first_pat and re.search(first_pat, clean, flags=re.IGNORECASE):
            if weight_pat and re.search(weight_pat, clean, flags=re.IGNORECASE):
                return clean
            return shorten_copy_line(f'{clean.rstrip(".")}. {suffix}')
        desire = _clean_str(persona.get(desire_field)).rstrip(".")
        if desire:
            desire_phrase = desire[:1].lower() + desire[1:] if len(desire) > 1 else desire.lower()
            return shorten_copy_line(desire_template.format(desire_phrase=desire_phrase))
        return fallback_text

    if first_pat and re.search(first_pat, clean):
        if weight_pat and re.search(weight_pat, clean):
            return clean
        return shorten_copy_line(f'{clean.rstrip("।")}। {suffix}')
    desire = _clean_str(persona.get(desire_field)).rstrip("।")
    if desire:
        return shorten_copy_line(desire_template.format(desire_phrase=desire))
    return fallback_text


def _persona_number_from_candidate(candidate: dict[str, Any]) -> int | None:
    persona = candidate.get("persona") if isinstance(candidate, dict) else None
    if not isinstance(persona, dict):
        return None
    val = persona.get("number")
    if isinstance(val, int):
        return val
    val = persona.get("persona_number")
    if isinstance(val, int):
        return val
    if isinstance(val, str) and val.strip().isdigit():
        return int(val.strip())
    return None


def _persona_name_from_candidate(candidate: dict[str, Any]) -> str:
    persona = candidate.get("persona") if isinstance(candidate, dict) else None
    if not isinstance(persona, dict):
        return ""
    return _clean_str(persona.get("name") or persona.get("persona_name") or "")


def _build_copy_skeleton(context: dict[str, Any], run_id: str) -> dict[str, Any]:
    ads: list[dict[str, Any]] = []
    token = run_id[-4:]
    for idx, item in enumerate(context["ads"], start=1):
        persona = item["persona"]
        fmt = item["format"]
        persona_num = int(persona["persona_number"])
        persona_name = persona["persona_name"]
        copy_req = item.get("copy_requirements") if isinstance(item.get("copy_requirements"), dict) else {}
        concept_ids = concept_ids_from_requirements(copy_req)

        pain_en = choose_text(persona.get("pain_points", []), f"Daily routine feels heavy and hard to sustain for persona {persona_num}.")
        desire_en = choose_text(persona.get("core_message", []), "A practical routine that feels easy to follow.")
        friction_en = choose_text(persona.get("objections", []), "Past plans felt too strict and difficult to maintain.")
        proof_en = choose_text(persona.get("trust_anchors", []), "Needs proof through clear structure and believable support.")
        tone_en = "Practical, empathetic, and confidence-building"
        pain_hi = "रोज की वजन-घटाने की दिनचर्या टूटना आसान है।"
        desire_hi = "ऐसा आसान सिस्टम चाहिए जो रोज निभ सके।"
        friction_hi = "पहले के प्लान बहुत सख्त और मुश्किल थे।"
        proof_hi = "साफ कदम, भरोसेमंद सपोर्ट और व्यावहारिक प्रमाण चाहिए।"
        tone_hi = "सरल, भरोसेमंद, और व्यावहारिक"

        if fmt in {"HERO", "UGC"}:
            copy_en: dict[str, Any] = {"headline": "", "support_line": "", "cta": ""}
            copy_hi: dict[str, Any] = {"headline": "", "support_line": "", "cta": ""}
            copy_hing: dict[str, Any] = {"headline": "", "support_line": "", "cta": ""}
        elif fmt in {"BA", "FEAT"}:
            copy_en = {"headline": "", "bullets": [], "cta": ""}
            copy_hi = {"headline": "", "bullets": [], "cta": ""}
            copy_hing = {"headline": "", "bullets": [], "cta": ""}
        else:
            copy_en = {"headline": "", "trust_line": "", "cta": ""}
            copy_hi = {"headline": "", "trust_line": "", "cta": ""}
            copy_hing = {"headline": "", "trust_line": "", "cta": ""}

        ad_payload = {
            "format": fmt,
            "headline_angle": "",
            "concept_angle": concept_ids["concept_angle"],
            "persona": {
                "number": persona_num,
                "name": persona_name,
                "pain_en": pain_en,
                "desire_en": desire_en,
                "friction_en": friction_en,
                "proof_needed_en": proof_en,
                "tone_cue_en": tone_en,
                "pain_hi": pain_hi,
                "desire_hi": desire_hi,
                "friction_hi": friction_hi,
                "proof_needed_hi": proof_hi,
                "tone_cue_hi": tone_hi,
            },
            "copy": {"EN": copy_en, "HI": copy_hi, "HINGLISH": copy_hing},
        }
        hypothesis = item.get("hypothesis") if isinstance(item.get("hypothesis"), dict) else None
        if hypothesis:
            ad_payload["hypothesis"] = hypothesis
        for key in [
            "visual_archetype",
            "visual_pattern_reused_from_run_id",
            "visual_pattern_reuse_key",
            "creative_index",
            "creative_total",
            "background_group_key",
        ]:
            if key in item:
                ad_payload[key] = item[key]
        ads.append(ad_payload)

    return {"default_aspect_ratio": "4:5", "ads": ads}


def normalize_generated_copy(
    generated: dict[str, Any] | None,
    context: dict[str, Any],
    run_id: str,
) -> dict[str, Any]:
    base = _build_copy_skeleton(context, run_id)
    generated = generated or {}
    ads_generated = generated.get("ads") if isinstance(generated.get("ads"), list) else None
    # Normalise flat responses: single ad wrapped in "ad" key, direct flat object, or array
    if ads_generated is None:
        single = generated.get("ad") if isinstance(generated.get("ad"), dict) else None
        if single:
            ads_generated = [single]
        elif isinstance(generated.get("format"), str) and generated.get("format").strip():
            ads_generated = [generated]
        elif isinstance(generated, list):
            ads_generated = generated
        else:
            ads_generated = []
    candidates = ads_generated if isinstance(ads_generated, list) else []
    for cand in candidates:
        if isinstance(cand, dict):
            for lang_copy in (cand.get("copy") or {}).values():
                if isinstance(lang_copy, dict) and "subheadline" in lang_copy:
                    lang_copy["support_line"] = lang_copy.pop("subheadline")

    used_indices: set[int] = set()

    def pick_candidate(fmt: str, persona_num: int, persona_name: str) -> dict[str, Any] | None:
        for idx, cand in enumerate(candidates):
            if idx in used_indices or not isinstance(cand, dict):
                continue
            cand_fmt = _clean_str(cand.get("format")).upper()
            if cand_fmt != fmt:
                continue
            cand_num = _persona_number_from_candidate(cand)
            cand_name = _persona_name_from_candidate(cand)
            if cand_num == persona_num or (cand_name and cand_name.lower() == persona_name.lower()):
                used_indices.add(idx)
                return cand

        for idx, cand in enumerate(candidates):
            if idx in used_indices or not isinstance(cand, dict):
                continue
            cand_fmt = _clean_str(cand.get("format")).upper()
            if cand_fmt == fmt:
                used_indices.add(idx)
                return cand
        return None

    for ad in base.get("ads", []):
        fmt = _clean_str(ad.get("format")).upper()
        persona = ad.get("persona") or {}
        persona_num = int(persona.get("number"))
        persona_name = _clean_str(persona.get("name"))
        candidate = pick_candidate(fmt, persona_num, persona_name)
        if not candidate:
            continue

        hypothesis = candidate.get("hypothesis") if isinstance(candidate.get("hypothesis"), dict) else None
        if hypothesis:
            ad["hypothesis"] = hypothesis

        angle = _clean_str(candidate.get("headline_angle"))
        if angle:
            ad["headline_angle"] = angle

        for key in ["concept_angle"]:
            value = _clean_str(candidate.get(key) or candidate.get("image_description"))
            if value:
                ad[key] = value

        cand_copy = candidate.get("copy") if isinstance(candidate.get("copy"), dict) else {}
        # If the candidate has flat fields at top level (no "copy" key), wrap them into a copy block
        copy_level_keys = {"headline", "subheadline", "support_line", "cta", "body", "bullets", "trust_line", "call_to_action", "image_description"}
        if not cand_copy and any(k in candidate for k in copy_level_keys):
            cand_flat = {k: candidate[k] for k in copy_level_keys if k in candidate}
            for k in list(copy_level_keys):
                candidate.pop(k, None)
            cand_copy = cand_flat
        # If the LLM returned flat copy (no EN/HI/HINGLISH wrapper), treat it as EN
        # and propagate to HI/HINGLISH as fallback so validation doesn't reject
        flat_keys = {"headline", "subheadline", "support_line", "cta", "body", "bullets", "trust_line"}
        if cand_copy and not any(k in cand_copy for k in {"EN", "HI", "HINGLISH"}):
            flat_en = dict(cand_copy)
            cand_copy = {"EN": flat_en, "HI": dict(flat_en), "HINGLISH": dict(flat_en)}
        for lang in ["EN", "HI", "HINGLISH"]:
            if lang not in ad["copy"]:
                continue
            base_lang = ad["copy"][lang]
            src_lang = cand_copy.get(lang) if isinstance(cand_copy.get(lang), dict) else {}

            headline = _clean_str(src_lang.get("headline"))
            cta = _clean_str(src_lang.get("cta") or src_lang.get("call_to_action"))
            if headline:
                base_lang["headline"] = shorten_copy_line(headline)
            if cta:
                base_lang["cta"] = cta

            if fmt == "TEST":
                if not _clean_str(src_lang.get("headline")):
                    base_lang["headline"] = ensure_testimonial_headline(base_lang.get("headline", ""), lang, persona)

            if fmt in {"HERO", "UGC"}:
                support = _clean_str(src_lang.get("support_line")) or _clean_str(src_lang.get("subheadline")) or _clean_str(src_lang.get("body"))
                if support:
                    base_lang["support_line"] = shorten_copy_line(support)
            elif fmt in {"BA", "FEAT"}:
                bullets = _clean_bullets(src_lang.get("bullets"))
                min_bullets = 4 if fmt == "BA" else 2
                if len(bullets) >= min_bullets:
                    if fmt == "BA":
                        bullets = [strip_ba_panel_label(b) for b in bullets]
                    base_lang["bullets"] = [shorten_copy_line(b) for b in bullets]
            else:
                trust = _clean_str(src_lang.get("trust_line") or src_lang.get("body"))
                if trust:
                    base_lang["trust_line"] = shorten_copy_line(trust)

    return base


def call_google_gemini(config: dict[str, Any], context: dict[str, Any], run_dir: Path, reserved_batch: str | None = None, language_mode: str | None = None) -> dict[str, Any] | None:
    api_key = (config.get("google_api_key") or "").strip() or os.getenv("GOOGLE_API_KEY", "").strip()
    model = (config.get("google_model") or "").strip() or DEFAULT_GOOGLE_MODEL
    if not api_key:
        print("[call_google_gemini] No Google API key configured", file=sys.stderr)
        return None

    print(f"[call_google_gemini] model={model}", file=sys.stderr)

    language_mode = resolve_language_mode(config)
    product_file = Path(str(context.get("product_file_path") or DEFAULT_PRODUCT_MASTER))
    generated_ads: list[dict[str, Any]] = []
    errors: list[str] = []
    warnings: list[str] = []

    if not product_file.exists() or not product_file.is_file():
        errors.append(f"Product master doc missing: {product_file}")
        (run_dir / "logs" / "opencode_error.txt").write_text("\n\n---\n\n".join(errors), encoding="utf-8")
        return None

    product_doc_content = product_file.read_text(encoding="utf-8")

    def call_llm(prompt: str, label: str = "ad_generation") -> tuple[dict[str, Any] | None, str, str, int]:
        sys_part = prompt.replace("SYSTEM:\n", "", 1).split("\nUSER_PAYLOAD_JSON:\n", 1)[0] if prompt.startswith("SYSTEM:\n") else prompt
        user_part = ""
        if "\nUSER_PAYLOAD_JSON:\n" in prompt:
            user_part = "USER_PAYLOAD_JSON:\n" + prompt.split("\nUSER_PAYLOAD_JSON:\n", 1)[1]

        body: dict[str, Any] = {
            "system_instruction": {"parts": [{"text": sys_part}]},
            "contents": [
                {"role": "user", "parts": [{"text": f"Product document:\n{product_doc_content}\n\n---\n\n{user_part}"}]},
            ],
            "generationConfig": {
                "temperature": 0.7,
                "maxOutputTokens": 8192,
            },
        }

        try:
            resp_raw = _gemini_generate(model, api_key, body, run_dir, label)
        except httpx.TimeoutException:
            return None, "", f"TIMEOUT after {OPENCODE_AD_TIMEOUT_SECONDS}s", -1
        except httpx.HTTPError as exc:
            return None, "", f"HTTP error: {exc}", -1

        if resp_raw is None:
            return None, "", "Gemini API returned empty", -1

        if hasattr(resp_raw, "status_code") and resp_raw.status_code != 200:
            try:
                err_body = resp_raw.json()
                err = err_body.get("error", {})
                code = err.get("code", resp_raw.status_code)
                msg = err.get("message", resp_raw.text[:500])
                status = err.get("status", "")
                detail = f"HTTP {code} ({status}): {msg}" if status else f"HTTP {code}: {msg}"
            except Exception:
                detail = f"HTTP {resp_raw.status_code}: {resp_raw.text[:300]}"
            return None, "", detail, -1

        try:
            data = resp_raw.json() if hasattr(resp_raw, "json") else resp_raw
        except (json.JSONDecodeError, AttributeError) as exc:
            return None, str(resp_raw), f"Parse error: {exc}", -1

        candidates = data.get("candidates") if isinstance(data, dict) else None
        if not candidates:
            err = data.get("error", {}).get("message", str(data))
            return None, json.dumps(data), f"Gemini error: {err}", -1

        try:
            content = candidates[0]["content"]["parts"][0]["text"]
        except (KeyError, IndexError, TypeError) as exc:
            return None, json.dumps(data), f"Gemini parse error: {exc}", -1

        parsed = parse_opencode_json_output(content)
        return parsed, content, "", 0

    with _opencode_queue_slot(f"copy_session {run_dir.name}"):
        _cancel_current_run.clear()
        cancel_event_for_run(run_dir.name).clear()

        all_items = context.get("ads") or []
        if not all_items:
            return {"ads": [], "default_aspect_ratio": "4:5"}

        payload = build_generation_payload_for_llm(context)

        formats_in_batch = sorted({str(a.get("format", "")).strip().upper() for a in all_items if isinstance(a, dict)})
        single_format = formats_in_batch[0] if len(formats_in_batch) == 1 else "ALL"
        skeleton_tail = build_ad_prompt_tail(single_format, formats=formats_in_batch)
        prompt = json.dumps(payload, ensure_ascii=False) + "\n\n" + skeleton_tail

        parsed, raw_content, err_msg, code = call_llm(prompt, label="ad_generation")

        if code == -1 and err_msg in ("CANCELLED",):
            pass
        elif parsed is None:
            errors.append(err_msg)
        else:
            ads_out = parsed.get("ads") or []
            for ad_idx, ad_item in enumerate(all_items):
                if ad_idx < len(ads_out):
                    generated_ads.append(ads_out[ad_idx])
                else:
                    warnings.append(f"No ad output for item {ad_idx}")

        result_payload: dict[str, Any] = {
            "ads": generated_ads,
            "default_aspect_ratio": "4:5",
        }
        if errors:
            result_payload["_opencode_failures"] = errors
        if warnings:
            result_payload["_opencode_warnings"] = warnings
        return result_payload


def _gemini_generate(model: str, api_key: str, body: dict[str, Any], run_dir: Path, label: str) -> Any:
    t0 = time.time()
    client = httpx.Client(timeout=httpx.Timeout(OPENCODE_AD_TIMEOUT_SECONDS, connect=5))
    _register_httpx_client(client)
    try:
        resp = client.post(
            f"{DEFAULT_GOOGLE_API_URL}/models/{model}:generateContent?key={api_key}",
            headers={"Content-Type": "application/json"},
            json=body,
        )
        elapsed = time.time() - t0
        if resp.status_code != 200:
            _log_llm_trace(run_dir.name, label, model, body, None, resp.status_code, elapsed, error=resp.text[:2000])
            return resp
        data = resp.json()
        _log_llm_trace(run_dir.name, label, model, body, data, resp.status_code, elapsed)
        return data
    except Exception:
        raise
    finally:
        _unregister_httpx_client()
        client.close()


def call_opencode_compatible(config: dict[str, Any], context: dict[str, Any], run_dir: Path, reserved_batch: str | None = None, language_mode: str | None = None) -> dict[str, Any] | None:
    api_url = (config.get("opencode_api_url") or "").strip() or os.getenv("OPENCODE_API_URL", "").strip() or DEFAULT_OPENCODE_API_URL
    api_key = (config.get("opencode_api_key") or "").strip() or os.getenv("OPENCODE_API_KEY", "").strip() or os.getenv("OPENCODE_SERVER_PASSWORD", "").strip()
    model = sanitize_dashboard_model((config.get("opencode_model") or "").strip(), list_opencode_models())
    config["opencode_model"] = model

    print(f"[call_opencode_compatible] api_url={api_url}, model={model}", file=sys.stderr)

    api_model = model.split("/", 1)[-1] if "/" in model else model
    language_mode = resolve_language_mode(config)
    product_file = Path(str(context.get("product_file_path") or DEFAULT_PRODUCT_MASTER))
    generated_ads: list[dict[str, Any]] = []
    errors: list[str] = []
    warnings: list[str] = []

    if not product_file.exists() or not product_file.is_file():
        errors.append(f"Product master doc missing: {product_file}")
        (run_dir / "logs" / "opencode_error.txt").write_text("\n\n---\n\n".join(errors), encoding="utf-8")
        return None

    product_doc_content = product_file.read_text(encoding="utf-8")

    def call_llm(prompt: str, label: str = "ad_generation") -> tuple[dict[str, Any] | None, str, str, int]:
        headers = {"Content-Type": "application/json"}
        if api_key:
            headers["Authorization"] = f"Bearer {api_key}"

        sys_part = prompt.replace("SYSTEM:\n", "", 1).split("\nUSER_PAYLOAD_JSON:\n", 1)[0] if prompt.startswith("SYSTEM:\n") else prompt
        user_part = ""
        if "\nUSER_PAYLOAD_JSON:\n" in prompt:
            user_part = "USER_PAYLOAD_JSON:\n" + prompt.split("\nUSER_PAYLOAD_JSON:\n", 1)[1]

        body = {
            "model": api_model,
            "messages": [
                {"role": "system", "content": sys_part},
                {"role": "user", "content": f"Product document:\n{product_doc_content}\n\n---\n\n{user_part}"},
            ],
            "response_format": {"type": "json_object"},
            "temperature": 0.7,
            "max_tokens": 8192,
        }

        if cancel_event_for_run(run_dir.name).is_set() or _cancel_current_run.is_set():
            return None, "", "CANCELLED", -1

        t0 = time.time()
        client = httpx.Client(timeout=httpx.Timeout(OPENCODE_AD_TIMEOUT_SECONDS, connect=2))
        _register_httpx_client(client)
        try:
            try:
                resp = client.post(f"{api_url}/chat/completions", headers=headers, json=body)
            except httpx.TimeoutException:
                _log_llm_trace(run_dir.name, label, model, body, None, -1, time.time() - t0, error="TIMEOUT")
                return None, "", f"TIMEOUT after {OPENCODE_AD_TIMEOUT_SECONDS}s", -1
            except httpx.HTTPError as exc:
                _log_llm_trace(run_dir.name, label, model, body, None, -1, time.time() - t0, error=str(exc))
                return None, "", f"HTTP error: {exc}", -1

            elapsed = time.time() - t0

            if resp.status_code != 200:
                _log_llm_trace(run_dir.name, label, model, body, None, resp.status_code, elapsed, error=resp.text[:2000])
                try:
                    err_body = resp.json()
                    err_msg = err_body.get("error", {}).get("message", "") or err_body.get("message", "")
                    detail = f"HTTP {resp.status_code}: {err_msg}" if err_msg else f"HTTP {resp.status_code}: {resp.text[:200]}"
                except Exception:
                    detail = f"HTTP {resp.status_code}: {resp.text[:200]}"
                return None, resp.text, detail, resp.status_code

            try:
                data = resp.json()
                content = data["choices"][0]["message"]["content"]
            except (json.JSONDecodeError, KeyError, IndexError) as exc:
                _log_llm_trace(run_dir.name, label, model, body, None, resp.status_code, elapsed, error=str(exc))
                return None, resp.text, f"Parse error: {exc}", -1
        finally:
            _unregister_httpx_client()
            client.close()

        _log_llm_trace(run_dir.name, label, model, body, data, resp.status_code, elapsed)
        parsed = parse_opencode_json_output(content)
        return parsed, content, "", 0

    with _opencode_queue_slot(f"copy_session {run_dir.name}"):
        _cancel_current_run.clear()
        cancel_event_for_run(run_dir.name).clear()

        all_items = context.get("ads") or []
        if not all_items:
            return {"ads": [], "default_aspect_ratio": "4:5"}

        payload = build_generation_payload_for_llm(context)

        # Build response skeleton instruction so the LLM knows the expected JSON format
        formats_in_batch = sorted({str(a.get("format", "")).strip().upper() for a in all_items if isinstance(a, dict)})
        single_format = formats_in_batch[0] if len(formats_in_batch) == 1 else "ALL"
        skeleton_tail = build_ad_prompt_tail(single_format, formats=formats_in_batch)
        prompt = json.dumps(payload, ensure_ascii=False) + "\n\n" + skeleton_tail

        parsed, raw_content, err_msg, code = call_llm(prompt, label="ad_generation")

        if code == -1 and err_msg in ("CANCELLED",):
            pass
        elif parsed is None:
            errors.append(err_msg)
        else:
            ads_out = parsed.get("ads") or []
            for ad_idx, ad_item in enumerate(all_items):
                if ad_idx < len(ads_out):
                    generated_ads.append(ads_out[ad_idx])
                else:
                    warnings.append(f"No ad output for item {ad_idx}")

        result_payload: dict[str, Any] = {
            "ads": generated_ads,
            "default_aspect_ratio": "4:5",
        }
        if errors:
            result_payload["_opencode_failures"] = errors
        if warnings:
            result_payload["_opencode_warnings"] = warnings
        return result_payload


def collect_run_result(run_dir: Path, batch_name: str, image_generated: bool) -> dict[str, Any]:
    output_dir = ROOT / "output" / batch_name
    prompt_files = []
    if output_dir.exists():
        for file in sorted(output_dir.glob("**/[A-Z]*_P*.txt")):
            prompt_files.append(str(file.relative_to(ROOT)))

    image_files: list[str] = []
    if image_generated:
        for generated_root in generated_image_roots():
            image_dir = generated_root / batch_name
            if not image_dir.exists():
                continue
            for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
                for file in sorted(image_dir.glob(f"**/{ext}")):
                    image_files.append(str(file.relative_to(ROOT)))

    result = {
        "run_id": run_dir.name,
        "batch": batch_name,
        "prompt_files": prompt_files,
        "image_files": image_files,
        "image_generated": image_generated,
        "updated_at": now_iso(),
    }
    (run_dir / "manifest.json").write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return result


_IMAGE_PATH_SORT_RE = re.compile(r"p(\d+)", re.IGNORECASE)


def _image_path_sort_key(rel: str):
    name = rel.rsplit("/", 1)[-1]
    m = _IMAGE_PATH_SORT_RE.search(name)
    persona = int(m.group(1)) if m else 0
    aspect = 0 if "/4_5/" in rel else 1
    return (persona, aspect, name)


def _collect_aspect_ratio_images(batch_name: str, aspect_ratio: str) -> list[str]:
    """Collect generated image paths for a specific aspect ratio.

    Searches both legacy and new unified output roots, looking under
    generated_images/{batch}/{aspect}/ for image files recursively.
    """
    aspect_folder = "4_5" if aspect_ratio == "4:5" else "9_16"
    image_files: list[str] = []
    for generated_root in generated_image_roots():
        image_dir = generated_root / batch_name / aspect_folder
        if not image_dir.exists():
            continue
        for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
            for file in image_dir.glob(f"**/{ext}"):
                rel = str(file.relative_to(ROOT)).replace("\\", "/")
                if "/debug/" in rel or "/.browser_downloads/" in rel:
                    continue
                image_files.append(rel)
    image_files.sort(key=_image_path_sort_key)
    return image_files


def scan_prompt_files_for_batch(batch_name: str) -> list[str]:
    output_dir = ROOT / "output" / batch_name
    prompt_files: list[str] = []
    if not output_dir.exists():
        return prompt_files
    # Match both the new slug format (<FMT>_<slug>_<LANG>_<angle>.txt)
    # and the legacy P{NN} format (<FMT>_P<NN>_<LANG>[_<angle>].txt).
    for file in sorted(output_dir.glob("**/*.txt")):
        name = file.name
        if not re.match(r"^(?:OUTPUT_|FINAL_)?[A-Z]+_", name):
            continue
        prompt_files.append(str(file.relative_to(ROOT)))
    return prompt_files


def scan_image_files_for_batch(batch_name: str) -> list[str]:
    image_files: list[str] = []
    seen: set[str] = set()
    for generated_root in generated_image_roots():
        image_dir = generated_root / batch_name
        if not image_dir.exists():
            continue
        for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
            for file in image_dir.glob(f"**/{ext}"):
                rel = str(file.relative_to(ROOT)).replace("\\", "/")
                if "/debug/" in rel or "/.browser_downloads/" in rel or "/to_be_regenerated/" in rel:
                    continue
                if rel in seen:
                    continue
                seen.add(rel)
                image_files.append(rel)
    image_files.sort(key=_image_path_sort_key)
    return image_files


def scan_regeneration_queue_files_for_batch(batch_name: str) -> list[str]:
    queue_files: list[str] = []
    seen: set[str] = set()
    for generated_root in generated_image_roots():
        for tbr_dir in sorted(generated_root.glob(f"{batch_name}/**/to_be_regenerated")):
            if not tbr_dir.is_dir():
                continue
            for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
                for file in tbr_dir.glob(f"**/{ext}"):
                    rel = str(file.relative_to(ROOT)).replace("\\", "/")
                    if rel in seen:
                        continue
                    seen.add(rel)
                    queue_files.append(rel)
    queue_files.sort(key=_image_path_sort_key)
    return queue_files


def _read_image_metadata(image_rel_path: str) -> dict[str, Any]:
    image_path = ROOT / image_rel_path
    meta_path = image_path.with_suffix(".json")
    if not meta_path.exists() or not meta_path.is_file():
        return {}
    try:
        payload = json.loads(meta_path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _prompt_stem_for_image(image_rel_path: str) -> str:
    stem = Path(image_rel_path).stem
    if stem.startswith("gemini-"):
        stem = stem.removeprefix("gemini-")
    elif stem.startswith("chatgpt-"):
        stem = stem.removeprefix("chatgpt-")
    # Strip trailing _4_5 / _9_16 aspect suffix added by the image generation scripts.
    stem = re.sub(r"_(?:4_5|9_16)$", "", stem)
    return stem.replace("-", "_").upper()


def _find_prompt_for_image(
    image_rel_path: str,
    prompt_files: list[str],
    metadata: dict[str, Any],
) -> str:
    prompt_from_name = _find_prompt_from_image_name(image_rel_path, prompt_files)
    if prompt_from_name:
        return prompt_from_name

    prompt_name = str(metadata.get("prompt_file_relative") or metadata.get("prompt_file") or "").strip().replace("\\", "/")
    if prompt_name:
        if prompt_name.startswith("output/") and prompt_name in prompt_files:
            return prompt_name
        by_name = [p for p in prompt_files if Path(p).name == Path(prompt_name).name]
        if by_name:
            if "/45/" in image_rel_path:
                return next((p for p in by_name if "/45/" in p), by_name[0])
            if "/9_16/" in image_rel_path or "/916/" in image_rel_path or "/96/" in image_rel_path:
                return next((p for p in by_name if "/916/" in p or "/96/" in p), by_name[0])
            return by_name[0]

    stem_key = _prompt_stem_for_image(image_rel_path)
    scored: list[tuple[int, str]] = []
    for prompt_file in prompt_files:
        parsed = parse_prompt_filename(prompt_file)
        if not parsed:
            continue
        fmt, lang, persona_num = parsed
        if persona_num is None:
            continue
        creative_index = parse_prompt_creative_index(prompt_file)
        tokens = [fmt.upper(), f"P{persona_num:02d}", lang.upper()]
        score = sum(1 for token in tokens if token in stem_key)
        if creative_index > 1 and f"A{creative_index:02d}" in stem_key:
            score += 1
        if "/45/" in prompt_file:
            score += 1
        if score >= 3:
            scored.append((score, prompt_file))
    if not scored:
        return ""
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored[0][1]


def _find_45_prompt_for_regeneration(prompt_file: str, prompt_files: list[str]) -> str:
    if not prompt_file:
        return ""
    if "/45/" in prompt_file:
        return prompt_file
    parsed = parse_prompt_filename(prompt_file)
    if not parsed:
        return ""
    fmt, lang, persona_num = parsed
    creative_index = parse_prompt_creative_index(prompt_file)
    for candidate in prompt_files:
        if "/45/" not in candidate:
            continue
        c_parsed = parse_prompt_filename(candidate)
        if not c_parsed:
            continue
        c_fmt, c_lang, c_persona_num = c_parsed
        if c_fmt == fmt and c_lang == lang and c_persona_num == persona_num and parse_prompt_creative_index(candidate) == creative_index:
            return candidate
    for candidate in prompt_files:
        if "/45/" in candidate and Path(candidate).name == Path(prompt_file).name:
            return candidate
    return ""


def _prompt_excerpt(prompt_file: str, max_chars: int | None = None) -> str:
    if not prompt_file:
        return ""
    prompt_path = ROOT / prompt_file
    if not prompt_path.exists() or not prompt_path.is_file():
        return ""
    text = prompt_path.read_text(encoding="utf-8", errors="ignore").strip()
    if max_chars is None or len(text) <= max_chars:
        return text
    return text[:max_chars].rstrip() + "\n..."


def _aspect_key_for_image(rel: str) -> str:
    return "9_16" if "/9_16/" in rel or "/916/" in rel or "/96/" in rel else "4_5"


def _prompt_matches_persona(prompt_file: str, fmt: str, lang: str, persona_num: int) -> bool:
    parsed = parse_prompt_filename(prompt_file)
    if not parsed:
        return False
    p_fmt, p_lang, p_persona = parsed
    return p_fmt.upper() == fmt.upper() and p_lang.upper() == lang.upper() and p_persona == persona_num


def _match_metadata_prompt(rel: str, prompt_files: list[str]) -> str:
    meta = _read_image_metadata(rel)
    prompt_name = str(meta.get("prompt_file") or "").strip()
    if not prompt_name:
        return ""
    matches = [p for p in prompt_files if Path(p).name == Path(prompt_name).name]
    if matches:
        return matches[0]
    return ""


def _find_prompt_for_group_bucket(
    rel: str,
    prompt_files: list[str],
    group_prompts: list[str],
    remaining: list[str],
) -> str:
    if not group_prompts and not remaining:
        return _match_metadata_prompt(rel, prompt_files)
    return ""


def _group_prompt_map_for_images(image_paths: list[str], prompt_files: list[str]) -> dict[str, dict[str, str]]:
    """Map generated images to prompts.

    Precedence:
      1. Sidecar metadata (the .json written alongside each image) stores the
         exact prompt_file name — use it when the file exists in prompt_files.
      2. Group-based heuristic for older images without metadata — assign within
         each (fmt, lang, persona, aspect) group so extra images don't shift later groups.
         Within a group, images and prompts are both sorted ascending by A-index.
    """
    direct: dict[str, dict[str, str]] = {}
    groups: dict[tuple[str, str, int, str], list[tuple[int, str]]] = {}
    out: dict[str, dict[str, str]] = {}

    for rel in image_paths:
        metadata_match = _match_metadata_prompt(rel, prompt_files)
        if metadata_match:
            direct[rel] = {"prompt_file": metadata_match, "mapping_status": ""}
            continue
        parsed = _parse_generated_image_name(rel)
        if not parsed:
            continue
        fmt = str(parsed.get("format") or "")
        lang = str(parsed.get("language") or "")
        persona_num = parsed.get("persona_number")
        image_index = parsed.get("image_index")
        if not fmt or not lang or not isinstance(persona_num, int):
            continue
        sort_index = int(image_index) if isinstance(image_index, int) else 0
        groups.setdefault((fmt, lang, persona_num, _aspect_key_for_image(rel)), []).append((sort_index, rel))

    for (fmt, lang, persona_num, _aspect), images in groups.items():
        prompts = sorted(
            [p for p in prompt_files if "/45/" in p and _prompt_matches_persona(p, fmt, lang, persona_num)],
            key=lambda p: (parse_prompt_creative_index(p), p),
        )
        if not prompts:
            prompts = sorted(
                [p for p in prompt_files if _prompt_matches_persona(p, fmt, lang, persona_num)],
                key=lambda p: (parse_prompt_creative_index(p), p),
            )
        if not prompts:
            continue

        images_sorted = [rel for _idx, rel in sorted(images, key=lambda item: (item[0], item[1]))]
        if len(images_sorted) > len(prompts):
            extras = images_sorted[: len(images_sorted) - len(prompts)]
            for rel in extras:
                out[rel] = {
                    "prompt_file": "",
                    "mapping_status": f"extra image: {len(images_sorted)} images for {len(prompts)} prompts in this persona",
                }
            images_sorted = images_sorted[-len(prompts):]

        for rel, prompt_file in zip(images_sorted, prompts):
            out[rel] = {"prompt_file": prompt_file, "mapping_status": ""}

    return {**direct, **out}


def _build_image_item(
    rel: str,
    prompt_files: list[str],
    *,
    is_queued: bool = False,
    prompt_file_override: str | None = None,
    mapping_status: str = "",
) -> dict[str, Any]:
    metadata = _read_image_metadata(rel)
    prompt_file = prompt_file_override if prompt_file_override is not None else _find_prompt_for_image(rel, prompt_files, metadata)
    regen_prompt_file = _find_45_prompt_for_regeneration(prompt_file, prompt_files) if not is_queued else prompt_file
    aspect = "9:16" if "/9_16/" in rel or "/916/" in rel or "/96/" in rel else "4:5"
    display_name = Path(rel).name
    if prompt_file:
        display_name = f"{Path(prompt_file).stem}{Path(rel).suffix}"
    return {
        "path": rel,
        "display_name": display_name,
        "aspect_ratio": aspect,
        "prompt_file": prompt_file,
        "regenerate_prompt_file": regen_prompt_file,
        "prompt_url": ("/output/" + prompt_file.replace("output/", "")) if prompt_file else "",
        "prompt_excerpt": _prompt_excerpt(prompt_file),
        "is_queued": is_queued,
        "mapping_status": mapping_status,
        "metadata": {
            "format": metadata.get("format", ""),
            "persona": metadata.get("persona", ""),
            "language": metadata.get("language", ""),
            "job_key": metadata.get("job_key", ""),
            "status": metadata.get("status", ""),
            "regenerated": bool(metadata.get("regenerated")) or (is_queued and "/to_be_regenerated/generated images/" in rel),
            "regenerated_at": metadata.get("regenerated_at", ""),
        },
    }


def _mark_image_metadata_regenerated(meta_path: Path, image_path: Path) -> None:
    if not meta_path.exists() or not meta_path.is_file():
        return
    try:
        payload = json.loads(meta_path.read_text(encoding="utf-8"))
    except Exception:
        return
    if not isinstance(payload, dict):
        return
    payload["regenerated"] = True
    payload["regenerated_at"] = now_iso()
    payload["regeneration_status"] = "pending_review"
    payload["saved_file"] = str(image_path)
    meta_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _image_sort_key(item: dict[str, Any]) -> tuple:
    """Sort key: aspect (4:5 before 9:16), persona number, creative index."""
    aspect = 0 if item.get("aspect_ratio") == "4:5" else 1
    prompt_file = item.get("prompt_file") or ""
    pf = parse_prompt_filename(prompt_file) if prompt_file else None
    persona = pf[2] if pf and pf[2] is not None else 999
    creative = parse_prompt_creative_index(prompt_file) if prompt_file else 999
    return (aspect, persona, creative)


def build_image_items_for_manifest(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    prompt_files = [str(path).replace("\\", "/") for path in (manifest.get("prompt_files") or [])]
    image_paths = [str(rel_raw).replace("\\", "/") for rel_raw in manifest.get("image_files") or []]
    prompt_map = _group_prompt_map_for_images(image_paths, prompt_files)
    image_items: list[dict[str, Any]] = []
    for rel in image_paths:
        mapped = prompt_map.get(rel, {})
        image_items.append(
            _build_image_item(
                rel,
                prompt_files,
                prompt_file_override=mapped.get("prompt_file") if mapped else None,
                mapping_status=str(mapped.get("mapping_status") or ""),
            )
        )
    image_items.sort(key=_image_sort_key)
    return image_items


def build_regeneration_queue_items_for_manifest(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    prompt_files = [str(path).replace("\\", "/") for path in (manifest.get("prompt_files") or [])]
    queue_paths = [str(rel_raw).replace("\\", "/") for rel_raw in manifest.get("regeneration_queue_files") or []]
    prompt_map = _group_prompt_map_for_images(queue_paths, prompt_files)
    queue_items: list[dict[str, Any]] = []
    for rel in queue_paths:
        mapped = prompt_map.get(rel, {})
        queue_items.append(
            _build_image_item(
                rel,
                prompt_files,
                is_queued=True,
                prompt_file_override=mapped.get("prompt_file") if mapped else None,
                mapping_status=str(mapped.get("mapping_status") or ""),
            )
        )
    queue_items.sort(key=_image_sort_key)
    return queue_items


def enrich_manifest_for_dashboard(manifest: dict[str, Any]) -> dict[str, Any]:
    enriched = dict(manifest)
    enriched["image_items"] = build_image_items_for_manifest(enriched)
    enriched["regeneration_queue_items"] = build_regeneration_queue_items_for_manifest(enriched)
    return enriched


def refresh_manifest_file_state(run_dir: Path, manifest: dict[str, Any]) -> dict[str, Any]:
    batch_name = str(manifest.get("batch") or "").strip()
    if not batch_name:
        return manifest

    prompt_files = scan_prompt_files_for_batch(batch_name)
    image_files = scan_image_files_for_batch(batch_name)
    regeneration_queue_files = scan_regeneration_queue_files_for_batch(batch_name)
    image_generated = bool(image_files) or bool(manifest.get("image_generated", False))
    previous_prompt_files = list(manifest.get("prompt_files") or [])
    previous_image_files = list(manifest.get("image_files") or [])
    previous_queue_files = list(manifest.get("regeneration_queue_files") or [])
    if (
        previous_prompt_files == prompt_files
        and previous_image_files == image_files
        and previous_queue_files == regeneration_queue_files
        and bool(manifest.get("image_generated", False)) == image_generated
    ):
        return manifest

    newest_mtime = 0.0
    for rel in prompt_files + image_files:
        try:
            path = ROOT / rel
            if path.exists():
                newest_mtime = max(newest_mtime, path.stat().st_mtime)
        except Exception:
            pass
    updated_at = (
        datetime.fromtimestamp(newest_mtime, tz=timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        if newest_mtime > 0
        else now_iso()
    )
    refreshed = {
        "run_id": run_dir.name,
        "batch": batch_name,
        "prompt_files": prompt_files,
        "image_files": image_files,
        "regeneration_queue_files": regeneration_queue_files,
        "image_generated": image_generated,
        "updated_at": updated_at,
    }
    merged = {**manifest, **refreshed}
    (run_dir / "manifest.json").write_text(json.dumps(merged, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return merged


def force_aspect_ratio(copy_json: dict[str, Any], aspect_ratio: str) -> dict[str, Any]:
    cloned = json.loads(json.dumps(copy_json, ensure_ascii=False))
    cloned["default_aspect_ratio"] = aspect_ratio
    ads = cloned.get("ads")
    if isinstance(ads, list):
        for ad in ads:
            if isinstance(ad, dict):
                ad["aspect_ratio"] = aspect_ratio
    return cloned


def _parse_prompt_field(prompt_text: str, label: str) -> str:
    match = re.search(rf"^\s*-\s*{re.escape(label)}:\s*(.+)$", prompt_text, flags=re.MULTILINE)
    return (match.group(1).strip() if match else "")


def parse_background_lock_from_prompt(prompt_text: str) -> tuple[str, int] | None:
    slot_match = re.search(r"^\s*-\s*Background\s+slot:\s*(BG-\d{3})\b", prompt_text, flags=re.MULTILINE | re.IGNORECASE)
    seed_match = re.search(r"^\s*-\s*Background\s+seed:\s*(\d+)\s*$", prompt_text, flags=re.MULTILINE | re.IGNORECASE)
    if not slot_match or not seed_match:
        return None
    return (slot_match.group(1).upper(), int(seed_match.group(1)))


def parse_prompt_filename(prompt_path: str) -> tuple[str, str, int | None] | None:
    """Parse a prompt file name. Returns (format, lang, persona_number) or None.

    Accepted canonical form:  <FMT>_<persona_slug>_<LANG>_<angle>[_A<NN>].txt
                              e.g. BA_always_hungry_EN_pain_point.txt,
                                   HERO_stress_snacker_HI_desired_outcome_A01.txt
    The persona_slug is looked up in persona_seeds.json to recover the persona_number.
    Also accepts forms without the concept_angle and/or variant.
    """
    name = Path(prompt_path).name
    patterns = [
        r"^(?:OUTPUT_|FINAL_)?([A-Z]+)_([a-z0-9][a-z0-9]*(?:_[a-z0-9]+)*)_(EN|HI|HINGLISH)_([a-z][a-z_]*?)(?:_(?:A|V)\d+)?\.txt$",
        r"^(?:OUTPUT_|FINAL_)?([A-Z]+)_([a-z0-9][a-z0-9]*(?:_[a-z0-9]+)*)_(EN|HI|HINGLISH)(?:_(?:A|V)\d+)?\.txt$",
        r"^(?:OUTPUT_|FINAL_)?([A-Z]+)_([a-z0-9][a-z0-9]*(?:_[a-z0-9]+)*)_(EN|HI|HINGLISH)\.txt$",
    ]
    for pat in patterns:
        m = re.match(pat, name, re.IGNORECASE)
        if m:
            slug = m.group(2).lower()
            pn = persona_number_from_slug(slug)
            return (m.group(1).upper(), m.group(3).upper(), pn)
    return None


def parse_prompt_filename_full(prompt_path: str) -> tuple[str, str, int, str, str] | None:
    """Like ``parse_prompt_filename`` but also extracts the concept_angle and variant.

    Returns ``(format, lang, persona_number, concept_angle, variant)`` or ``None``.
    ``concept_angle`` defaults to ``""`` if the filename has no angle component.
    ``variant`` is the ``A01``/``V01`` string, or ``""`` if absent.
    The persona_number is recovered from the persona slug via persona_seeds.json.
    """
    name = Path(prompt_path).name
    patterns = [
        # canonical: <FMT>_<slug>_<LANG>_<angle>[_A<NN>].txt
        r"^(?:OUTPUT_|FINAL_)?(?P<fmt>[A-Z]+)_(?P<slug>[a-z0-9][a-z0-9]*(?:_[a-z0-9]+)*)_(?P<lang>EN|HI|HINGLISH)_(?P<angle>[a-z][a-z_]*?)(?:_(?P<variant>A\d+|V\d+))?\.txt$",
        r"^(?:OUTPUT_|FINAL_)?(?P<fmt>[A-Z]+)_(?P<slug>[a-z0-9][a-z0-9]*(?:_[a-z0-9]+)*)_(?P<lang>EN|HI|HINGLISH)(?:_(?P<variant>A\d+|V\d+))?\.txt$",
    ]
    for pat in patterns:
        m = re.match(pat, name, re.IGNORECASE)
        if m:
            pn = persona_number_from_slug(m.group("slug").lower())
            if pn is None:
                return None
            return (
                m.group("fmt").upper(),
                m.group("lang").upper(),
                pn,
                m.groupdict().get("angle", "") or "",
                m.groupdict().get("variant", "") or "",
            )
    return None


def parse_prompt_creative_index(prompt_path: str) -> int:
    match = re.search(r"_A(\d+)\.txt$", Path(prompt_path).name, flags=re.IGNORECASE)
    return int(match.group(1)) if match else 1


def _extract_persona_slug_from_prompt_filename(prompt_path: str) -> str:
    """Extract the persona slug from a prompt filename (e.g. 'always_hungry' from
    'BA_always_hungry_EN_pain_point.txt'). Returns '' if no slug is found.
    For legacy 'BA_P03_EN_pain_point.txt' prompts, returns '' (caller should
    fall back to P<NN>)."""
    if not prompt_path:
        return ""
    name = Path(prompt_path).name
    patterns = [
        r"^(?:OUTPUT_|FINAL_)?[A-Z]+_(?P<slug>[a-z0-9][a-z0-9]*(?:_[a-z0-9]+)*)_(?:EN|HI|HINGLISH)_",
        r"^(?:OUTPUT_|FINAL_)?[A-Z]+_(?P<slug>[a-z0-9][a-z0-9]*(?:_[a-z0-9]+)*)_(?:EN|HI|HINGLISH)(?:\.txt|$)",
    ]
    for pat in patterns:
        m = re.match(pat, name, re.IGNORECASE)
        if m:
            return m.group("slug").lower()
    return ""


def _parse_generated_image_name(image_rel_path: str) -> dict[str, Any]:
    """Parse a generated image filename. Returns dict with format, persona_number,
    language, concept_angle, image_index, aspect; missing keys mean the stem is unparseable.

    Canonical form: ``<FMT>_<persona_slug>_<LANG>_<angle>[_A<NN>][_<aspect>].<ext>``
                    e.g. BA_always_hungry_EN_pain_point_4_5.png,
                         HERO_stress_snacker_HI_desired_outcome_A01_9_16.jpg
    The optional trailing ``_4_5`` / ``_9_16`` marks the aspect ratio (added by
    the chatgpt/gemini scripts).  The persona_number is recovered from the persona
    slug via persona_seeds.json.
    """
    stem = Path(image_rel_path).stem
    # Strip a trailing _4_5 / _9_16 aspect suffix so the rest of the regex can
    # match the canonical prompt-derived stem.  This also lets us report the
    # detected aspect back to callers.
    detected_aspect = ""
    aspect_match = re.search(r"_(?P<aspect>4_5|9_16)$", stem)
    if aspect_match:
        detected_aspect = aspect_match.group("aspect")
        stem = stem[: aspect_match.start()]
    patterns = [
        # canonical: <FMT>_<slug>_<LANG>_<angle>[_A<NN>]
        r"^(?P<fmt>[A-Z]+)_(?P<slug>[a-z0-9][a-z0-9]*(?:_[a-z0-9]+)*)_(?P<lang>EN|HI|HINGLISH)_(?P<angle>[a-z][a-z_]*?)(?:_A(?P<image_index>\d+))?$",
        # angle-less: <FMT>_<slug>_<LANG>[_A<NN>]
        r"^(?P<fmt>[A-Z]+)_(?P<slug>[a-z0-9][a-z0-9]*(?:_[a-z0-9]+)*)_(?P<lang>EN|HI|HINGLISH)(?:_A(?P<image_index>\d+))?$",
    ]
    for pat in patterns:
        m = re.search(pat, stem, flags=re.IGNORECASE)
        if not m:
            continue
        pn = persona_number_from_slug(m.group("slug").lower())
        if pn is None:
            continue
        return {
            "format": m.group("fmt").upper(),
            "persona_number": int(pn),
            "language": m.group("lang").upper(),
            "concept_angle": m.groupdict().get("angle", "") or "",
            "image_index": int(m.group("image_index")) if m.groupdict().get("image_index") else None,
            "aspect": detected_aspect or "",
        }
    return {}


def _sorted_prompt_candidates(
    prompt_files: list[str],
    *,
    fmt: str,
    lang: str,
    prefer_45: bool,
) -> list[str]:
    candidates: list[str] = []
    for prompt_file in prompt_files:
        parsed = parse_prompt_filename(prompt_file)
        if not parsed:
            continue
        p_fmt, p_lang, persona_num = parsed
        if persona_num is None:
            continue
        if p_fmt.upper() != fmt.upper() or p_lang.upper() != lang.upper():
            continue
        if prefer_45 and "/45/" not in prompt_file:
            continue
        candidates.append(prompt_file)
    return sorted(
        candidates,
        key=lambda prompt_file: (
            parse_prompt_filename(prompt_file)[2] or 0,
            parse_prompt_creative_index(prompt_file),
            prompt_file,
        ),
    )


def _find_prompt_from_image_name(image_rel_path: str, prompt_files: list[str]) -> str:
    parsed_image = _parse_generated_image_name(image_rel_path)
    if not parsed_image:
        return ""

    fmt = str(parsed_image.get("format") or "")
    lang = str(parsed_image.get("language") or "")
    persona_num = parsed_image.get("persona_number")
    image_index = parsed_image.get("image_index")
    prefer_45 = True

    candidates = _sorted_prompt_candidates(prompt_files, fmt=fmt, lang=lang, prefer_45=prefer_45)
    if not candidates and prefer_45:
        candidates = _sorted_prompt_candidates(prompt_files, fmt=fmt, lang=lang, prefer_45=False)

    if isinstance(image_index, int):
        # Current/future naming: image a02 should map to persona Pxx prompt A02.
        for prompt_file in candidates:
            parsed_prompt = parse_prompt_filename(prompt_file)
            if not parsed_prompt:
                continue
            _fmt, _lang, prompt_persona = parsed_prompt
            if prompt_persona == persona_num and parse_prompt_creative_index(prompt_file) == image_index:
                return prompt_file

        # Older generated images used a global sequence in filenames: P02 a04
        # means the 4th prompt in sorted FORMAT/LANG order, not prompt A04.
        if 1 <= image_index <= len(candidates):
            global_prompt = candidates[image_index - 1]
            parsed_global = parse_prompt_filename(global_prompt)
            if parsed_global and parsed_global[2] == persona_num:
                return global_prompt

    persona_candidates = [
        prompt_file
        for prompt_file in candidates
        if (parse_prompt_filename(prompt_file) or (None, None, None))[2] == persona_num
    ]
    if len(persona_candidates) == 1:
        return persona_candidates[0]
    return ""


def parse_persona_number_from_prompt(prompt_text: str) -> int | None:
    match = re.search(r"\(\s*Persona\s*(\d+)\s*\)", prompt_text, flags=re.IGNORECASE)
    if not match:
        return None
    return int(match.group(1))


EXACT_COPY_BLOCK_RE = re.compile(
    r"EXACT ON-IMAGE COPY - DO NOT ALTER ANYTHING\s*\n(?P<block>.+?)\n\s*Render every character exactly as written",
    flags=re.DOTALL,
)


def extract_on_image_copy_lines(prompt_text: str) -> list[dict[str, str]]:
    """
    Legacy-ish extractor used by the dashboard editor.

    It DOES NOT preserve exact spacing/linebreaks; it trims lines into {label,value}.
    Keep this for backward compatibility.
    """
    block = EXACT_COPY_BLOCK_RE.search(prompt_text)
    if not block:
        return []

    out: list[dict[str, str]] = []
    for line in block.group("block").splitlines():
        raw = line.strip()
        if not raw:
            continue
        parsed = re.match(r"^-\s*([^:]+):\s*(.*)$", raw)
        if not parsed:
            continue
        out.append({"label": parsed.group(1).strip(), "value": parsed.group(2).strip()})
    return out


def extract_exact_on_image_copy_block(prompt_text: str, *, warn_log_path: Path | None = None) -> str | None:
    """
    Task 5: Extract ONLY the content inside:
      EXACT ON-IMAGE COPY - DO NOT ALTER ANYTHING
      ...
      Render every character exactly as written

    Rules:
    - preserve exact text including punctuation/case/spacing/line breaks
    - no normalization (no strip, no join)
    - if block missing: optionally log warning; return None
    """
    pattern = (
        r"EXACT ON-IMAGE COPY - DO NOT ALTER ANYTHING\s*\n"
        r"(?P<block>.+?)\n\s*Render every character exactly as written"
    )
    m = re.search(pattern, prompt_text, flags=re.DOTALL)
    if not m:
        if warn_log_path is not None:
            warn_log_path.parent.mkdir(parents=True, exist_ok=True)
            warn_log_path.write_text(
                "WARNING: EXACT ON-IMAGE COPY block missing; skipping this prompt.\n",
                encoding="utf-8",
            )
        return None

    # Return exactly what was captured: no strip().
    return m.group("block")


def load_run_language_mode(run_dir: Path) -> str:
    run_context_path = run_dir / "context" / "run_context.json"
    assembler_mode = "BOTH"
    if not run_context_path.exists():
        return assembler_mode
    try:
        run_context = json.loads(run_context_path.read_text(encoding="utf-8"))
        lang_mode = str(run_context.get("language_mode") or "ALL").upper()
        if lang_mode == "EN":
            return "EN"
        if lang_mode == "HI":
            return "HI"
    except Exception:
        return assembler_mode
    return assembler_mode


def rerender_prompts_for_run(run_dir: Path, batch: str, copy_file: Path, language_mode: str) -> None:
    result = run_cmd(
        [
            "python3",
            "scripts/generate_ads.py",
            "--copy-file",
            str(copy_file),
            "--batch",
            batch,
            "--language-mode",
            language_mode,
        ],
        cwd=ROOT,
        run_id=run_dir.name,
    )
    if result.returncode != 0:
        if cancel_event_for_run(run_dir.name).is_set() or _cancel_current_run.is_set():
            raise HTTPException(status_code=499, detail="Cancelled by user")
        error_text = result.stderr or result.stdout
        (run_dir / "logs" / "assembler_edit_error.txt").write_text(error_text, encoding="utf-8")
        short_error = "\n".join([line for line in error_text.splitlines() if line.strip()][-12:])
        raise HTTPException(status_code=500, detail=f"Prompt regeneration failed: {short_error}")


def merge_manifest(run_dir: Path, previous_manifest: dict[str, Any], refreshed: dict[str, Any]) -> dict[str, Any]:
    merged = {**previous_manifest, **refreshed}
    (run_dir / "manifest.json").write_text(json.dumps(merged, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return merged


def generate_916_for_run(run_dir: Path, manifest: dict[str, Any]) -> dict[str, Any]:
    copy_path = run_dir / "context" / "copy_batch.json"
    if not copy_path.exists():
        raise HTTPException(status_code=404, detail="copy_batch.json not found for run")

    batch = (manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    copy_json = json.loads(copy_path.read_text(encoding="utf-8"))
    copy_916 = force_aspect_ratio(copy_json, "9:16")
    visual_locks = collect_45_visual_locks(batch)
    if visual_locks:
        copy_916 = apply_visual_locks(copy_916, visual_locks)
    copy_916_path = run_dir / "context" / "copy_batch_916.json"
    copy_916_path.write_text(json.dumps(copy_916, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    assembler_mode = load_run_language_mode(run_dir)
    result = run_cmd(
        [
            "python3",
            "scripts/generate_ads.py",
            "--copy-file",
            str(copy_916_path),
            "--batch",
            batch,
            "--language-mode",
            assembler_mode,
            "--no-registry-write",
            "--skip-uniqueness-check",
        ],
        cwd=ROOT,
        run_id=run_dir.name,
    )

    if result.returncode != 0:
        if cancel_event_for_run(run_dir.name).is_set() or _cancel_current_run.is_set():
            raise HTTPException(status_code=499, detail="Cancelled by user")
        error_text = result.stderr or result.stdout
        (run_dir / "logs" / "assembler_916_error.txt").write_text(error_text, encoding="utf-8")
        short_error = "\n".join([line for line in error_text.splitlines() if line.strip()][-12:])
        raise HTTPException(status_code=500, detail=f"9:16 generation failed: {short_error}")

    refreshed = collect_run_result(run_dir, batch, bool(manifest.get("image_generated", False)))
    refreshed["generated_variant"] = "9:16"
    return merge_manifest(run_dir, manifest, refreshed)


def collect_45_visual_locks(batch: str) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    ratio_dir = ROOT / "output" / batch / "45"
    if not ratio_dir.exists():
        return out
    for prompt_file in sorted(ratio_dir.glob("*_EN.txt")) + sorted(ratio_dir.glob("*_HI.txt")):
        parsed = parse_prompt_filename(prompt_file.name)
        if not parsed:
            continue
        fmt, _lang, persona_number = parsed
        key = f"{fmt}::P{persona_number}" if isinstance(persona_number, int) else fmt
        current = out.get(key, {})
        text = prompt_file.read_text(encoding="utf-8", errors="ignore")
        if persona_number is None:
            inferred = parse_persona_number_from_prompt(text)
            if isinstance(inferred, int):
                persona_number = inferred
                key = f"{fmt}::P{persona_number}"
                current = out.get(key, current)
        lock = parse_background_lock_from_prompt(text)
        if lock:
            current["background_slot"] = lock[0]
            current["background_seed"] = lock[1]

        visual_lock = {
            "seeded_background_direction": _parse_prompt_field(text, "Seeded background direction (single sentence, exact)"),
            "subject": _parse_prompt_field(text, "Subject"),
            "action": _parse_prompt_field(text, "Action"),
            "camera": _parse_prompt_field(text, "Camera"),
            "lighting": _parse_prompt_field(text, "Lighting"),
            "props": _parse_prompt_field(text, "Props"),
            "surfaces": _parse_prompt_field(text, "Surfaces"),
            "mood": _parse_prompt_field(text, "Mood"),
            "realism": _parse_prompt_field(text, "Realism"),
        }
        visual_lock = {k: v for k, v in visual_lock.items() if v}
        if visual_lock:
            current["visual_lock"] = visual_lock

        if current:
            out[key] = current
    return out


def apply_visual_locks(copy_json: dict[str, Any], locks: dict[str, dict[str, Any]]) -> dict[str, Any]:
    cloned = json.loads(json.dumps(copy_json, ensure_ascii=False))
    ads = cloned.get("ads")
    if not isinstance(ads, list):
        return cloned
    for ad in ads:
        if not isinstance(ad, dict):
            continue
        fmt = str(ad.get("format") or "").strip().upper()
        persona_no = None
        persona = ad.get("persona")
        if isinstance(persona, dict):
            raw_no = persona.get("number")
            if isinstance(raw_no, int):
                persona_no = raw_no
        lock_key = f"{fmt}::P{persona_no}" if isinstance(persona_no, int) else ""
        lock = (locks.get(lock_key) if lock_key else None) or locks.get(fmt) or {}
        if not lock:
            continue
        if isinstance(lock.get("background_slot"), str):
            ad["background_slot"] = lock["background_slot"]
        if isinstance(lock.get("background_seed"), int):
            ad["background_seed"] = lock["background_seed"]
        if isinstance(lock.get("visual_lock"), dict):
            ad["visual_lock"] = lock["visual_lock"]
    return cloned


def _background_reuse_keys(fmt: str, persona_no: int | None, visual_archetype: str, share_across_personas: bool) -> list[str]:
    fmt = fmt.strip().upper()
    persona = f"P{persona_no:02d}" if isinstance(persona_no, int) else ""
    arch = visual_archetype.strip()
    if share_across_personas:
        return [key for key in [f"{fmt}::{arch}" if arch else "", fmt] if key]
    return [key for key in [f"{fmt}::{persona}::{arch}" if persona and arch else "", f"{fmt}::{persona}" if persona else ""] if key]


def collect_background_reuse_locks(source_run_id: str) -> dict[str, dict[str, Any]]:
    source_run_id = str(source_run_id or "").strip()
    if not source_run_id:
        return {}
    _source_dir, manifest, _has_storage_manifest = load_manifest_for_run(source_run_id)
    locks: dict[str, dict[str, Any]] = {}
    for rel_path in manifest.get("prompt_files") or []:
        rel = str(rel_path).replace("\\", "/")
        if "/916/" in rel or "/96/" in rel:
            continue
        parsed = parse_prompt_filename(rel)
        if not parsed:
            continue
        fmt, _lang, persona_no = parsed
        prompt_path = ROOT / rel
        if not prompt_path.exists():
            continue

        slot = ""
        seed: int | None = None
        visual_archetype = ""
        sidecar = prompt_path.with_suffix(".json")
        if sidecar.exists():
            try:
                meta = json.loads(sidecar.read_text(encoding="utf-8"))
            except Exception:
                meta = {}
            bg = meta.get("background") if isinstance(meta.get("background"), dict) else {}
            slot = str(bg.get("slot") or "").strip()
            raw_seed = bg.get("seed")
            if isinstance(raw_seed, int):
                seed = raw_seed
            visual = meta.get("visual_archetype") if isinstance(meta.get("visual_archetype"), dict) else {}
            visual_archetype = str(visual.get("id") or "").strip()

        if not slot or not isinstance(seed, int):
            text = prompt_path.read_text(encoding="utf-8", errors="ignore")
            lock = parse_background_lock_from_prompt(text)
            if lock:
                slot, seed = lock
            if not visual_archetype:
                visual_archetype = _parse_prompt_field(text, "Selected visual archetype").split(" - ", 1)[0].strip()

        if not slot or not isinstance(seed, int):
            continue

        lock_payload = {
            "background_slot": slot,
            "background_seed": seed,
            "background_reused_from_run_id": source_run_id,
        }
        for key in _background_reuse_keys(fmt, persona_no, visual_archetype, False):
            locks.setdefault(key, lock_payload)
        for key in _background_reuse_keys(fmt, persona_no, visual_archetype, True):
            locks.setdefault(key, lock_payload)
    return locks


def apply_background_reuse_locks(
    copy_json: dict[str, Any],
    locks: dict[str, dict[str, Any]],
    *,
    share_across_personas: bool,
) -> tuple[dict[str, Any], int]:
    cloned = json.loads(json.dumps(copy_json, ensure_ascii=False))
    ads = cloned.get("ads")
    if not isinstance(ads, list) or not locks:
        return cloned, 0
    applied = 0
    for ad in ads:
        if not isinstance(ad, dict):
            continue
        fmt = str(ad.get("format") or "").strip().upper()
        persona_no = None
        persona = ad.get("persona")
        if isinstance(persona, dict) and isinstance(persona.get("number"), int):
            persona_no = int(persona["number"])
        visual_archetype = str(ad.get("visual_archetype") or "").strip()
        lock = None
        reuse_key = ""
        for key in _background_reuse_keys(fmt, persona_no, visual_archetype, share_across_personas):
            if key in locks:
                lock = locks[key]
                reuse_key = key
                break
        if not lock:
            continue
        ad["background_slot"] = lock["background_slot"]
        ad["background_seed"] = lock["background_seed"]
        ad["background_reused_from_run_id"] = lock.get("background_reused_from_run_id", "")
        ad["background_reuse_key"] = reuse_key
        applied += 1
    return cloned, applied


def collect_visual_pattern_reuse_locks(source_run_id: str) -> dict[str, dict[str, Any]]:
    source_run_id = str(source_run_id or "").strip()
    if not source_run_id:
        return {}
    _source_dir, manifest, _has_storage_manifest = load_manifest_for_run(source_run_id)
    locks: dict[str, dict[str, Any]] = {}
    for rel_path in manifest.get("prompt_files") or []:
        rel = str(rel_path).replace("\\", "/")
        if "/916/" in rel or "/96/" in rel:
            continue
        parsed = parse_prompt_filename(rel)
        if not parsed:
            continue
        fmt, _lang, persona_no = parsed
        prompt_path = ROOT / rel
        if not prompt_path.exists():
            continue

        visual_archetype = ""
        sidecar = prompt_path.with_suffix(".json")
        if sidecar.exists():
            try:
                meta = json.loads(sidecar.read_text(encoding="utf-8"))
            except Exception:
                meta = {}
            visual = meta.get("visual_archetype") if isinstance(meta.get("visual_archetype"), dict) else {}
            visual_archetype = str(visual.get("id") or "").strip()

        if not visual_archetype:
            text = prompt_path.read_text(encoding="utf-8", errors="ignore")
            visual_archetype = _parse_prompt_field(text, "Selected visual archetype").split(" - ", 1)[0].strip()

        if not visual_archetype:
            continue

        lock_payload = {
            "visual_archetype": visual_archetype,
            "visual_pattern_reused_from_run_id": source_run_id,
        }
        for key in _background_reuse_keys(fmt, persona_no, visual_archetype, False):
            locks.setdefault(key, lock_payload)
        for key in _background_reuse_keys(fmt, persona_no, visual_archetype, True):
            locks.setdefault(key, lock_payload)
    return locks


def apply_visual_pattern_reuse_to_plan(
    plan: list[dict[str, Any]],
    locks: dict[str, dict[str, Any]],
    *,
    share_across_personas: bool,
) -> tuple[list[dict[str, Any]], int]:
    if not locks:
        return plan, 0
    out: list[dict[str, Any]] = []
    applied = 0
    for item in plan:
        entry = dict(item)
        fmt = str(entry.get("format") or "").strip().upper()
        persona_no = int(entry.get("persona")) if entry.get("persona") is not None else None
        lock = None
        reuse_key = ""
        keys = []
        if share_across_personas:
            keys.append(fmt)
        else:
            keys.append(f"{fmt}::P{persona_no:02d}" if isinstance(persona_no, int) else fmt)
        for key in keys:
            if key in locks:
                lock = locks[key]
                reuse_key = key
                break
        if lock:
            entry["visual_archetype"] = lock["visual_archetype"]
            entry["visual_pattern_reused_from_run_id"] = lock.get("visual_pattern_reused_from_run_id", "")
            entry["visual_pattern_reuse_key"] = reuse_key
            applied += 1
        out.append(entry)
    return out, applied


def resolve_format_plan(config: dict[str, Any]) -> list[dict[str, Any]]:
    personas = config.get("selected_personas") or []
    if not personas:
        raise RuntimeError("selected_personas is required")

    all_formats = [fmt for fmt in (config.get("global_formats") or []) if fmt in FORMATS]
    format_map = config.get("formats_by_persona") or {}
    archetype_map = config.get("visual_archetypes_by_format") or {}
    share_bg_across_personas = bool(config.get("share_background_across_personas"))
    try:
        multiplier = max(1, min(20, int(config.get("multiplier") or 1)))
    except (TypeError, ValueError):
        multiplier = 1

    out: list[dict[str, Any]] = []
    for raw_persona in personas:
        persona_num = int(raw_persona)
        per_formats = [fmt for fmt in (format_map.get(str(persona_num)) or format_map.get(persona_num) or []) if fmt in FORMATS]
        formats = per_formats if per_formats else all_formats
        if not formats:
            formats = ["HERO"]
        for fmt in formats:
            forced_archetype = str(archetype_map.get(fmt) or "").strip()
            background_group_key = fmt if share_bg_across_personas else f"{fmt}::P{persona_num:02d}"
            for creative_index in range(1, multiplier + 1):
                item = {
                    "persona": persona_num,
                    "format": fmt,
                    "creative_index": creative_index,
                    "creative_total": multiplier,
                    "background_group_key": background_group_key,
                    "share_background_across_personas": share_bg_across_personas,
                }
                if forced_archetype:
                    item["visual_archetype"] = forced_archetype
                out.append(item)
    return out


def expand_plan_with_hypothesis(plan: list[dict[str, Any]], hypothesis_cfg: dict[str, Any]) -> list[dict[str, Any]]:
    """Expand ad plan to include hypothesis style.

    When a hypothesis is active, generates ads using that specific style/variant.
    """
    hyp_type = str(hypothesis_cfg.get("type") or "none").strip().lower()
    if hyp_type == "none" or hyp_type not in HYPOTHESIS_VARIABLES:
        return plan

    variable_def = HYPOTHESIS_VARIABLES[hyp_type]
    selected_variant = str(hypothesis_cfg.get("variant") or "").strip()
    available_options = [opt["id"] for opt in variable_def.get("options", [])]

    if not available_options:
        return plan

    # Use the selected variant if valid, otherwise use first available
    variant_to_use = selected_variant if selected_variant in available_options else available_options[0]

    out: list[dict[str, Any]] = []
    for item in plan:
        entry = dict(item)
        entry["hypothesis"] = {
            "type": hyp_type,
            "variable_label": variable_def["label"],
            "variant": variant_to_use,
            "hypothesis_id": f"{hyp_type}-{variant_to_use}",
        }
        base_group_key = str(entry.get("background_group_key") or f"{entry.get('format')}::P{int(entry.get('persona')):02d}")
        entry["background_group_key"] = f"{base_group_key}::{hyp_type}::{variant_to_use}"
        out.append(entry)
    return out


app = FastAPI(title="Ad Dashboard API", version="1.0.0")

# Load settings-based CORS
from dashboard.backend.db.settings import settings as app_settings, validate_production_settings

app.add_middleware(
    CORSMiddleware,
    allow_origins=app_settings.cors_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
)


# ── Auth middleware (protects old routes without modifying them) ────────────
from dashboard.backend.auth.service import get_current_user_from_cookie

PUBLIC_API_PREFIXES = ("/api/auth/",)


@app.middleware("http")
async def auth_middleware(request: Request, call_next) -> Response:
    if app_settings.is_production:
        path = request.url.path
        if path.startswith("/api/") and not path.startswith(PUBLIC_API_PREFIXES):
            session_token = request.cookies.get("session")
            user = get_current_user_from_cookie(session_token)
            if user is None:
                return JSONResponse({"detail": "Not authenticated"}, status_code=401)
            request.state.user = user
    response = await call_next(request)
    return response


_opencode_catalog_cache: dict[str, Any] = {}
_opencode_catalog_lock = threading.Lock()


def _build_opencode_catalog_cached():
    global _opencode_catalog_cache
    try:
        catalog = build_opencode_catalog()
    except Exception:
        catalog = {
            "api_url": DEFAULT_OPENCODE_API_URL,
            "providers": [],
            "models_by_provider": {},
            "default_model": "",
        }
    with _opencode_catalog_lock:
        _opencode_catalog_cache = catalog


def _get_opencode_catalog():
    with _opencode_catalog_lock:
        return dict(_opencode_catalog_cache)


@app.on_event("startup")
def startup() -> None:
    load_env_file(ENV_PATH)
    ensure_dirs()

    # Validate production settings (exits if critical vars missing)
    validate_production_settings()

    # Initialize MongoDB indexes (best-effort)
    try:
        from dashboard.backend.db.indexes import create_indexes
        idx_result = create_indexes()
        created = sum(v for v in idx_result.values() if v > 0)
        failed = sum(1 for v in idx_result.values() if v < 0)
        if created or failed:
            print(f"[startup] MongoDB indexes: {created} created, {failed} failed")
    except Exception as e:
        msg = str(e)
        if app_settings.is_production:
            print(f"[startup] FATAL: MongoDB connection failed in production: {msg}", file=sys.stderr)
            sys.exit(1)
        print(f"[startup] MongoDB index init skipped (dev): {e}")

    threading.Thread(target=_build_opencode_catalog_cached, daemon=True).start()


def api_defaults() -> dict[str, Any]:
    personas = parse_persona_library()
    opencode = _get_opencode_catalog()
    if not opencode.get("providers") and not opencode.get("models_by_provider"):
        try:
            opencode = build_opencode_catalog()
            with _opencode_catalog_lock:
                _opencode_catalog_cache.clear()
                _opencode_catalog_cache.update(opencode)
        except Exception:
            opencode = {
                "api_url": DEFAULT_OPENCODE_API_URL,
                "providers": [],
                "models_by_provider": {},
                "default_model": "",
            }
    return {
        "personas": personas,
        "formats": FORMATS,
        "format_patterns": load_format_visual_archetypes(),
        "image_sources": read_active_images(default_image_sources_file()),
        "input_images": list_input_images(),
        "product_doc": default_product_doc_info(),
        "default_files": {
            "product_info": str(DEFAULT_PRODUCT_MASTER.relative_to(ROOT)),

        },
        "opencode": opencode,
        "provider": {
            "current": (os.getenv("LLM_PROVIDER") or "opencode").strip().lower(),
            "google_api_key": bool(os.getenv("GOOGLE_API_KEY", "")),
            "opencode_api_url": os.getenv("OPENCODE_API_URL", "") or DEFAULT_OPENCODE_API_URL,
            "google_model": os.getenv("GOOGLE_MODEL", "") or DEFAULT_GOOGLE_MODEL,
            "google_models": api_google_models(),
        },
        "hypothesis": {
            "variables": HYPOTHESIS_VARIABLES,
            "default": {"type": "none", "variant": ""},
        },
        "batch_size": 10,
    }


def api_progress(batch_key: str) -> dict[str, Any]:
    batch_key_clean = str(batch_key).strip()
    for root in generated_image_roots():
        log_path = root / batch_key_clean / "_headless_progress.json"
        if not log_path.exists():
            continue
        lines = log_path.read_text(encoding="utf-8", errors="ignore").splitlines()
        entries = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            try:
                entries.append(json.loads(line))
            except json.JSONDecodeError:
                pass
        if not entries:
            continue
        latest = entries[-1]
        return {
            "batch_key": batch_key_clean,
            "step": latest.get("step", ""),
            "message": latest.get("message", ""),
            "time": latest.get("time", 0),
            "entries": entries,
        }
    raise HTTPException(status_code=404, detail=f"No progress found for batch: {batch_key_clean}")


def api_opencode_catalog() -> dict[str, Any]:
    try:
        catalog = build_opencode_catalog()
        with _opencode_catalog_lock:
            _opencode_catalog_cache.clear()
            _opencode_catalog_cache.update(catalog)
        return catalog
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to discover OpenCode catalog: {exc}") from exc


def api_save_provider_config(payload: dict[str, Any]) -> dict[str, Any]:
    # Read current env file
    current = {}
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" in line:
                k, v = line.split("=", 1)
                current[k.strip()] = v.strip()

    # Override with payload fields
    def merge(key, payload_key=None):
        pk = payload_key or key.lower()
        if pk in payload:
            val = str(payload[pk]).strip()
            if val:
                current[key] = val
            else:
                current.pop(key, None)
        return current.get(key, "")

    merge("LLM_PROVIDER", "provider")
    merge("GOOGLE_API_KEY")
    merge("OPENCODE_API_URL")
    merge("OPENCODE_API_KEY")
    merge("GOOGLE_MODEL")

    try:
        ENV_PATH.write_text(
            "\n".join(f"{k}={v}" for k, v in current.items() if v) + "\n",
            encoding="utf-8",
        )
        for k, v in current.items():
            if v:
                os.environ[k] = v
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Failed to save config: {exc}")

    return {"status": "ok", "provider": current.get("LLM_PROVIDER", "opencode")}


def api_google_models(api_key: str = "") -> list[str]:
    key = api_key.strip() or os.getenv("GOOGLE_API_KEY", "").strip()
    if not key:
        return ["gemini-2.0-flash", "gemini-2.5-flash", "gemini-2.5-pro"]
    url = f"{DEFAULT_GOOGLE_API_URL}/models?key={key}"
    try:
        resp = httpx.get(url, timeout=15)
        if resp.status_code != 200:
            return ["gemini-2.0-flash", "gemini-2.5-flash", "gemini-2.5-pro"]
        data = resp.json()
        raw = data.get("models") or []
        models = []
        for m in raw:
            name = m.get("name", "")
            methods = m.get("supportedGenerationMethods") or []
            if "generateContent" in methods:
                models.append(name.replace("models/", "", 1))
        return sorted(models) if models else ["gemini-2.0-flash", "gemini-2.5-flash", "gemini-2.5-pro"]
    except Exception:
        return ["gemini-2.0-flash", "gemini-2.5-flash", "gemini-2.5-pro"]


def _extract_backfill_batch(run_id: str) -> str | None:
    match = re.match(r"^batch_(v\d+)$", str(run_id or "").strip(), flags=re.IGNORECASE)
    if not match:
        return None
    return match.group(1)


def _build_backfill_manifest(run_id: str, batch: str) -> dict[str, Any]:
    prompt_files = scan_prompt_files_for_batch(batch)
    if not prompt_files:
        raise HTTPException(status_code=404, detail=f"No prompt files found in output/{batch}")
    image_files = scan_image_files_for_batch(batch)
    batch_dir = ROOT / "output" / batch
    updated_at = now_iso()
    if batch_dir.exists():
        updated_at = datetime.fromtimestamp(batch_dir.stat().st_mtime, tz=timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    regeneration_queue_files = scan_regeneration_queue_files_for_batch(batch)
    return {
        "run_id": run_id,
        "batch": batch,
        "prompt_files": prompt_files,
        "image_files": image_files,
        "regeneration_queue_files": regeneration_queue_files,
        "image_generated": bool(image_files),
        "updated_at": updated_at,
        "source": "output_backfill",
    }


def load_manifest_for_run(run_id: str) -> tuple[Path | None, dict[str, Any], bool]:
    run_dir = RUNS_ROOT / run_id
    manifest_path = run_dir / "manifest.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        refreshed = refresh_manifest_file_state(run_dir, manifest)
        return run_dir, refreshed, True

    backfill_batch = _extract_backfill_batch(run_id)
    if backfill_batch:
        return None, _build_backfill_manifest(run_id, backfill_batch), False

    raise HTTPException(status_code=404, detail="Run not found")


def collect_backfill_result(run_id: str, batch: str) -> dict[str, Any]:
    manifest = _build_backfill_manifest(run_id, batch)
    manifest["generated_variant"] = "4:5"
    return manifest


def api_runs() -> dict[str, Any]:
    ensure_dirs()
    runs: list[dict[str, Any]] = []
    seen_run_ids: set[str] = set()
    seen_batches: set[str] = set()
    for run_dir in sorted(RUNS_ROOT.glob("run_*"), reverse=True):
        manifest = run_dir / "manifest.json"
        if manifest.exists():
            payload = json.loads(manifest.read_text(encoding="utf-8"))
            refreshed = refresh_manifest_file_state(run_dir, payload)
            run_id = str(refreshed.get("run_id") or run_dir.name)
            batch = str(refreshed.get("batch") or "").strip()
            if run_id in seen_run_ids:
                continue
            seen_run_ids.add(run_id)
            if batch:
                seen_batches.add(batch)
            runs.append(enrich_manifest_for_dashboard(refreshed))

    # Backfill batches that exist on disk but have no run manifest
    # (e.g., older/generated output imported from another machine).
    output_root = ROOT / "output"
    if output_root.exists():
        for batch_dir in sorted(output_root.glob("v*"), reverse=True):
            if not batch_dir.is_dir():
                continue
            batch_name = batch_dir.name
            if batch_name in seen_batches:
                continue
            prompt_files = scan_prompt_files_for_batch(batch_name)
            if not prompt_files:
                continue
            image_files = scan_image_files_for_batch(batch_name)
            regeneration_queue_files = scan_regeneration_queue_files_for_batch(batch_name)
            updated_at = datetime.fromtimestamp(batch_dir.stat().st_mtime, tz=timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
            runs.append(
                enrich_manifest_for_dashboard({
                    "run_id": f"batch_{batch_name}",
                    "batch": batch_name,
                    "prompt_files": prompt_files,
                    "image_files": image_files,
                    "regeneration_queue_files": regeneration_queue_files,
                    "image_generated": bool(image_files),
                    "updated_at": updated_at,
                    "source": "output_backfill",
                })
            )

    def batch_sort_key(run: dict[str, Any]) -> tuple[int, float]:
        batch = str(run.get("batch") or "").strip().lower()
        match = re.match(r"^v(\d+)$", batch)
        batch_num = int(match.group(1)) if match else -1
        updated = str(run.get("updated_at") or "")
        ts = 0.0
        if updated:
            try:
                ts = datetime.fromisoformat(updated.replace("Z", "+00:00")).timestamp()
            except Exception:
                ts = 0.0
        return (batch_num, ts)

    runs.sort(key=batch_sort_key, reverse=True)
    return {"runs": runs}


def api_run(run_id: str) -> dict[str, Any]:
    _run_dir, manifest, _has_storage_manifest = load_manifest_for_run(run_id)
    return enrich_manifest_for_dashboard(manifest)


def api_run_partial(run_id: str) -> dict[str, Any]:
    run_dir = RUNS_ROOT / run_id
    error_file = run_dir / "partial" / "error.txt"
    if error_file.exists():
        return {"ads": [], "progress": "0/0", "error": error_file.read_text(encoding="utf-8").strip()}
    partial_json = run_dir / "partial" / "ads.json"
    if not partial_json.exists():
        return {"ads": [], "progress": "0/0"}
    ads = json.loads(partial_json.read_text(encoding="utf-8"))
    progress_file = run_dir / "partial" / "progress.txt"
    progress = progress_file.read_text(encoding="utf-8").strip() if progress_file.exists() else "0/0"
    ads["progress"] = progress
    return ads


def api_run_prompt_copies(run_id: str) -> dict[str, Any]:
    _run_dir, manifest, _has_storage_manifest = load_manifest_for_run(run_id)
    prompt_files_all = manifest.get("prompt_files") or []
    prompt_files = [path for path in prompt_files_all if "/45/" in str(path)] or prompt_files_all
    records: list[dict[str, Any]] = []
    for rel_path in prompt_files:
        prompt_path = ROOT / rel_path
        if not prompt_path.exists() or not prompt_path.is_file():
            continue
        text = prompt_path.read_text(encoding="utf-8", errors="ignore")
        parsed_name = parse_prompt_filename(rel_path)
        persona_number = parsed_name[2] if parsed_name else None
        if persona_number is None:
            persona_number = parse_persona_number_from_prompt(text)
        records.append(
            {
                "prompt_file": rel_path,
                "format": parsed_name[0] if parsed_name else "",
                "language": parsed_name[1] if parsed_name else "",
                "persona_number": persona_number,
                "review_url": "/output/" + rel_path.replace("output/", ""),
                "copy_lines": extract_on_image_copy_lines(text),
            }
        )

    return {"run_id": run_id, "batch": manifest.get("batch"), "prompts": records}


def api_run_update_prompt_copies(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_dir = RUNS_ROOT / run_id
    manifest_path = run_dir / "manifest.json"
    copy_path = run_dir / "context" / "copy_batch.json"

    if not manifest_path.exists():
        raise HTTPException(status_code=404, detail="Run not found")
    if not copy_path.exists():
        raise HTTPException(status_code=404, detail="copy_batch.json not found for run")

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    batch = (manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    edits = payload.get("edits")
    if not isinstance(edits, list) or not edits:
        raise HTTPException(status_code=400, detail="edits must be a non-empty array")

    copy_json = json.loads(copy_path.read_text(encoding="utf-8"))
    ads = copy_json.get("ads")
    if not isinstance(ads, list) or not ads:
        raise HTTPException(status_code=400, detail="Invalid copy batch payload")

    updated_count = 0
    for entry in edits:
        if not isinstance(entry, dict):
            continue
        prompt_file = str(entry.get("prompt_file") or "").strip()
        if not prompt_file:
            continue
        parsed_name = parse_prompt_filename(prompt_file)
        if not parsed_name:
            continue
        fmt, lang, parsed_persona = parsed_name
        persona_number = entry.get("persona_number")
        if not isinstance(persona_number, int):
            persona_number = parsed_persona
        line_items = entry.get("copy_lines")
        if not isinstance(line_items, list) or not line_items:
            continue

        target_ad = None
        for ad in ads:
            if not isinstance(ad, dict):
                continue
            if str(ad.get("format") or "").strip().upper() != fmt:
                continue
            if isinstance(persona_number, int):
                persona = ad.get("persona")
                ad_persona_no = None
                if isinstance(persona, dict) and isinstance(persona.get("number"), int):
                    ad_persona_no = int(persona.get("number"))
                if ad_persona_no != persona_number:
                    continue
            target_ad = ad
            break
        if not isinstance(target_ad, dict):
            continue

        ad_copy = target_ad.setdefault("copy", {})
        if not isinstance(ad_copy, dict):
            continue
        lang_copy = ad_copy.setdefault(lang, {})
        if not isinstance(lang_copy, dict):
            continue

        for line_item in line_items:
            if not isinstance(line_item, dict):
                continue
            label = str(line_item.get("label") or "").strip()
            value = str(line_item.get("value") or "").strip()
            if not label:
                continue
            key = label.lower()

            if key == "headline":
                lang_copy["headline"] = value
            elif key == "subheadline":
                lang_copy["subheadline"] = value
            elif key == "support line":
                lang_copy["support_line"] = value
            elif key == "context line":
                lang_copy["context_line"] = value
            elif key == "cta":
                lang_copy["cta"] = value
            elif key == "attribution":
                lang_copy["attribution"] = value
            elif key == "trust line":
                lang_copy["trust_line"] = value
            elif key.startswith("bullet "):
                match = re.match(r"^bullet\s+(\d+)$", key)
                if not match:
                    continue
                index = int(match.group(1)) - 1
                if index < 0:
                    continue
                bullets = lang_copy.get("bullets")
                if not isinstance(bullets, list):
                    bullets = []
                while len(bullets) <= index:
                    bullets.append("")
                bullets[index] = value
                lang_copy["bullets"] = bullets
            elif key.startswith("left situation ") or key.startswith("right shift "):
                match = re.match(r"^(left situation|right shift)\s+(\d+)$", key)
                if not match:
                    continue
                side = match.group(1)
                ordinal = int(match.group(2))
                if ordinal <= 0:
                    continue
                if side == "left situation":
                    index = ordinal - 1
                else:
                    index = ordinal + 1
                bullets = lang_copy.get("bullets")
                if not isinstance(bullets, list):
                    bullets = []
                while len(bullets) <= index:
                    bullets.append("")
                bullets[index] = value
                lang_copy["bullets"] = bullets

        updated_count += 1

    if updated_count == 0:
        raise HTTPException(status_code=400, detail="No valid prompt edits were provided")

    copy_path.write_text(json.dumps(copy_json, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    rerender_prompts_for_run(run_dir, batch, copy_path, load_run_language_mode(run_dir))

    _append_audit_log(
        run_dir,
        "prompt_updates",
        {
            "run_id": run_id,
            "batch": batch,
            "updated_count": updated_count,
        },
    )

    has_916 = any("/96/" in str(path) for path in (manifest.get("prompt_files") or []))
    if has_916:
        manifest = generate_916_for_run(run_dir, manifest)

    refreshed = collect_run_result(run_dir, batch, bool(manifest.get("image_generated", False)))
    refreshed["copy_edits_applied"] = updated_count
    merged = merge_manifest(run_dir, manifest, refreshed)
    return merged


# ──────────────────────────────────────────────────────────────────────────────
# Task 6/7/8: Export/Import on-image copy (EXACT ON-IMAGE COPY block)
# ──────────────────────────────────────────────────────────────────────────────

EXACT_COPY_SHEET_COLUMNS = [
    "prompt_id",
    "vn",
    "format",
    "persona_name",
    "persona_pain",
    "persona_desire",
    "persona_friction",
    "persona_proof",
    "persona_tone",
    "concept_angle",
    "concept_angle_definition",
    "hypothesis_type",
    "hypothesis_variant",
    "headline_copy",
    "exact_on_image_copy_block",
    "created_at",
]

def _extract_vn_from_prompt_rel_path(prompt_rel_path: str) -> str:
    # Expected pattern: output/v{N}/...
    # Keep backward compatible: if not found, return empty string.
    m = re.search(r"/output/(v\d+)(/|$)", prompt_rel_path.replace("\\", "/"))
    return m.group(1) if m else ""


def _extract_vn_from_image_path(image_path: str) -> str:
    # Expected pattern: generated_images/v{N}/...
    m = re.search(r"/generated_images/(v\d+)(/|$)", image_path.replace("\\", "/"))
    return m.group(1) if m else ""


def _extract_aspect_from_image_path(image_path: str) -> str:
    # Extract aspect folder e.g. 4_5 or 9_16 from generated_images/v77/4_5/...
    m = re.search(r"/(\d+_\d+)/", image_path.replace("\\", "/"))
    return m.group(1) if m else ""


def _extract_created_at_iso_from_file(file_path: Path) -> str:
    try:
        ts = file_path.stat().st_mtime
        return datetime.fromtimestamp(ts, tz=timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    except Exception:
        return ""


def _parse_exact_block_headline_value(block_text: str) -> str | None:
    """
    Preserve EXACT headline value text as written in the exact block.

    We intentionally do NOT trim/normalize:
    - keep any spaces immediately after the colon
    - keep punctuation/case/capitalization
    """
    for raw_line in (block_text or "").splitlines():
        # Allow optional whitespace before "-" and around "-", but preserve everything after "Headline:"
        m = re.match(r"^\s*-\s*Headline:(.*)$", raw_line)
        if not m:
            continue
        return m.group(1)
    return None


def _replace_exact_copy_block(prompt_text: str, new_block_text: str) -> str | None:
    m = EXACT_COPY_BLOCK_RE.search(prompt_text or "")
    if not m:
        return None
    start_idx = m.start("block")
    end_idx = m.end("block")
    return (prompt_text[:start_idx] + new_block_text + prompt_text[end_idx:])


def _load_run_prompt_files(run_id: str, aspect_ratios: list[str] | None = None) -> list[str]:
    _run_dir, manifest, _has_storage_manifest = load_manifest_for_run(run_id)
    prompt_files_all = manifest.get("prompt_files") or []
    if not aspect_ratios:
        return prompt_files_all
    result: list[str] = []
    for p in prompt_files_all:
        for ar in aspect_ratios:
            if f"/{ar}/" in str(p):
                result.append(p)
                break
    return result or prompt_files_all


def _get_architecture_definition(arch: dict[str, Any], group: str, variant: str) -> str:
    """Get the intent summary for a concept_variation field from copy_architecture.json."""
    if not arch or not group or not variant:
        return ""
    headline_archs = arch.get("headline_architectures") or {}
    group_data = headline_archs.get(group) or {}
    variant_data = group_data.get(variant) or {}
    return str(variant_data.get("meaning") or variant_data.get("intent") or variant_data.get("direction") or variant_data.get("template") or "").strip()


def _extract_prompt_row_metadata(run_id: str, copy_batch: dict[str, Any], prompt_rel_path: str, batch_vn: str = "") -> dict[str, Any]:
    prompt_path = ROOT / prompt_rel_path
    text = prompt_path.read_text(encoding="utf-8", errors="ignore")

    block = extract_exact_on_image_copy_block(text)
    headline_copy = None
    exact_block = ""
    if block is not None:
        headline_copy = _parse_exact_block_headline_value(block)
        exact_block = block.strip()
    if headline_copy is None:
        headline_copy = ""

    vn = _extract_vn_from_prompt_rel_path(prompt_rel_path)
    if not vn and batch_vn:
        vn = batch_vn
    created_at = _extract_created_at_iso_from_file(prompt_path)

    parsed = parse_prompt_filename(prompt_rel_path)
    fmt = parsed[0] if parsed else ""
    persona_number = parsed[2] if parsed else None
    creative_index = parse_prompt_creative_index(prompt_rel_path)
    if persona_number is None:
        persona_number = parse_persona_number_from_prompt(text)

    persona_name = ""
    persona_pain = ""
    persona_desire = ""
    persona_friction = ""
    persona_proof = ""
    persona_tone = ""
    concept_angle = ""
    hypothesis_type = ""
    hypothesis_variant = ""

    if isinstance(persona_number, int):
        # Pull full persona data from persona_seeds.json
        seed = PERSONA_SEED_INPUTS.get(persona_number) or {}
        persona_name = str(seed.get("persona_name") or f"Persona {persona_number}")
        persona_pain = str(seed.get("core_pattern", ""))
        persona_desire = str(seed.get("common_indian_moments", ""))
        persona_friction = str(seed.get("objections_raw", ""))
        persona_proof = str(seed.get("common_indian_moments", ""))
        persona_tone = str(seed.get("guardrail", ""))

        for ad in copy_batch.get("ads") or []:
            if not isinstance(ad, dict):
                continue
            if str(ad.get("format") or "").strip().upper() != fmt:
                continue
            persona_obj = ad.get("persona")
            if isinstance(persona_obj, dict):
                if isinstance(persona_obj.get("number"), int) and int(persona_obj.get("number")) == persona_number:
                    if int(ad.get("creative_index") or 1) != creative_index:
                        continue
                    # Override persona_name from copy_batch if available
                    pn = str(persona_obj.get("persona_name") or persona_obj.get("name") or "")
                    if pn:
                        persona_name = pn
                    concept_angle = str(ad.get("concept_angle") or ad.get("headline_angle") or "")
                    hyp = ad.get("hypothesis") if isinstance(ad.get("hypothesis"), dict) else {}
                    if hyp:
                        hypothesis_type = str(hyp.get("type") or hyp.get("variable_label") or "")
                        hypothesis_variant = str(hyp.get("variant") or "")
                    break

    arch = COPY_ARCH
    concept_angle_def = _get_architecture_definition(arch, "concept_angle", concept_angle)

    return {
        "prompt_id": prompt_rel_path,
        "vn": vn,
        "format": fmt,
        "persona_name": persona_name,
        "persona_pain": persona_pain,
        "persona_desire": persona_desire,
        "persona_friction": persona_friction,
        "persona_proof": persona_proof,
        "persona_tone": persona_tone,
        "concept_angle": concept_angle,
        "concept_angle_definition": concept_angle_def,
        "hypothesis_type": hypothesis_type,
        "hypothesis_variant": hypothesis_variant,
        "headline_copy": headline_copy,
        "exact_on_image_copy_block": exact_block,
        "created_at": created_at,
    }


def _append_audit_log(run_dir: Path, event_type: str, payload: dict[str, Any]) -> None:
    pass


def api_export_on_image_copy(run_id: str) -> StreamingResponse:
    from openpyxl import Workbook
    import io

    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)

    copy_batch: dict[str, Any] = {"ads": []}
    if has_storage_manifest and run_dir is not None:
        copy_path = run_dir / "context" / "copy_batch.json"
        if copy_path.exists():
            copy_batch = json.loads(copy_path.read_text(encoding="utf-8"))

    prompt_files = _load_run_prompt_files(run_id)

    unique_vns = set()
    for rel in prompt_files:
        prompt_rel_path = str(rel).replace("\\", "/")
        vn = _extract_vn_from_prompt_rel_path(prompt_rel_path)
        if vn:
            unique_vns.add(vn)

    batch = manifest.get("batch", "")

    if not unique_vns:
        if batch and batch.startswith("v"):
            unique_vns.add(batch)

    if unique_vns:
        vn_suffix = "-".join(sorted(unique_vns))
    else:
        vn_suffix = None

    wb = Workbook()
    ws = wb.active
    ws.title = "on-image-copy"

    ws.append(EXACT_COPY_SHEET_COLUMNS)
    for rel in prompt_files:
        prompt_rel_path = str(rel).replace("\\", "/")
        if not (ROOT / prompt_rel_path).exists():
            continue
        row = _extract_prompt_row_metadata(run_id, copy_batch, prompt_rel_path, batch)
        ws.append([row.get(col, "") for col in EXACT_COPY_SHEET_COLUMNS])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if run_dir is not None:
        _append_audit_log(run_dir, "export_on_image_copy", {"run_id": run_id, "prompt_rows": len(prompt_files)})

    if vn_suffix:
        filename = f"on-image-copy-{vn_suffix}.xlsx"
    else:
        filename = f"on-image-copy-{run_id}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def api_import_on_image_copy(
    run_id: str,
    file: UploadFile = File(...),
    confirm: bool = Form(False),
) -> dict[str, Any]:
    from openpyxl import load_workbook

    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)

    # Parse xlsx (no prompt regeneration; only exact-block replacement)
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing upload filename")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty upload")

    import_root = (run_dir / "imports") if run_dir is not None else (RUNTIME_ROOT / "imports")
    tmp_path = import_root / f"upload-{int(time.time())}-{file.filename}"
    tmp_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path.write_bytes(content)

    wb = load_workbook(tmp_path)
    ws = wb.active

    # Build column index
    header = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(header) if name}

    # Required columns for import (core copy fields)
    REQUIRED_IMPORT_COLUMNS = ["prompt_id", "headline_copy", "exact_on_image_copy_block"]
    missing_required = [c for c in REQUIRED_IMPORT_COLUMNS if c not in col_idx]
    if missing_required:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing_required}")

    # Optional context columns (from export, used for validation/reference)
    OPTIONAL_CONTEXT_COLUMNS = [
        "format", "persona_name", "persona_pain", "persona_desire", "persona_friction",
        "persona_proof", "persona_tone",
        "concept_angle", "concept_angle_definition",
        "hypothesis_type", "hypothesis_variant",
        "vn", "created_at",
    ]

    seen_prompt_ids: set[str] = set()
    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for excel_row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in excel_row]
        prompt_id = str(values[col_idx["prompt_id"]] or "").replace("\\", "/").strip()
        if not prompt_id:
            continue
        if prompt_id in seen_prompt_ids:
            errors.append({"prompt_id": prompt_id, "error": "duplicate_prompt_id"})
            continue
        seen_prompt_ids.add(prompt_id)

        headline_copy = str(values[col_idx["headline_copy"]] or "").strip()
        full_block = str(values[col_idx.get("exact_on_image_copy_block", -1)] or "").strip() if "exact_on_image_copy_block" in col_idx else ""

        # Read optional context columns if present
        ctx: dict[str, str] = {}
        for opt_col in OPTIONAL_CONTEXT_COLUMNS:
            if opt_col in col_idx:
                ctx[opt_col] = str(values[col_idx[opt_col]] or "").strip()

        if not headline_copy.strip() and not full_block:
            errors.append({"prompt_id": prompt_id, "error": "empty_headline_copy_and_block"})
            continue

        rows.append(
            {
                "prompt_id": prompt_id,
                "headline_copy": headline_copy,
                "full_block": full_block,
                "context": ctx,
            }
        )

    # Validate prompt_id exists
    for r in rows:
        p = ROOT / r["prompt_id"]
        if not p.exists() or not p.is_file():
            errors.append({"prompt_id": r["prompt_id"], "error": "prompt_id_not_found"})
    if errors:
        if run_dir is not None:
            _append_audit_log(run_dir, "import_on_image_copy_validation_failed", {"run_id": run_id, "errors": errors, "confirm": confirm})
        raise HTTPException(status_code=400, detail={"validation_errors": errors})

    # Preview diffs
    preview_items: list[dict[str, Any]] = []
    applied_count = 0
    skipped_count = 0

    for r in rows:
        prompt_rel_path = r["prompt_id"]
        prompt_path = ROOT / prompt_rel_path
        old_text = prompt_path.read_text(encoding="utf-8", errors="ignore")

        old_block = extract_exact_on_image_copy_block(old_text, warn_log_path=None)
        if old_block is None:
            skipped_count += 1
            preview_items.append({"prompt_id": prompt_rel_path, "status": "skipped_missing_exact_block"})
            continue

        full_block = r.get("full_block", "")
        new_block = None

        if full_block:
            new_block = full_block
            old_copy = old_block.strip()
            new_copy = new_block.strip()
        else:
            headline_copy = r.get("headline_copy", "")
            new_lines: list[str] = []
            headline_replaced = False
            for line in old_block.splitlines():
                m = re.match(r"^(\s*-\s*Headline:)(.*)$", line)
                if m:
                    new_lines.append(m.group(1) + headline_copy)
                    headline_replaced = True
                else:
                    new_lines.append(line)

            if not headline_replaced:
                skipped_count += 1
                preview_items.append({"prompt_id": prompt_rel_path, "status": "skipped_headline_line_not_found"})
                continue

            new_block = "\n".join(new_lines)
            old_copy = _parse_exact_block_headline_value(old_block) or ""
            new_copy = _parse_exact_block_headline_value(new_block) or ""

        preview_items.append(
            {
                "prompt_id": prompt_rel_path,
                "status": "ready_to_apply" if confirm else "preview",
                "old_copy": old_copy[:100] + "..." if len(old_copy) > 100 else old_copy,
                "new_copy": new_copy[:100] + "..." if len(new_copy) > 100 else new_copy,
            }
        )

        if confirm:
            updated_text = _replace_exact_copy_block(old_text, new_block)
            if updated_text is None:
                skipped_count += 1
                preview_items[-1]["status"] = "skipped_replace_failed"
                continue
            prompt_path.write_text(updated_text, encoding="utf-8")
            applied_count += 1

    if run_dir is not None:
        _append_audit_log(
            run_dir,
            "import_on_image_copy",
            {"run_id": run_id, "confirm": confirm, "rows": len(rows), "applied": applied_count, "skipped": skipped_count},
        )

    if not confirm:
        return {
            "run_id": run_id,
            "preview": True,
            "changed_rows_count": applied_count,
            "skipped_rows": skipped_count,
            "failed_rows": len(errors),
            "items": preview_items,
        }

    # Re-assemble prompts side-effects: since we directly edited prompt text,
    # we do not mutate copy_batch.json metadata (per requirements).
    # However, manifest/prompt_files state should be refreshed.
    merged: dict[str, Any] | None = None
    if run_dir is not None and has_storage_manifest:
        manifest = json.loads((run_dir / "manifest.json").read_text(encoding="utf-8"))
        batch = (manifest.get("batch") or "").strip()
        refreshed = collect_run_result(run_dir, batch, bool(manifest.get("image_generated", False)))
        refreshed["on_image_copy_import_applied"] = applied_count
        merged = merge_manifest(run_dir, manifest, refreshed)
    else:
        batch = str(manifest.get("batch") or "")
        merged = collect_backfill_result(run_id, batch) if batch else manifest
        merged["on_image_copy_import_applied"] = applied_count

    return {
        "run_id": run_id,
        "preview": False,
        "changed_rows_count": applied_count,
        "skipped_rows": skipped_count,
        "failed_rows": len(errors),
        "items": preview_items,
        "manifest": merged,
    }


def api_run_generate_916(run_id: str) -> dict[str, Any]:
    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
    if not has_storage_manifest or run_dir is None:
        raise HTTPException(status_code=400, detail="This endpoint requires run context in dashboard_storage. Use generate-images-916-from-45 for output-only batches.")
    return generate_916_for_run(run_dir, manifest)


def api_run_generate_916_selected(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
    if not has_storage_manifest or run_dir is None:
        raise HTTPException(status_code=400, detail="This endpoint requires run context in dashboard_storage for copy_batch filtering.")
    copy_path = run_dir / "context" / "copy_batch.json"
    if not copy_path.exists():
        raise HTTPException(status_code=404, detail="copy_batch.json not found for run")
    batch = str(manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    prompt_files = payload.get("prompt_files")
    if not isinstance(prompt_files, list) or not prompt_files:
        raise HTTPException(status_code=400, detail="prompt_files must be a non-empty array")

    selected_45 = validate_selected_45_prompts(batch, prompt_files)
    if not selected_45:
        raise HTTPException(status_code=400, detail="No valid 4:5 prompt files selected")

    selected_keys = extract_selected_ad_keys_from_45_prompts(selected_45)
    if not selected_keys:
        raise HTTPException(status_code=400, detail="Could not resolve selected persona/format keys")

    copy_json = json.loads(copy_path.read_text(encoding="utf-8"))
    selected_copy = filter_copy_json_for_selected_ads(copy_json, selected_keys)
    ads = selected_copy.get("ads")
    if not isinstance(ads, list) or not ads:
        raise HTTPException(status_code=400, detail="No ads matched selected prompts")

    copy_916 = force_aspect_ratio(selected_copy, "9:16")
    visual_locks = collect_45_visual_locks(batch)
    if visual_locks:
        copy_916 = apply_visual_locks(copy_916, visual_locks)
    copy_916_path = run_dir / "context" / "copy_batch_916_selected.json"
    copy_916_path.write_text(json.dumps(copy_916, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    result = run_cmd(
        [
            "python3",
            "scripts/generate_ads.py",
            "--copy-file",
            str(copy_916_path),
            "--batch",
            batch,
            "--language-mode",
            load_run_language_mode(run_dir),
            "--no-registry-write",
            "--skip-uniqueness-check",
        ],
        cwd=ROOT,
        run_id=run_id,
    )
    if result.returncode != 0:
        if cancel_event_for_run(run_id).is_set() or _cancel_current_run.is_set():
            raise HTTPException(status_code=499, detail="Cancelled by user")
        error_text = result.stderr or result.stdout
        (run_dir / "logs" / "assembler_916_selected_error.txt").write_text(error_text, encoding="utf-8")
        short_error = "\n".join([line for line in error_text.splitlines() if line.strip()][-12:])
        raise HTTPException(status_code=500, detail=f"Selective 9:16 generation failed: {short_error}")

    refreshed = collect_run_result(run_dir, batch, bool(manifest.get("image_generated", False)))
    refreshed["generated_variant"] = "9:16"
    refreshed["generated_916_for_prompts"] = selected_45
    return merge_manifest(run_dir, manifest, refreshed)


def validate_selected_45_prompts(batch: str, prompt_files: list[Any]) -> list[str]:
    valid_prompt_files: list[str] = []
    for prompt_file in prompt_files:
        rel = str(prompt_file or "").strip().replace("\\", "/")
        if not rel or not rel.startswith("output/"):
            continue
        if "/45/" not in rel:
            continue
        candidate = ROOT / rel
        if not candidate.exists() or not candidate.is_file():
            continue
        if f"output/{batch}/" not in rel:
            continue
        valid_prompt_files.append(rel)
    return valid_prompt_files


def map_45_to_96_prompts(selected_45: list[str]) -> list[str]:
    out: list[str] = []
    for rel in selected_45:
        rel_96 = rel.replace("/45/", "/96/")
        file_96 = ROOT / rel_96
        if file_96.exists() and file_96.is_file():
            out.append(rel_96)
    return out


def extract_selected_ad_keys_from_45_prompts(selected_45: list[str]) -> set[tuple[str, int | None]]:
    keys: set[tuple[str, int | None]] = set()
    for rel in selected_45:
        parsed = parse_prompt_filename(rel)
        if not parsed:
            continue
        fmt, _lang, persona_number = parsed
        keys.add((fmt, persona_number))
    return keys


def filter_copy_json_for_selected_ads(copy_json: dict[str, Any], selected_keys: set[tuple[str, int | None]]) -> dict[str, Any]:
    ads = copy_json.get("ads")
    if not isinstance(ads, list):
        return copy_json
    selected_ads: list[dict[str, Any]] = []
    for ad in ads:
        if not isinstance(ad, dict):
            continue
        fmt = str(ad.get("format") or "").strip().upper()
        persona_number = None
        persona = ad.get("persona")
        if isinstance(persona, dict) and isinstance(persona.get("number"), int):
            persona_number = int(persona.get("number"))
        if (fmt, persona_number) in selected_keys or (fmt, None) in selected_keys:
            selected_ads.append(ad)
    cloned = json.loads(json.dumps(copy_json, ensure_ascii=False))
    cloned["ads"] = selected_ads
    return cloned


def api_run_generate_images_45(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
    batch = str(manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    prompt_files = payload.get("prompt_files")
    if not isinstance(prompt_files, list) or not prompt_files:
        raise HTTPException(status_code=400, detail="prompt_files must be a non-empty array")

    selected_45 = validate_selected_45_prompts(batch, prompt_files)
    if not selected_45:
        raise HTTPException(status_code=400, detail="No valid 4:5 prompt files selected")

    headless = bool(payload.get("headless", False))
    engine = str(payload.get("engine") or "gemini").strip().lower()
    try:
        if engine == "chatgpt":
            result = run_chatgpt_generation(
                batch=batch,
                prompt_files=selected_45,
                aspect_ratio="4:5",
                image_sources_file=None,
                headless=headless,
                run_dir=run_dir,
            )
        else:
            result = run_gemini_generation(
                batch=batch,
                prompt_files=selected_45,
                aspect_ratio="4:5",
                image_sources_file=None,
                headless=headless,
                run_dir=run_dir,
            )
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if result.returncode != 0:
        error_text = result.stderr or result.stdout
        engine_label = "ChatGPT" if engine == "chatgpt" else "Gemini"
        log_path = RUNTIME_ROOT / "generation_logs" / f"gen_{batch}_4_5{'_chatgpt' if engine == 'chatgpt' else ''}.log"
        if run_dir is not None:
            (run_dir / "logs" / f"image_generation_45_error{'_chatgpt' if engine == 'chatgpt' else ''}.txt").write_text(error_text, encoding="utf-8")
        short_error = "\n".join([line for line in error_text.splitlines() if line.strip()][-6:])
        raise HTTPException(status_code=500, detail=f"{engine_label} image generation failed (4:5). Log: {log_path}\n{short_error}")

    if not has_storage_manifest or run_dir is None:
        refreshed = collect_backfill_result(run_id, batch)
        refreshed["generated_images_for_prompts_45"] = selected_45
        return refreshed

    refreshed = collect_run_result(run_dir, batch, True)
    refreshed["generated_images_for_prompts_45"] = selected_45
    merged = merge_manifest(run_dir, manifest, refreshed)
    return merged


def api_run_generate_images_916_from_45(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
    batch = str(manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    prompt_files = payload.get("prompt_files")
    if not isinstance(prompt_files, list) or not prompt_files:
        raise HTTPException(status_code=400, detail="prompt_files must be a non-empty array")

    selected_45 = validate_selected_45_prompts(batch, prompt_files)
    if not selected_45:
        raise HTTPException(status_code=400, detail="No valid 4:5 prompt files for 9:16 generation")

    selected_keys = extract_selected_ad_keys_from_45_prompts(selected_45)
    all_jobs = collect_45_reference_jobs_for_batch(batch)
    selected_jobs = [
        job
        for job in all_jobs
        if (job["format"], int(job["persona_number"])) in selected_keys or (job["format"], None) in selected_keys
    ]
    if not selected_jobs:
        raise HTTPException(status_code=400, detail="No usable 4:5 reference images matched selected prompts")

    headless = bool(payload.get("headless", False))
    engine = str(payload.get("engine") or "gemini").strip().lower()
    if engine not in {"gemini", "chatgpt"}:
        raise HTTPException(status_code=400, detail="engine must be gemini or chatgpt")
    result = run_916_conversion_from_45_for_batch(batch=batch, headless=headless, run_dir=run_dir, engine=engine, jobs=selected_jobs)

    if not has_storage_manifest or run_dir is None:
        refreshed = collect_backfill_result(run_id, batch)
        refreshed["generated_images_for_prompts_916"] = result.get("prompt_files_used", [])
        refreshed["generated_variant"] = "9:16"
        refreshed["conversion_failures"] = result.get("failures", [])
        return refreshed

    refreshed = collect_run_result(run_dir, batch, True)
    refreshed["generated_images_for_prompts_916"] = result.get("prompt_files_used", [])
    refreshed["generated_variant"] = "9:16"
    refreshed["conversion_failures"] = result.get("failures", [])
    merged = merge_manifest(run_dir, manifest, refreshed)
    return merged


def api_batch_generate_images_45(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_ids = payload.get("run_ids")
    if not isinstance(run_ids, list) or not run_ids:
        raise HTTPException(status_code=400, detail="run_ids must be a non-empty array")

    all_prompt_files: list[str] = []
    run_info: list[dict[str, Any]] = []

    primary_run_dir: Path | None = None
    for run_id in run_ids:
        try:
            run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
        except HTTPException:
            continue
        batch = str(manifest.get("batch") or "").strip()
        if not batch:
            continue
        prompt_files_all = manifest.get("prompt_files") or []
        prompt_files_45 = [path for path in prompt_files_all if "/45/" in str(path)]
        if not prompt_files_45:
            continue
        all_prompt_files.extend(prompt_files_45)
        if has_storage_manifest and run_dir is not None and primary_run_dir is None:
            primary_run_dir = run_dir
        run_info.append({
            "run_id": run_id,
            "batch": batch,
            "prompt_count": len(prompt_files_45),
        })

    if not all_prompt_files:
        raise HTTPException(status_code=400, detail="No 4:5 prompt files found for any run")

    batch_names = sorted({r["batch"] for r in run_info})
    batch_name = batch_names[0] if len(batch_names) == 1 else "_".join(batch_names)
    work_id = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
    engine = str(payload.get("engine") or "gemini").strip().lower()
    engine_label = "ChatGPT" if engine == "chatgpt" else "Gemini"
    prompt_work_dir = RUNTIME_ROOT / f"{engine.lower()}_selected_prompts" / f"{batch_name}_{work_id}"
    prompt_work_dir.mkdir(parents=True, exist_ok=True)
    starting_prompt = ""
    starting_prompt_path = ROOT / "input" / "startingprompt.txt"
    if starting_prompt_path.exists():
        starting_prompt = starting_prompt_path.read_text(encoding="utf-8").strip()
    prompt_files_created: list[str] = []
    for src_pf in all_prompt_files:
        src = Path(src_pf)
        if not src.is_absolute():
            src = ROOT / src
        src = src.resolve()
        if not src.exists():
            continue
        prompt_text = src.read_text(encoding="utf-8")
        combined = f"{starting_prompt}\n\n{prompt_text.strip()}\n" if starting_prompt else prompt_text
        dest = prompt_work_dir / src.name
        dest.write_text(combined, encoding="utf-8")
        sidecar = src.with_suffix(".json")
        if sidecar.exists():
            (prompt_work_dir / sidecar.name).write_text(sidecar.read_text(encoding="utf-8"), encoding="utf-8")
        prompt_files_created.append(str(dest))
    headless = bool(payload.get("headless", False))
    out_dir = GENERATED_IMAGES_ROOT / batch_name / "4_5"
    out_dir.mkdir(parents=True, exist_ok=True)

    if engine == "chatgpt":
        cmd = [
            sys.executable,
            "scripts/chatgpt_web_sutomation.py",
            "--prompt-dir",
            str(prompt_work_dir),
            "--prompt-glob",
            "*.txt",
            "--out-dir",
            str(out_dir),
            "--timeout",
            str(int(os.getenv("CHATGPT_GENERATION_TIMEOUT_SECONDS") or "420")),
            "--download-timeout",
            str(int(os.getenv("CHATGPT_DOWNLOAD_TIMEOUT_SECONDS") or "90")),
            "--manual-login-timeout",
            str(int(os.getenv("CHATGPT_MANUAL_LOGIN_TIMEOUT_SECONDS") or "180")),
            "--upload-dir",
            str(INPUT_IMAGES_DIR),
            "--aspect-ratio", "4:5",
        ]
        # Pass CDP URL if running in WSL
        if Path("/mnt/c").exists():
            cmd.extend(["--cdp-url", wsl_chrome_cdp_url()])
    else:
        cmd = [
            sys.executable,
            "scripts/gemini_web_automation.py",
            "--prompt-dir",
            str(prompt_work_dir),
            "--prompt-glob",
            "*.txt",
            "--out-dir",
            str(out_dir),
            "--timeout",
            str(int(os.getenv("GEMINI_GENERATION_TIMEOUT_SECONDS") or "420")),
            "--manual-login-timeout",
            str(int(os.getenv("GEMINI_MANUAL_LOGIN_TIMEOUT_SECONDS") or "180")),
            "--upload-dir",
            str(INPUT_IMAGES_DIR),
        ]
    if headless:
        cmd.append("--headless")

    result = run_cmd(cmd, cwd=ROOT)
    if result.returncode != 0:
        error_text = result.stderr or result.stdout
        short_error = "\n".join([line for line in error_text.splitlines() if line.strip()][-30:])
        raise HTTPException(status_code=500, detail=f"Batch 4:5 generation failed ({engine_label}):\n{short_error}")

    return {
        "status": "completed",
        "batch_key": batch_name,
        "total_prompts": len(prompt_files_created),
        "run_count": len(run_ids),
    }


def api_batch_generate_images_both(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """First generate 4:5 images, then generate 9:16 from them."""
    run_ids = payload.get("run_ids")
    if not isinstance(run_ids, list) or not run_ids:
        raise HTTPException(status_code=400, detail="run_ids must be a non-empty array")

    headless = bool(payload.get("headless", False))
    engine = str(payload.get("engine") or "gemini").strip().lower()
    if engine not in {"gemini", "chatgpt"}:
        raise HTTPException(status_code=400, detail="engine must be gemini or chatgpt")
    engine_label = "ChatGPT" if engine == "chatgpt" else "Gemini"

    # ---- Step 1: Generate 4:5 images ----
    all_prompt_files: list[str] = []
    run_info: list[dict[str, Any]] = []
    primary_run_dir: Path | None = None
    for run_id in run_ids:
        try:
            run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
        except HTTPException:
            continue
        batch = str(manifest.get("batch") or "").strip()
        if not batch:
            continue
        prompt_files_all = manifest.get("prompt_files") or []
        prompt_files_45 = [path for path in prompt_files_all if "/45/" in str(path)]
        if not prompt_files_45:
            continue
        all_prompt_files.extend(prompt_files_45)
        if has_storage_manifest and run_dir is not None and primary_run_dir is None:
            primary_run_dir = run_dir
        run_info.append({"run_id": run_id, "batch": batch, "prompt_count": len(prompt_files_45)})

    if not all_prompt_files:
        raise HTTPException(status_code=400, detail="No 4:5 prompt files found for any run")

    batch_names = sorted({r["batch"] for r in run_info})
    batch_name = batch_names[0] if len(batch_names) == 1 else "_".join(batch_names)
    work_id = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
    prompt_work_dir = RUNTIME_ROOT / f"{engine.lower()}_selected_prompts" / f"{batch_name}_{work_id}"
    prompt_work_dir.mkdir(parents=True, exist_ok=True)
    starting_prompt = ""
    starting_prompt_path = ROOT / "input" / "startingprompt.txt"
    if starting_prompt_path.exists():
        starting_prompt = starting_prompt_path.read_text(encoding="utf-8").strip()
    for src_pf in all_prompt_files:
        src = Path(src_pf)
        if not src.is_absolute():
            src = ROOT / src
        src = src.resolve()
        if not src.exists():
            continue
        prompt_text = src.read_text(encoding="utf-8")
        combined = f"{starting_prompt}\n\n{prompt_text.strip()}\n" if starting_prompt else prompt_text
        dest = prompt_work_dir / src.name
        dest.write_text(combined, encoding="utf-8")
        sidecar = src.with_suffix(".json")
        if sidecar.exists():
            (prompt_work_dir / sidecar.name).write_text(sidecar.read_text(encoding="utf-8"), encoding="utf-8")

    out_dir_45 = GENERATED_IMAGES_ROOT / batch_name / "4_5"
    out_dir_45.mkdir(parents=True, exist_ok=True)

    if engine == "chatgpt":
        cmd = [
            sys.executable, "scripts/chatgpt_web_sutomation.py",
            "--prompt-dir", str(prompt_work_dir), "--prompt-glob", "*.txt",
            "--out-dir", str(out_dir_45),
            "--timeout", str(int(os.getenv("CHATGPT_GENERATION_TIMEOUT_SECONDS") or "420")),
            "--download-timeout", str(int(os.getenv("CHATGPT_DOWNLOAD_TIMEOUT_SECONDS") or "90")),
            "--manual-login-timeout", str(int(os.getenv("CHATGPT_MANUAL_LOGIN_TIMEOUT_SECONDS") or "180")),
            "--upload-dir", str(INPUT_IMAGES_DIR),
            "--aspect-ratio", "4:5",
        ]
        if Path("/mnt/c").exists():
            cmd.extend(["--cdp-url", wsl_chrome_cdp_url()])
    else:
        cmd = [
            sys.executable, "scripts/gemini_web_automation.py",
            "--prompt-dir", str(prompt_work_dir), "--prompt-glob", "*.txt",
            "--out-dir", str(out_dir_45),
            "--timeout", str(int(os.getenv("GEMINI_GENERATION_TIMEOUT_SECONDS") or "420")),
            "--manual-login-timeout", str(int(os.getenv("GEMINI_MANUAL_LOGIN_TIMEOUT_SECONDS") or "180")),
            "--upload-dir", str(INPUT_IMAGES_DIR),
        ]
    if headless:
        cmd.append("--headless")

    result = run_cmd(cmd, cwd=ROOT)
    if result.returncode != 0:
        error_text = result.stderr or result.stdout
        short_error = "\n".join([line for line in error_text.splitlines() if line.strip()][-30:])
        raise HTTPException(status_code=500, detail=f"4:5 generation failed ({engine_label}):\n{short_error}")

    # ---- Step 2: Generate 9:16 from 4:5 images ----
    batch_errors: list[str] = []
    total_completed = 0
    total_attempted = 0
    processed_batches: list[str] = []
    batch_to_run_dir: dict[str, Path | None] = {}
    for run_id in run_ids:
        try:
            run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
        except HTTPException:
            continue
        batch = str(manifest.get("batch") or "").strip()
        if not batch:
            continue
        if has_storage_manifest and run_dir is not None:
            batch_to_run_dir[batch] = run_dir
        elif batch not in batch_to_run_dir:
            batch_to_run_dir[batch] = None

    for batch, run_dir in sorted(batch_to_run_dir.items()):
        try:
            result = run_916_conversion_from_45_for_batch(batch=batch, headless=headless, run_dir=run_dir, engine=engine)
        except HTTPException as exc:
            batch_errors.append(f"{batch}: {exc.detail}")
            continue
        processed_batches.append(batch)
        total_attempted += int(result.get("attempted") or 0)
        total_completed += int(result.get("completed") or 0)
        if run_dir is not None:
            manifest_path = run_dir / "manifest.json"
            if manifest_path.exists():
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
                refreshed = collect_run_result(run_dir, batch, True)
                refreshed["generated_variant"] = "9:16"
                refreshed["generated_images_for_prompts_916"] = result.get("prompt_files_used", [])
                merge_manifest(run_dir, manifest, refreshed)

    if total_completed == 0:
        detail = "4:5 images generated but 9:16 conversion failed"
        if batch_errors:
            detail += ": " + " | ".join(batch_errors[:3])
        raise HTTPException(status_code=400, detail=detail)

    return {
        "status": "completed",
        "batch_key": ",".join(processed_batches),
        "message": f"4:5 + 9:16 images generated for {len(processed_batches)} batch(es)",
        "total_45_prompts": len(all_prompt_files),
        "total_916_completed": total_completed,
        "run_count": len(run_ids),
        "errors": batch_errors,
    }


def _resolve_916_generation_for_run(run_dir: Path, manifest: dict[str, Any]) -> list[dict[str, Any]]:
    """For a single run, build the list of {prompt_96, image_sources} entries for 9:16 generation.

    Checks manifest for existing 9:16 prompt files, falls back to deriving from 4:5 prompts.
    Uses load_batch_image_summary to find existing 4:5 images as references.
    """
    batch = (manifest.get("batch") or "").strip()
    if not batch:
        return []

    prompt_files_all = manifest.get("prompt_files") or []

    # First try to use existing 9:16 prompt files from the manifest
    prompt_files_96 = [p for p in prompt_files_all if "/96/" in str(p)]

    if prompt_files_96:
        # We have 9:16 prompts already; find their corresponding 4:5 images
        image_summary = load_batch_image_summary(batch)
        prompt_to_images: dict[str, list[str]] = {}
        for entry in image_summary:
            pf = entry.get("prompt_file") or ""
            saved = entry.get("saved_files") or []
            if pf and saved:
                prompt_to_images[pf] = saved

        entries: list[dict[str, Any]] = []
        for pf96 in prompt_files_96:
            rel_96 = str(pf96).replace("\\", "/")
            parsed = parse_prompt_filename(rel_96)
            if not parsed:
                continue
            fmt, lang, persona_num = parsed

            # Look for 4:5 image by matching format+persona
            image_sources: list[str] = []
            persona_slug_str = persona_slug(persona_num) if isinstance(persona_num, int) else ""
            for pf45, imgs in prompt_to_images.items():
                pf_upper = str(pf45).upper()
                if (f"{fmt}_P{persona_num:02d}" in pf_upper) or (persona_slug_str and persona_slug_str.upper() in pf_upper):
                    image_sources = list(imgs)
                    break

            # Fallback: search image roots directly for the 4:5 image
            if not image_sources:
                slug = persona_slug(persona_num) if isinstance(persona_num, int) else ""
                patterns = [f"*{slug}*", f"*p{persona_num:02d}*"] if slug else [f"*p{persona_num:02d}*"]
                for img_root in generated_image_roots():
                    ref_dir = img_root / batch / "4_5"
                    if not ref_dir.exists():
                        continue
                    for ext in ("png", "jpg", "jpeg", "webp"):
                        seen_in_pattern: set[str] = set()
                        for pattern in patterns:
                            for f in sorted(ref_dir.glob(f"**/{pattern}.{ext}")):
                                rel = str(f.relative_to(ROOT))
                                if rel in seen_in_pattern:
                                    continue
                                seen_in_pattern.add(rel)
                                if rel not in image_sources:
                                    image_sources.append(rel)
                        if image_sources:
                            break
                    if image_sources:
                        break

            if not image_sources:
                continue

            pf96_path = f"output/{batch}/96/{Path(pf96).name}"
            entries.append({
                "prompt_96": pf96_path,
                "image_sources": image_sources,
            })
        return entries

    # Fallback: derive 9:16 prompts from 4:5 prompts (if 96 outputs exist on disk)
    prompt_files_45 = [p for p in prompt_files_all if "/45/" in str(p)]
    image_summary = load_batch_image_summary(batch)
    prompt_to_images: dict[str, list[str]] = {}
    for entry in image_summary:
        pf = entry.get("prompt_file") or ""
        saved = entry.get("saved_files") or []
        if pf and saved:
            prompt_to_images[pf] = saved

    entries = []
    for pf in prompt_files_45:
        rel_45 = str(pf).replace("\\", "/")
        parsed = parse_prompt_filename(rel_45)
        if not parsed:
            continue
        fmt, lang, persona_num = parsed

        # 9:16 prompt expected at output/{batch}/96/
        persona_slug_str = persona_slug(persona_num) if isinstance(persona_num, int) else ""
        prompt_96_pattern = f"output/{batch}/96/{fmt}_{persona_slug_str}_{lang}*.txt"
        prompt_96_matches = sorted(ROOT.glob(prompt_96_pattern))
        if not prompt_96_matches:
            continue
        pf_filename = prompt_96_matches[0].name

        image_sources = list(prompt_to_images.get(rel_45, []))

        # Fallback: search image roots directly
        if not image_sources:
            slug = persona_slug(persona_num) if isinstance(persona_num, int) else ""
            patterns = [f"*{slug}*", f"*p{persona_num:02d}*"] if slug else [f"*p{persona_num:02d}*"]
            for img_root in generated_image_roots():
                ref_dir = img_root / batch / "4_5"
                if not ref_dir.exists():
                    continue
                for ext in ("png", "jpg", "jpeg", "webp"):
                    seen_in_pattern: set[str] = set()
                    found_one = False
                    for pattern in patterns:
                        for f in sorted(ref_dir.glob(f"**/{pattern}.{ext}")):
                            rel = str(f.relative_to(ROOT))
                            if rel in seen_in_pattern:
                                continue
                            seen_in_pattern.add(rel)
                            if rel not in image_sources:
                                image_sources.append(rel)
                            found_one = True
                        if found_one:
                            break
                    if image_sources:
                        break
                if image_sources:
                    break

        if not image_sources:
            continue

        entries.append({
            "prompt_96": prompt_96,
            "image_sources": image_sources,
        })

    return entries


def run_916_conversion_from_45_for_batch(
    *,
    batch: str,
    headless: bool,
    run_dir: Path | None,
    engine: str = "gemini",
    jobs: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    resolved_jobs = jobs if isinstance(jobs, list) else collect_45_reference_jobs_for_batch(batch)
    if not resolved_jobs:
        raise HTTPException(status_code=400, detail=f"No usable 4:5 reference images found for batch {batch}")

    template_path = ensure_916_conversion_template()
    template_text = template_path.read_text(encoding="utf-8").strip()
    prompt_root = RUNTIME_ROOT / "conversion_916_prompts" / f"{batch}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
    prompt_root.mkdir(parents=True, exist_ok=True)

    failures: list[str] = []
    completed = 0
    prompt_files_used: list[str] = []

    for index, job in enumerate(resolved_jobs, start=1):
        source_stem = ""
        image_path = job.get("image_abs") or job.get("image_rel") or ""
        if image_path:
            raw_stem = Path(str(image_path)).stem
            # The 4:5 image stem carries the aspect suffix (e.g. ..._pain_point_4_5).
            # Strip it so the 9:16 prompt name matches the original 4:5 prompt stem.
            source_stem = re.sub(r"_(?:4_5|9_16)$", "", raw_stem)
        prompt_name = build_916_conversion_prompt_job(
            job["format"],
            int(job["persona_number"]),
            job["language"],
            index,
            source_stem=source_stem,
        )
        prompt_path = prompt_root / prompt_name
        prompt_path.write_text(template_text + "\n", encoding="utf-8")

        source_file = prompt_root / f"{prompt_path.stem}.images.txt"
        source_file.write_text(str(job["image_abs"]) + "\n", encoding="utf-8")

        if engine == "chatgpt":
            result = run_chatgpt_generation(
                batch=batch,
                prompt_files=[str(prompt_path)],
                aspect_ratio="9:16",
                image_sources_file=str(source_file),
                headless=headless,
                run_dir=run_dir,
                prepend_starting_prompt=False,
                first_tab_mode="new",
            )
        else:
            result = run_gemini_generation(
                batch=batch,
                prompt_files=[str(prompt_path)],
                aspect_ratio="9:16",
                image_sources_file=str(source_file),
                headless=headless,
                run_dir=run_dir,
                prepend_starting_prompt=False,
                first_tab_mode="new",
            )

        if result.returncode != 0:
            failures.append(f"{prompt_name}: {(result.stderr or result.stdout or '').strip()[:300]}")
            continue

        completed += 1
        prompt_files_used.append(str(prompt_path))

    if completed == 0:
        short = "\n".join(failures[:3])
        engine_label = "ChatGPT" if engine == "chatgpt" else "Gemini"
        raise HTTPException(status_code=500, detail=f"9:16 conversion failed for batch {batch} ({engine_label}). {short}")

    return {
        "batch": batch,
        "completed": completed,
        "attempted": len(resolved_jobs),
        "failures": failures,
        "prompt_files_used": prompt_files_used,
    }


def api_batch_generate_images_916(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_ids = payload.get("run_ids")
    if not isinstance(run_ids, list) or not run_ids:
        raise HTTPException(status_code=400, detail="run_ids must be a non-empty array")

    headless = bool(payload.get("headless", False))
    engine = str(payload.get("engine") or "gemini").strip().lower()
    if engine not in {"gemini", "chatgpt"}:
        raise HTTPException(status_code=400, detail="engine must be gemini or chatgpt")
    batch_to_run_dir: dict[str, Path | None] = {}
    for run_id in run_ids:
        try:
            run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
        except HTTPException:
            continue
        batch = str(manifest.get("batch") or "").strip()
        if not batch:
            continue
        if has_storage_manifest and run_dir is not None:
            batch_to_run_dir[batch] = run_dir
        elif batch not in batch_to_run_dir:
            batch_to_run_dir[batch] = None

    if not batch_to_run_dir:
        raise HTTPException(status_code=400, detail="No valid batches found for selected runs")

    total_attempted = 0
    total_completed = 0
    processed_batches: list[str] = []
    batch_errors: list[str] = []

    for batch, run_dir in sorted(batch_to_run_dir.items()):
        try:
            result = run_916_conversion_from_45_for_batch(batch=batch, headless=headless, run_dir=run_dir, engine=engine)
        except HTTPException as exc:
            batch_errors.append(f"{batch}: {exc.detail}")
            continue

        processed_batches.append(batch)
        total_attempted += int(result.get("attempted") or 0)
        total_completed += int(result.get("completed") or 0)

        if run_dir is not None:
            manifest_path = run_dir / "manifest.json"
            if manifest_path.exists():
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
                refreshed = collect_run_result(run_dir, batch, True)
                refreshed["generated_variant"] = "9:16"
                refreshed["generated_images_for_prompts_916"] = result.get("prompt_files_used", [])
                merge_manifest(run_dir, manifest, refreshed)

    if total_completed == 0:
        detail = "No 9:16 conversions succeeded"
        if batch_errors:
            detail += ": " + " | ".join(batch_errors[:3])
        raise HTTPException(status_code=400, detail=detail)

    return {
        "status": "completed",
        "batch_key": ",".join(processed_batches),
        "total_prompts": total_completed,
        "attempted_prompts": total_attempted,
        "run_count": len(processed_batches),
        "errors": batch_errors,
    }


def extract_persona_input_block(prompt_text: str) -> str:
    markers = ["EXACT ON-IMAGE COPY", "PERSONA INPUT", "PERSONA:", "INPUT:"]
    for marker in markers:
        if marker in prompt_text.upper():
            start = prompt_text.upper().find(marker)
            if start != -1:
                return prompt_text[start:].strip()
    if len(prompt_text) > 50:
        return prompt_text.strip()
    return ""


def api_file_content(path: str, max_lines: int = 400) -> dict[str, Any]:
    file_path = resolve_safe_path(path)
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    lines = file_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    clipped = lines[:max_lines]
    return {
        "path": str(file_path.relative_to(ROOT)),
        "total_lines": len(lines),
        "shown_lines": len(clipped),
        "content": "\n".join(clipped),
    }


async def api_run_execute(
    config: str = Form(...),
    product_info_file: UploadFile | None = File(None),
    mechanism_file: UploadFile | None = File(None),
    faq_file: UploadFile | None = File(None),
    image_source_file: UploadFile | None = File(None),
    input_image_files: list[UploadFile] | None = File(None),
    clear_input_images: bool = Form(False),
    user_id: str = "dev_user",
) -> dict[str, Any]:
    ensure_dirs()
    try:
        cfg = json.loads(config)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid config JSON") from exc
    if str((cfg.get("server_type") or "opencode")).strip().lower() != "opencode":
        raise HTTPException(status_code=400, detail="Unsupported server type. Use OpenCode.")

    run_id = make_run_id()
    run_dir = RUNS_ROOT / run_id

    # Record ownership + run config in MongoDB
    _record_run_owner(run_id, user_id, config=cfg)
    (run_dir / "inputs").mkdir(parents=True, exist_ok=True)
    (run_dir / "logs").mkdir(parents=True, exist_ok=True)
    (run_dir / "context").mkdir(parents=True, exist_ok=True)

    product_path = save_upload(run_dir / "inputs" / "product master doc.txt", product_info_file)
    image_sources_path = save_upload(run_dir / "inputs" / "image_sources.txt", image_source_file)
    saved_input_images = store_uploaded_input_images(input_image_files or [], clear_input_images)

    product_file = coalesce_path(product_path, DEFAULT_PRODUCT_MASTER)
    image_sources_file_path = coalesce_path(image_sources_path, default_image_sources_file())

    try:
        base_plan = resolve_format_plan(cfg)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    hypothesis_cfg = cfg.get("hypothesis") or {}
    plan = expand_plan_with_hypothesis(base_plan, hypothesis_cfg)

    reuse_visual_patterns_from_run_id = str(cfg.get("reuse_visual_patterns_from_run_id") or "").strip()
    if reuse_visual_patterns_from_run_id:
        pattern_locks = collect_visual_pattern_reuse_locks(reuse_visual_patterns_from_run_id)
        plan, applied_patterns = apply_visual_pattern_reuse_to_plan(
            plan,
            pattern_locks,
            share_across_personas=bool(cfg.get("share_background_across_personas")),
        )
        (run_dir / "context" / "visual_pattern_reuse.json").write_text(
            json.dumps({"source_run_id": reuse_visual_patterns_from_run_id, "available_locks": len(pattern_locks), "applied_ads": applied_patterns, "share_background_across_personas": bool(cfg.get("share_background_across_personas"))}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8",
        )

    if hypothesis_cfg:
        (run_dir / "context" / "hypothesis_config.json").write_text(
            json.dumps(hypothesis_cfg, ensure_ascii=False, indent=2) + "\n", encoding="utf-8",
        )

    product_ctx_source = "attached_product_master_doc"
    extractor_model = "none"
    provider = (cfg.get("provider") or "").strip().lower()
    if provider == "google":
        execution_model = (cfg.get("google_model") or "").strip() or DEFAULT_GOOGLE_MODEL
        cfg["opencode_model"] = ""
    else:
        execution_model = sanitize_dashboard_model((cfg.get("opencode_model") or "").strip(), list_opencode_models())
        cfg["opencode_model"] = execution_model

    persona_library = parse_persona_library()
    ads_context: list[dict[str, Any]] = []
    format_seen_counts: dict[str, int] = {}
    for item in plan:
        persona_no = item["persona"]
        fmt = item["format"]
        format_seen_counts[fmt] = format_seen_counts.get(fmt, 0) + 1
        persona_payload = build_persona_payload(persona_no, persona_library)
        format_payload = {"format": fmt, "rules": []}
        copy_req = build_copy_requirements(persona_no, fmt, format_seen_counts[fmt], run_id)
        hyp_meta = item.get("hypothesis")
        concept = {}
        hyp_type = str(hyp_meta.get("type") or "").strip().lower() if isinstance(hyp_meta, dict) else ""
        variant = str(hyp_meta.get("variant") or "").strip() if isinstance(hyp_meta, dict) else ""
        if isinstance(hyp_meta, dict) and hyp_type and hyp_type != "none":
            guidance = _hypothesis_guidance(hyp_type, variant) if variant else ""
            copy_req["hypothesis"] = {"type": hyp_type, "variant": variant, "hypothesis_id": hyp_meta.get("hypothesis_id") or f"{hyp_type}-{variant}", "intent": guidance, "do_not_force_template": True}
            concept = copy_req.get("concept_variation") or {}
            if hyp_type == "concept_angle" and variant:
                concept["concept_angle"] = _framework_item("concept_angle", variant)
            elif hyp_type == "hook_structure" and variant:
                concept["hook_structure_override"] = variant
            copy_req["concept_variation"] = concept
            copy_req["selection_mode"] = "locked"
        if not concept.get("concept_angle"):
            concept["concept_angle"] = {"id": "auto"}
        copy_req["concept_variation"] = concept
        ads_context.append({"persona": persona_payload, "format_rules": format_payload, "format": fmt, "copy_requirements": copy_req, "hypothesis": hyp_meta, "visual_archetype": item.get("visual_archetype"), "visual_pattern_reused_from_run_id": item.get("visual_pattern_reused_from_run_id"), "visual_pattern_reuse_key": item.get("visual_pattern_reuse_key"), "creative_index": item.get("creative_index", 1), "creative_total": item.get("creative_total", 1), "background_group_key": item.get("background_group_key"), "share_background_across_personas": item.get("share_background_across_personas", False)})

    full_context = {"generated_at": now_iso(), "run_id": run_id, "language_mode": resolve_language_mode(cfg), "context_source": product_ctx_source, "context_extractor_model": extractor_model, "opencode_model": execution_model, "product_file_path": str(product_file), "ads": ads_context}
    (run_dir / "context" / "run_context.json").write_text(json.dumps({k: full_context[k] for k in ["generated_at", "run_id", "language_mode", "context_source", "context_extractor_model", "opencode_model", "product_file_path"]}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    # Run pipeline in background thread so frontend can poll partial results
    bg_kwargs = dict(run_dir=run_dir, cfg=cfg, full_context=full_context, image_sources_file_path=image_sources_file_path, saved_input_images=saved_input_images, reuse_visual_patterns_from_run_id=reuse_visual_patterns_from_run_id, product_ctx_source=product_ctx_source, extractor_model=extractor_model, execution_model=execution_model, ads_context=ads_context, user_id=user_id)
    threading.Thread(target=_run_pipeline_background, kwargs=bg_kwargs, daemon=True).start()

    return {"run_id": run_id, "status": "started"}


def _list_output_batches() -> list[int]:
    output_dir = ROOT / "output"
    if not output_dir.exists():
        return []
    out: list[int] = []
    for child in output_dir.iterdir():
        if child.is_dir():
            m = re.match(r"^v(\d+)$", child.name)
            if m:
                out.append(int(m.group(1)))
    return sorted(out)


def _reserve_batch_name() -> str:
    batches = _list_output_batches()
    return "v1" if not batches else f"v{batches[-1] + 1}"


def _run_pipeline_background(
    run_dir: Path, cfg: dict, full_context: dict,
    image_sources_file_path: Path, saved_input_images: list,
    reuse_visual_patterns_from_run_id: str,
    product_ctx_source: str, extractor_model: str,
    execution_model: str,
    ads_context: list,
    user_id: str = "",
) -> None:
    """Run the full pipeline in a background thread, writing results incrementally."""
    run_id = run_dir.name
    _update_run_status_db(run_id, "running")
    try:
        print(f"[PIPELINE] Starting background pipeline for run {run_id}", file=sys.stderr)
        # Reserve batch number early so incremental assembler runs write to the same batch dir
        reserved_batch = _reserve_batch_name()
        language_mode = assembler_language_mode(cfg)
        provider = (cfg.get("provider") or "").strip().lower()
        if provider == "google":
            llm_mode = "google_gemini"
            copy_json = call_google_gemini(cfg, full_context, run_dir, reserved_batch=reserved_batch, language_mode=language_mode)
        else:
            llm_mode = "opencode"
            copy_json = call_opencode_compatible(cfg, full_context, run_dir, reserved_batch=reserved_batch, language_mode=language_mode)
        if not copy_json:
            provider_label = "Google Gemini" if llm_mode == "google_gemini" else "OpenCode"
            error_msg = f"{provider_label} copy generation unavailable (no LLM response) and fallback template has been removed."
            (run_dir / "partial").mkdir(parents=True, exist_ok=True)
            (run_dir / "partial" / "error.txt").write_text(error_msg, encoding="utf-8")
            print(f"[PIPELINE ERROR] {error_msg}", file=sys.stderr)
            return
        opencode_failures = copy_json.pop("_opencode_failures", []) if isinstance(copy_json, dict) else []
        opencode_warnings = copy_json.pop("_opencode_warnings", []) if isinstance(copy_json, dict) else []
        opencode_session_rollovers = int(copy_json.pop("_opencode_session_rollovers", 0) or 0) if isinstance(copy_json, dict) else 0
        if opencode_failures:
            llm_mode = "opencode_partial_fallback"
        copy_json = normalize_generated_copy(copy_json, full_context, run_dir.name)
        copy_json = strip_internal_markers_from_payload(copy_json)
        copy_json = enforce_unique_ctas(copy_json, full_context)
        copy_json = scrub_on_image_copy(copy_json)
        reuse_backgrounds_from_run_id = str(cfg.get("reuse_backgrounds_from_run_id") or "").strip()
        if reuse_backgrounds_from_run_id:
            locks = collect_background_reuse_locks(reuse_backgrounds_from_run_id)
            copy_json, applied_locks = apply_background_reuse_locks(copy_json, locks, share_across_personas=bool(cfg.get("share_background_across_personas")))
            (run_dir / "context" / "background_reuse.json").write_text(json.dumps({"source_run_id": reuse_backgrounds_from_run_id, "available_locks": len(locks), "applied_ads": applied_locks, "share_background_across_personas": bool(cfg.get("share_background_across_personas"))}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        generated_copy_error = validate_generated_copy_payload(copy_json, ads_context, language_mode)
        if generated_copy_error:
            (run_dir / "logs" / "opencode_error.txt").write_text(generated_copy_error + "\n\nGenerated payload:\n" + json.dumps(copy_json, ensure_ascii=False, indent=2), encoding="utf-8")
            (run_dir / "partial").mkdir(parents=True, exist_ok=True)
            provider_label = "Google Gemini" if llm_mode == "google_gemini" else "OpenCode"
            if opencode_failures:
                error_detail = "; ".join(str(e).splitlines()[0] for e in opencode_failures[:3])
                (run_dir / "partial" / "error.txt").write_text(f"{provider_label} API error: {error_detail}", encoding="utf-8")
            else:
                (run_dir / "partial" / "error.txt").write_text(f"{provider_label} copy generation returned incomplete copy: {generated_copy_error}", encoding="utf-8")
            print(f"[PIPELINE ERROR] {generated_copy_error}", file=sys.stderr)
            return

        copy_file = run_dir / "context" / "copy_batch.json"
        copy_file.write_text(json.dumps(copy_json, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        assembler_result = run_cmd(["python3", "scripts/generate_ads.py", "--copy-file", str(copy_file), "--batch", reserved_batch, "--language-mode", language_mode], cwd=ROOT, run_id=run_dir.name)
        if assembler_result.returncode != 0:
            if cancel_event_for_run(run_dir.name).is_set() or _cancel_current_run.is_set():
                print(f"[PIPELINE] Assembler cancelled by user for run {run_dir.name}", file=sys.stderr)
                return
            assembler_error = assembler_result.stderr or assembler_result.stdout
            (run_dir / "logs" / "assembler_error.txt").write_text(assembler_error, encoding="utf-8")
            (run_dir / "partial").mkdir(parents=True, exist_ok=True)
            (run_dir / "partial" / "error.txt").write_text(f"Prompt assembly failed: {assembler_error}", encoding="utf-8")
            print(f"[PIPELINE ERROR] Prompt assembly failed: {assembler_error}", file=sys.stderr)
            return

        batch = reserved_batch

        manifest = collect_run_result(run_dir, batch, image_generated=False)
        manifest["llm_mode"] = llm_mode
        if llm_mode == "google_gemini":
            manifest["copy_source"] = f"google gemini — {execution_model}"
        else:
            manifest["copy_source"] = "opencode generated copy"
        if opencode_failures:
            manifest["copy_generation_failures"] = len(opencode_failures)
            manifest["copy_generation_notes"] = [str(e).splitlines()[0] for e in opencode_failures[:5]]
        if opencode_warnings:
            manifest["copy_generation_warnings"] = len(opencode_warnings)
            manifest["copy_warning_log"] = str((run_dir / "logs" / "opencode_error.txt").relative_to(ROOT))
            manifest["copy_generation_notes"] = [str(item).splitlines()[0] for item in opencode_warnings[:3]]
        if opencode_session_rollovers:
            manifest["copy_session_rollovers"] = opencode_session_rollovers
            manifest["copy_session_schedule"] = OPENCODE_ADS_PER_SESSION_SCHEDULE
            manifest["copy_session_log"] = str((run_dir / "logs" / "opencode_session.log").relative_to(ROOT))
        manifest["context_source"] = product_ctx_source
        manifest["context_extractor_model"] = extractor_model
        manifest["opencode_model"] = execution_model
        manifest["llm_mode"] = llm_mode
        manifest["image_sources_file"] = str(image_sources_file_path)
        manifest["input_images_dir"] = str(INPUT_IMAGES_DIR.relative_to(ROOT)).replace("\\", "/")
        manifest["input_images_uploaded"] = saved_input_images
        if reuse_visual_patterns_from_run_id:
            manifest["visual_pattern_reuse_from_run_id"] = reuse_visual_patterns_from_run_id
        (run_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        _update_run_status_db(run_id, "completed", extra={"manifest_summary": {"format_count": len(manifest.get("results", [])), "batch": batch}})
        _store_output_mapping(run_id, user_id, batch, manifest)
        partial_dir = run_dir / "partial"
        if partial_dir.exists():
            import shutil
            shutil.rmtree(partial_dir)
        print(f"[PIPELINE DONE] Run {run_id} completed, batch={batch}", file=sys.stderr)
    except Exception as exc:
        _update_run_status_db(run_id, "error", extra={"error": str(exc)[:500]})
        (run_dir / "logs" / "pipeline_error.txt").write_text(f"Pipeline background task failed: {exc}\n{traceback.format_exc()}", encoding="utf-8")
        (run_dir / "partial").mkdir(parents=True, exist_ok=True)
        (run_dir / "partial" / "error.txt").write_text(f"Pipeline failed: {exc}", encoding="utf-8")
        print(f"[PIPELINE ERROR] {exc}", file=sys.stderr)


# Chrome process tracking
_chrome_process: subprocess.Popen | None = None


def api_launch_visible_browser() -> dict[str, Any]:
    """Launch a visible Chrome instance with CDP enabled so the user can log in
    before automation begins."""
    global _chrome_process

    # Determine CDP URL based on whether we're in WSL
    is_wsl = Path("/mnt/c").exists()
    if is_wsl:
        try:
            ip_route = subprocess.run(["ip", "route"], capture_output=True, text=True, timeout=5)
            gw_line = [l for l in ip_route.stdout.splitlines() if "default" in l]
            win_host_ip = gw_line[0].split()[2] if gw_line else "127.0.0.1"
        except Exception:
            win_host_ip = "127.0.0.1"
    else:
        win_host_ip = "127.0.0.1"

    cdp_base_url = f"http://{win_host_ip}:9222"
    cdp_url = f"{cdp_base_url}/json/version"

    # Check if CDP is already responding
    try:
        resp = urllib.request.urlopen(cdp_url, timeout=2)
        if resp.status == 200:
            return {
                "status": "already_running",
                "cdp_url": cdp_base_url,
                "message": "Chrome with CDP is already running. Log in to ChatGPT if needed, then trigger generation.",
            }
    except Exception:
        pass

    # Kill only the Chrome process holding port 9222 (if any) — do NOT nuke all Chrome instances.
    import socket
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    port_held = False
    try:
        sock.bind(("127.0.0.1", 9222))
        sock.close()
    except OSError:
        port_held = True
        sock.close()

    if port_held and Path("/mnt/c").exists():
        try:
            netstat = subprocess.run(
                ["netstat.exe", "-ano", "-p", "TCP"],
                capture_output=True, text=True, timeout=10,
            )
            pids_to_kill: set[str] = set()
            for line in (netstat.stdout or "").splitlines():
                if ":9222" in line and "LISTENING" in line:
                    parts = line.split()
                    if parts:
                        pids_to_kill.add(parts[-1])
            for pid in pids_to_kill:
                try:
                    subprocess.run(
                        ["taskkill.exe", "/F", "/PID", pid],
                        capture_output=True, timeout=10,
                    )
                except Exception:
                    pass
            if pids_to_kill:
                time.sleep(2)
        except Exception:
            pass

    # Verify port 9222 is actually free
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.bind(("127.0.0.1", 9222))
        sock.close()
    except OSError:
        sock.close()
        raise HTTPException(status_code=500, detail="Port 9222 is still in use. Wait 10 seconds and try again.")

    chrome_bin = None
    use_wsl_launch = False
    for candidate in [
        "/usr/bin/google-chrome",
        "/usr/bin/google-chrome-stable",
        "/snap/bin/chromium",
        "/usr/bin/chromium-browser",
        "/usr/bin/chromium",
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/Applications/Chromium.app/Contents/MacOS/Chromium",
    ]:
        if Path(candidate).exists():
            chrome_bin = candidate
            break

    if not chrome_bin:
        wsl_user = detect_wsl_user()
        user_specific = (
            [f"/mnt/c/Users/{wsl_user}/AppData/Local/Google/Chrome/Application/chrome.exe"]
            if wsl_user
            else []
        )
        wsl_candidates = user_specific + [
            "/mnt/c/Program Files/Google/Chrome/Application/chrome.exe",
            "/mnt/c/Program Files (x86)/Google/Chrome/Application/chrome.exe",
        ]
        for candidate in wsl_candidates:
            if Path(candidate).exists():
                chrome_bin = candidate
                use_wsl_launch = True
                break

    if not chrome_bin:
        raise HTTPException(status_code=500, detail="Chrome binary not found on system (tried Linux and WSL paths)")

    user_data_dir = os.path.expandvars("$HOME/.config/google-chrome-cdp")

    if use_wsl_launch:
        # Use PowerShell script to launch Chrome (handles WSL2 networking issues)
        script_path = Path(__file__).resolve().parent.parent.parent / "scripts" / "launch_chrome_cdp.ps1"
        if not script_path.exists():
            raise HTTPException(status_code=500, detail=f"Chrome launch script not found: {script_path}")

        # Convert to Windows path for PowerShell
        win_script_path = str(script_path).replace("/mnt/c/", "C:\\").replace("/", "\\")

        print(f"[chrome-launch] Running PowerShell script: {win_script_path}")

        ps_path = "/mnt/c/Windows/System32/WindowsPowerShell/v1.0/powershell.exe"
        result = subprocess.run(
            [ps_path, "-ExecutionPolicy", "Bypass", "-File", win_script_path],
            capture_output=True, text=True, timeout=60,
        )

        print(f"[chrome-launch] PowerShell stdout: {result.stdout.strip()}")
        if result.stderr:
            print(f"[chrome-launch] PowerShell stderr: {result.stderr.strip()}")

        if result.returncode != 0 or "SUCCESS" not in result.stdout:
            raise HTTPException(
                status_code=500,
                detail=f"Chrome launch failed. stdout: {result.stdout}, stderr: {result.stderr}"
            )

        # Chrome is running and CDP is responding on Windows localhost
        # For WSL2, we need to use the Windows host IP to reach CDP
        try:
            ip_route = subprocess.run(["ip", "route"], capture_output=True, text=True, timeout=5)
            gw_line = [l for l in ip_route.stdout.splitlines() if "default" in l]
            win_host_ip = gw_line[0].split()[2] if gw_line else "127.0.0.1"
        except Exception:
            win_host_ip = "127.0.0.1"

        cdp_url = f"http://{win_host_ip}:9223/json/version"
        print(f"[chrome-launch] CDP URL: {cdp_url}")

        # Verify CDP is reachable from WSL
        try:
            resp = urllib.request.urlopen(cdp_url, timeout=5)
            if resp.status == 200:
                return {
                    "status": "launched",
                    "cdp_url": f"http://{win_host_ip}:9223",
                    "message": "Chrome launched. Log in to ChatGPT, then trigger image generation.",
                }
        except Exception as e:
            print(f"[chrome-launch] CDP verification failed: {e}")

        # If direct CDP fails, Chrome is still running - user can proceed manually
        return {
            "status": "launched",
            "cdp_url": f"http://{win_host_ip}:9223",
            "message": "Chrome launched. Log in to ChatGPT, then trigger image generation.",
        }
    else:
        cmd = [
            chrome_bin,
            "--remote-debugging-port=9222",
            f"--user-data-dir={user_data_dir}",
            "--no-first-run",
            "--no-default-browser-check",
        ]
        _chrome_process = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    # Wait for Chrome to initialize
    time.sleep(5)
    for attempt in range(20):
        try:
            resp = urllib.request.urlopen("http://127.0.0.1:9222/json/version", timeout=2)
            if resp.status == 200:
                return {
                    "status": "launched",
                    "cdp_url": "http://127.0.0.1:9222",
                    "message": "Chrome launched. Log in to ChatGPT, then trigger image generation.",
                }
        except Exception as e:
            print(f"[chrome-launch] CDP attempt {attempt+1} failed: {e}")
            time.sleep(1)

    proc_alive = _chrome_process.poll() is None if _chrome_process else False
    raise HTTPException(
        status_code=500,
        detail=f"Chrome launched but CDP not responding on port 9222. Process alive: {proc_alive}. Close all Chrome windows and try again."
    )


def api_kill_chrome() -> dict[str, Any]:
    """Kill the Chrome process started by launch-visible-browser and stop any running automation."""
    global _chrome_process
    killed = False
    if _chrome_process and _chrome_process.poll() is None:
        try:
            _chrome_process.terminate()
            try:
                _chrome_process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                _chrome_process.kill()
                _chrome_process.wait(timeout=3)
            killed = True
        except Exception:
            pass
        _chrome_process = None

    # Also kill any running gemini automation processes
    gemini_killed = 0
    for proc in psutil.process_iter(["pid", "name", "cmdline"]):
        try:
            cmdline = proc.info.get("cmdline") or []
            if any("gemini_web_automation" in c for c in cmdline):
                proc.kill()
                gemini_killed += 1
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

    # Also kill any running chatgpt automation processes
    chatgpt_killed = 0
    for proc in psutil.process_iter(["pid", "name", "cmdline"]):
        try:
            cmdline = proc.info.get("cmdline") or []
            if any("chatgpt_web_sutomation" in c for c in cmdline):
                proc.kill()
                chatgpt_killed += 1
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

    # Kill Windows Chrome CDP instances (only relevant when running inside WSL)
    win_chrome_killed = 0
    if Path("/mnt/c").exists():
        for candidate in ["taskkill.exe", "/mnt/c/Windows/System32/taskkill.exe"]:
            if Path(candidate).exists() or shutil.which(candidate):
                try:
                    subprocess.run(
                        [candidate, "/F", "/IM", "chrome.exe", "/FI", "WINDOWTITLE eq ChromeCDP*"],
                        capture_output=True, timeout=5,
                    )
                    win_chrome_killed = 1
                except Exception:
                    pass
                break

    return {"status": "killed", "chrome": killed, "gemini_processes": gemini_killed, "chatgpt_processes": chatgpt_killed, "windows_chrome": win_chrome_killed}


def api_stop_generation() -> dict[str, Any]:
    """Kill any running generation/assembly scripts (chatgpt, gemini, generate_ads, opencode)."""
    targets = ["chatgpt_web_sutomation", "gemini_web_automation", "generate_ads.py", "opencode"]
    counts: dict[str, int] = {t: 0 for t in targets}
    for proc in psutil.process_iter(["pid", "name", "cmdline"]):
        try:
            cmdline = proc.info.get("cmdline") or []
            joined = " ".join(cmdline)
            for target in targets:
                if target in joined:
                    proc.kill()
                    counts[target] += 1
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue
    return {"status": "killed", **counts}


def api_edit_prompt(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Edit a prompt file in-place, replacing only the EXACT ON-IMAGE COPY block."""
    run_dir = RUNS_ROOT / run_id
    prompt_path = payload.get("prompt_file", "")
    new_text = payload.get("text", "")
    if not prompt_path or not new_text:
        raise HTTPException(status_code=400, detail="prompt_file and text are required")

    full_path = ROOT / prompt_path
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="Prompt file not found")

    old_text = full_path.read_text(encoding="utf-8")
    updated_text = _replace_exact_copy_block(old_text, new_text)
    if updated_text is None:
        raise HTTPException(status_code=400, detail="No EXACT ON-IMAGE COPY block found in prompt file")
    full_path.write_text(updated_text, encoding="utf-8")
    return {"status": "saved", "prompt_file": prompt_path}


def api_delete_run(run_id: str) -> dict[str, Any]:
    run_dir = RUNS_ROOT / run_id
    if not run_dir.exists():
        raise HTTPException(status_code=404, detail="Run not found")

    manifest_path = run_dir / "manifest.json"
    batch = None
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        batch = manifest.get("batch")

    import shutil
    shutil.rmtree(run_dir)

    deleted_images = False
    deleted_prompts = False
    if batch:
        other_runs_with_same_batch = False
        for d in RUNS_ROOT.glob("run_*"):
            if d.name == run_id:
                continue
            mf = d / "manifest.json"
            if mf.exists():
                try:
                    m = json.loads(mf.read_text(encoding="utf-8"))
                    if m.get("batch") == batch:
                        other_runs_with_same_batch = True
                        break
                except (json.JSONDecodeError, OSError):
                    continue

        if not other_runs_with_same_batch:
            batch_images_dir = GENERATED_IMAGES_ROOT / batch
            if batch_images_dir.exists():
                shutil.rmtree(batch_images_dir)
                deleted_images = True

            batch_prompts_dir = ROOT / "output" / batch
            if batch_prompts_dir.exists():
                shutil.rmtree(batch_prompts_dir)
                deleted_prompts = True

    return {"status": "deleted", "run_id": run_id, "batch": batch, "deleted_images": deleted_images, "deleted_prompts": deleted_prompts}


def api_delete_prompt(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Delete a prompt file and remove it from the run manifest."""
    run_dir = RUNS_ROOT / run_id
    manifest_path = run_dir / "manifest.json"
    prompt_path = payload.get("prompt_file", "")
    if not prompt_path:
        raise HTTPException(status_code=400, detail="prompt_file is required")

    full_path = ROOT / prompt_path
    if full_path.exists():
        full_path.unlink()

    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest["prompt_files"] = [p for p in manifest.get("prompt_files", []) if p != prompt_path]
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return {"status": "deleted", "prompt_file": prompt_path}


def api_delete_image(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Delete a generated image and its metadata JSON."""
    run_dir = RUNS_ROOT / run_id
    manifest_path = run_dir / "manifest.json"
    image_path = payload.get("image_file", "")
    if not image_path:
        raise HTTPException(status_code=400, detail="image_file is required")

    full_path = ROOT / image_path
    if full_path.exists():
        full_path.unlink()

    # Also delete companion JSON metadata if it exists
    for json_path in (full_path.with_suffix(".json"), full_path.with_suffix(full_path.suffix + ".json")):
        if json_path.exists():
            json_path.unlink()

    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest["image_files"] = [p for p in manifest.get("image_files", []) if p != image_path]
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return {"status": "deleted", "image_file": image_path}


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(2, 10_000):
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise HTTPException(status_code=500, detail=f"Could not create unique path for {path.name}")


def _regeneration_archive_path(full_path: Path) -> Path:
    parent = full_path.parent
    if parent.name == "generated images":
        archive_dir = parent.parent / "to_be_regenerated"
    else:
        archive_dir = parent / "to_be_regenerated"
    archive_dir.mkdir(parents=True, exist_ok=True)
    return _unique_path(archive_dir / full_path.name)


def api_mark_images_to_regenerate(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Move bad generated images out of the active gallery before regeneration."""
    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
    batch = str(manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    image_files = payload.get("image_files")
    if not isinstance(image_files, list) or not image_files:
        raise HTTPException(status_code=400, detail="image_files must be a non-empty array")

    moved: list[dict[str, str]] = []
    skipped: list[dict[str, str]] = []
    generated_root = GENERATED_IMAGES_ROOT.resolve()

    for raw in image_files:
        rel = str(raw or "").strip().replace("\\", "/")
        if not rel:
            continue
        full_path = resolve_safe_path(rel)
        resolved = full_path.resolve()
        if generated_root not in resolved.parents:
            skipped.append({"image_file": rel, "reason": "not under generated_images"})
            continue
        if f"/generated_images/{batch}/" not in f"/{rel}":
            skipped.append({"image_file": rel, "reason": "not in this run batch"})
            continue
        if "/to_be_regenerated/" in rel:
            skipped.append({"image_file": rel, "reason": "already archived"})
            continue
        if not full_path.exists() or not full_path.is_file():
            skipped.append({"image_file": rel, "reason": "missing"})
            continue

        archive_path = _regeneration_archive_path(full_path)
        shutil.move(str(full_path), str(archive_path))
        archive_rel = str(archive_path.relative_to(ROOT)).replace("\\", "/")
        moved.append({"image_file": rel, "archived_file": archive_rel})

        meta_path = full_path.with_suffix(".json")
        if meta_path.exists() and meta_path.is_file():
            meta_archive = archive_path.with_suffix(".json")
            shutil.move(str(meta_path), str(_unique_path(meta_archive)))

    refreshed = collect_backfill_result(run_id, batch)
    if has_storage_manifest and run_dir is not None:
        refreshed = collect_run_result(run_dir, batch, True)
        refreshed = merge_manifest(run_dir, manifest, refreshed)

    return {
        "status": "archived",
        "moved": moved,
        "skipped": skipped,
        "manifest": enrich_manifest_for_dashboard(refreshed),
    }


def _original_path_for_queued_image(rel: str) -> str | None:
    misplaced = re.match(r"^(generated_images/[^/]+)/to_be_regenerated/generated images/(.+)$", rel)
    if misplaced:
        metadata = _read_image_metadata(rel)
        aspect = str((metadata.get("test_variables") or {}).get("aspect_ratio") or "4:5") if isinstance(metadata.get("test_variables"), dict) else "4:5"
        aspect_dir = "9_16" if aspect == "9:16" else "4_5"
        return f"{misplaced.group(1)}/{aspect_dir}/generated images/{misplaced.group(2)}"
    original = rel.replace("/to_be_regenerated/", "/")
    if original == rel:
        return None
    return original


def api_restore_images_from_regeneration_queue(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Move images from to_be_regenerated back to their original location."""
    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
    batch = str(manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    image_files = payload.get("image_files")
    if not isinstance(image_files, list) or not image_files:
        raise HTTPException(status_code=400, detail="image_files must be a non-empty array")

    restored: list[dict[str, str]] = []
    skipped: list[dict[str, str]] = []
    generated_root = GENERATED_IMAGES_ROOT.resolve()

    for raw in image_files:
        rel = str(raw or "").strip().replace("\\", "/")
        if not rel:
            continue
        if "/to_be_regenerated/" not in rel:
            skipped.append({"image_file": rel, "reason": "not in to_be_regenerated"})
            continue
        full_path = resolve_safe_path(rel)
        resolved = full_path.resolve()
        if generated_root not in resolved.parents:
            skipped.append({"image_file": rel, "reason": "not under generated_images"})
            continue
        if not full_path.exists() or not full_path.is_file():
            skipped.append({"image_file": rel, "reason": "missing"})
            continue

        original_rel = _original_path_for_queued_image(rel)
        if not original_rel:
            skipped.append({"image_file": rel, "reason": "could not resolve original path"})
            continue

        original_abs = (ROOT / original_rel).resolve()
        original_abs.parent.mkdir(parents=True, exist_ok=True)

        if original_abs.exists():
            # A new image already occupies this slot — move it aside
            backup = _unique_path(original_abs)
            shutil.move(str(original_abs), str(backup))

        shutil.move(str(resolved), str(original_abs))
        restored.append({
            "restored_file": original_rel,
            "archived_file": rel,
        })

        meta_path = full_path.with_suffix(".json")
        if meta_path.exists() and meta_path.is_file():
            original_meta = original_abs.with_suffix(".json")
            if original_meta.exists():
                backup_meta = _unique_path(original_meta)
                shutil.move(str(original_meta), str(backup_meta))
            shutil.move(str(meta_path), str(original_meta))

    refreshed = collect_backfill_result(run_id, batch)
    if has_storage_manifest and run_dir is not None:
        refreshed = collect_run_result(run_dir, batch, True)
        refreshed = merge_manifest(run_dir, manifest, refreshed)

    return {
        "status": "restored",
        "restored": restored,
        "skipped": skipped,
        "manifest": enrich_manifest_for_dashboard(refreshed),
    }


# ── Queue regeneration helpers ──────────────────────────────────────────────


def _find_prompt_by_name(prompt_name: str, prompt_files: list[str]) -> str:
    """Find a prompt path in prompt_files whose filename matches prompt_name."""
    target = Path(prompt_name).name
    for pf in prompt_files:
        if Path(pf).name == target:
            return pf
    return ""


def _build_output_stem_from_prompt(prompt_path: str, engine: str, aspect_dir: str = "") -> str:
    name = Path(prompt_path).stem
    aspect_suffix = f"_{aspect_dir}" if aspect_dir in ("4_5", "9_16") else ""
    # New format: <FMT>_<slug>_<LANG>_<angle>[_A<NN>]
    new_match = re.match(
        r"^(?:OUTPUT_|FINAL_)?([A-Z]+)_([a-z0-9][a-z0-9]*(?:_[a-z0-9]+)*)_(EN|HI|HINGLISH)_(?P<angle>[a-z][a-z_]*?)(?:_([AV]\d+))?$",
        name,
        flags=re.IGNORECASE,
    )
    if new_match:
        fmt = new_match.group(1).upper()
        slug = new_match.group(2).lower()
        lang = new_match.group(3).upper()
        angle = new_match.group("angle")
        variant = new_match.group(5).upper() if new_match.group(5) else ""
        variant_suffix = f"_{variant}" if variant else ""
        return f"{fmt}_{slug}_{lang}_{angle}{variant_suffix}{aspect_suffix}"
    # Angle-less new format: <FMT>_<slug>_<LANG>[_A<NN>]
    new_no_angle = re.match(
        r"^(?:OUTPUT_|FINAL_)?([A-Z]+)_([a-z0-9][a-z0-9]*(?:_[a-z0-9]+)*)_(EN|HI|HINGLISH)(?:_([AV]\d+))?$",
        name,
        flags=re.IGNORECASE,
    )
    if new_no_angle:
        fmt = new_no_angle.group(1).upper()
        slug = new_no_angle.group(2).lower()
        lang = new_no_angle.group(3).upper()
        variant = new_no_angle.group(4).upper() if new_no_angle.group(4) else ""
        variant_suffix = f"_{variant}" if variant else ""
        return f"{fmt}_{slug}_{lang}{variant_suffix}{aspect_suffix}"
    # Legacy format: <FMT>_P<NN>_<LANG>[_<angle>][_A<NN>]
    legacy_match = re.match(
        r"^(?:OUTPUT_|FINAL_)?([A-Za-z0-9]+)_P(\d+)_([A-Za-z0-9]+)(?:_([AV]\d+))?(?:_[a-z_]+)?$",
        name,
        flags=re.IGNORECASE,
    )
    if legacy_match:
        fmt = legacy_match.group(1).upper()
        persona = f"P{int(legacy_match.group(2)):02d}"
        lang = legacy_match.group(3).upper()
        angle_match = re.search(r"_[a-z_]+$", name)
        angle = angle_match.group(0)[1:] if angle_match else ""
        variant = legacy_match.group(4).upper() if legacy_match.group(4) else ""
        variant_suffix = f"_{variant}" if variant else ""
        if angle:
            return f"{fmt}_{persona}_{lang}_{angle}{variant_suffix}{aspect_suffix}"
        return f"{fmt}_{persona}_{lang}{variant_suffix}{aspect_suffix}"
    return ""


def _build_expected_output_path(batch: str, prompt_path: str, aspect_dir: str, engine: str) -> Path | None:
    """Compute the expected full output path for a generated image."""
    stem = _build_output_stem_from_prompt(prompt_path, engine, aspect_dir=aspect_dir)
    if not stem:
        return None
    return GENERATED_IMAGES_ROOT / batch / aspect_dir / "generated images" / f"{stem}.png"


def _find_45_parent_for_prompt(batch: str, prompt_path: str, engine: str) -> Path | None:
    """Find the 4:5 reference image for a given prompt."""
    expected = _build_expected_output_path(batch, prompt_path, "4_5", engine)
    if expected and expected.exists() and expected.is_file():
        return expected
    # Try the other engine
    other_engine = "gemini" if engine == "chatgpt" else "chatgpt"
    expected_other = _build_expected_output_path(batch, prompt_path, "4_5", other_engine)
    if expected_other and expected_other.exists() and expected_other.is_file():
        return expected_other
    # Scan the directory for any image whose metadata links to this prompt
    four_five_dir = GENERATED_IMAGES_ROOT / batch / "4_5" / "generated images"
    if four_five_dir.exists() and four_five_dir.is_dir():
        target_name = Path(prompt_path).name
        for child in four_five_dir.iterdir():
            if child.suffix.lower() not in {".png", ".jpg", ".jpeg", ".webp"}:
                continue
            meta = child.with_suffix(".json")
            if meta.exists() and meta.is_file():
                try:
                    meta_data = json.loads(meta.read_text(encoding="utf-8"))
                    if Path(meta_data.get("prompt_file", "")).name == target_name:
                        return child
                except Exception:
                    continue
    return None


def api_regenerate_queued_images(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Regenerate images already in the to_be_regenerated queue.

    Single endpoint that handles both 4:5 and 9:16 images in one call.
    New images stay in to_be_regenerated for user review before restore.
    """
    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
    batch = str(manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    image_files = payload.get("image_files")
    if not isinstance(image_files, list) or not image_files:
        raise HTTPException(status_code=400, detail="image_files must be a non-empty array")

    headless = bool(payload.get("headless", False))
    engine = str(payload.get("engine") or "gemini").strip().lower()
    if engine not in {"gemini", "chatgpt"}:
        raise HTTPException(status_code=400, detail="engine must be gemini or chatgpt")

    prompt_files_list = list(manifest.get("prompt_files") or [])
    generated_root = GENERATED_IMAGES_ROOT.resolve()

    prompts_45: set[str] = set()
    jobs_916: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []

    for raw in image_files:
        rel = str(raw or "").strip().replace("\\", "/")
        if not rel:
            continue
        if "/to_be_regenerated/" not in rel:
            skipped.append({"image_file": rel, "reason": "not in to_be_regenerated"})
            continue

        meta = _read_image_metadata(rel)
        prompt_name = str(meta.get("prompt_file") or "").strip()
        if not prompt_name:
            skipped.append({"image_file": rel, "reason": "no prompt_file in metadata"})
            continue
        prompt_path = _find_prompt_by_name(prompt_name, prompt_files_list)
        if not prompt_path:
            skipped.append({"image_file": rel, "reason": f"prompt {prompt_name} not found in manifest"})
            continue

        aspect_dir = "4_5" if "/4_5/" in rel else ("9_16" if "/9_16/" in rel else "")
        if not aspect_dir:
            skipped.append({"image_file": rel, "reason": "unknown aspect ratio"})
            continue

        if aspect_dir == "4_5":
            prompts_45.add(prompt_path)
        else:
            parent_path = _find_45_parent_for_prompt(batch, prompt_path, engine)
            if not parent_path:
                skipped.append({"image_file": rel, "reason": "no 4:5 parent image found"})
                continue
            parsed = parse_prompt_filename(prompt_path)
            if not parsed:
                skipped.append({"image_file": rel, "reason": "could not parse prompt filename"})
                continue
            p_fmt, p_lang, persona_num = parsed
            jobs_916.append({
                "format": p_fmt.upper(),
                "persona_number": int(persona_num) if persona_num else 0,
                "language": p_lang.upper(),
                "image_rel": str(parent_path.relative_to(ROOT)).replace("\\", "/"),
                "image_abs": str(parent_path.resolve()),
            })

    if not prompts_45 and not jobs_916:
        raise HTTPException(status_code=400, detail=f"No valid queued images to regenerate ({len(skipped)} skipped)")

    generated_files: list[str] = []

    # ── Regenerate 4:5 images ──────────────────────────────────────────
    if prompts_45:
        result: subprocess.CompletedProcess[str] | None = None
        try:
            if engine == "chatgpt":
                result = run_chatgpt_generation(
                    batch=batch,
                    prompt_files=list(prompts_45),
                    aspect_ratio="4:5",
                    image_sources_file=None,
                    headless=headless,
                    run_dir=run_dir,
                )
            else:
                result = run_gemini_generation(
                    batch=batch,
                    prompt_files=list(prompts_45),
                    aspect_ratio="4:5",
                    image_sources_file=None,
                    headless=headless,
                    run_dir=run_dir,
                )
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"4:5 regeneration failed: {exc}")
        generation_error = ""
        if result is not None and result.returncode != 0:
            generation_error = (result.stderr or result.stdout or "").strip()

        for pf in prompts_45:
            expected = _build_expected_output_path(batch, pf, "4_5", engine)
            if expected and expected.exists() and expected.is_file():
                archive_dir = GENERATED_IMAGES_ROOT / batch / "4_5" / "to_be_regenerated"
                archive_dir.mkdir(parents=True, exist_ok=True)
                dest = _unique_path(archive_dir / expected.name)
                shutil.move(str(expected), str(dest))
                generated_files.append(str(dest.relative_to(ROOT)).replace("\\", "/"))
                # Move sidecar
                meta_src = expected.with_suffix(".json")
                if meta_src.exists() and meta_src.is_file():
                    meta_dest = dest.with_suffix(".json")
                    shutil.move(str(meta_src), str(meta_dest))
                    _mark_image_metadata_regenerated(meta_dest, dest)
            else:
                skipped.append({"image_file": pf, "reason": "regenerated image was not downloaded/found"})
        if generation_error and not generated_files:
            short = "\n".join([line for line in generation_error.splitlines() if line.strip()][-6:])
            raise HTTPException(status_code=500, detail=f"4:5 regeneration failed before any downloads were found. {short}")
        if generation_error:
            skipped.append({"image_file": "4:5 generation", "reason": "generator exited with errors after partial output"})

    # ── Regenerate 9:16 images ─────────────────────────────────────────
    if jobs_916:
        try:
            run_916_conversion_from_45_for_batch(
                batch=batch,
                headless=headless,
                run_dir=run_dir,
                engine=engine,
                jobs=jobs_916,
            )
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"9:16 regeneration failed: {exc}")

        for job in jobs_916:
            job_persona_slug = persona_slug(int(job["persona_number"])) if job.get("persona_number") is not None else ""
            prompt_path = str(ROOT / "output" / batch / "45" / f"{job['format']}_{job_persona_slug}_{job['language']}.txt")
            # Try to find exact prompt path
            pname = f"{job['format']}_{job_persona_slug}_{job['language']}"
            candidates = [pf for pf in prompt_files_list if pname in Path(pf).name]
            if candidates:
                prompt_path = candidates[0]
            expected = _build_expected_output_path(batch, prompt_path, "9_16", engine)
            if expected and expected.exists() and expected.is_file():
                archive_dir = GENERATED_IMAGES_ROOT / batch / "9_16" / "to_be_regenerated"
                archive_dir.mkdir(parents=True, exist_ok=True)
                dest = _unique_path(archive_dir / expected.name)
                shutil.move(str(expected), str(dest))
                generated_files.append(str(dest.relative_to(ROOT)).replace("\\", "/"))
                meta_src = expected.with_suffix(".json")
                if meta_src.exists() and meta_src.is_file():
                    meta_dest = dest.with_suffix(".json")
                    shutil.move(str(meta_src), str(meta_dest))
                    _mark_image_metadata_regenerated(meta_dest, dest)
            else:
                skipped.append({"image_file": str(job.get("prompt_96") or job.get("format") or "9:16"), "reason": "regenerated image was not downloaded/found"})

    refreshed = collect_backfill_result(run_id, batch)
    if has_storage_manifest and run_dir is not None:
        refreshed = collect_run_result(run_dir, batch, True)
        refreshed = merge_manifest(run_dir, manifest, refreshed)

    return {
        "status": "regenerated",
        "generated_files": generated_files,
        "skipped": skipped,
        "manifest": enrich_manifest_for_dashboard(refreshed),
    }


async def api_replace_image(run_id: str, image_file: str = Form(...), replacement_file: UploadFile = File(...)) -> dict[str, Any]:
    run_dir = RUNS_ROOT / run_id
    full_path = resolve_safe_path(image_file)
    generated_root = GENERATED_IMAGES_ROOT.resolve()
    if generated_root not in full_path.resolve().parents:
        raise HTTPException(status_code=400, detail="image_file must be under generated_images")
    if not full_path.exists() or not full_path.is_file():
        raise HTTPException(status_code=404, detail="Generated image not found")

    allowed = {".png", ".jpg", ".jpeg", ".webp"}
    upload_name = Path(replacement_file.filename or "").name
    upload_ext = Path(upload_name).suffix.lower()
    if upload_ext not in allowed:
        raise HTTPException(status_code=400, detail="Replacement must be png, jpg, jpeg, or webp")

    data = await replacement_file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Replacement file is empty")
    full_path.write_bytes(data)

    meta_path = full_path.with_suffix(".json")
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            meta = {}
    else:
        meta = {"type": "ad_image", "status": "success", "saved_file": str(full_path)}
    replacements = meta.setdefault("replacements", [])
    if isinstance(replacements, list):
        replacements.append(
            {
                "timestamp": int(time.time()),
                "source_filename": upload_name,
                "size_bytes": len(data),
            }
        )
    meta["status"] = "success"
    meta["saved_file"] = str(full_path)
    meta["replaced"] = True
    meta["replacement_timestamp"] = int(time.time())
    meta["replacement_source_filename"] = upload_name
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    manifest_path = run_dir / "manifest.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest.setdefault("image_files", [])
        if image_file not in manifest["image_files"]:
            manifest["image_files"].append(image_file)
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return {"status": "replaced", "image_file": image_file, "size_bytes": len(data)}


def _parse_image_naming(image_path_str: str, run_dir: Path | None) -> dict[str, str]:
    """Extract format, persona, language, concept_angle from an image's companion
    JSON metadata and build a human-readable stem for download naming.

    Stem format mirrors the canonical prompt filename (using the persona slug,
    e.g. ``HERO_stress_snacker_EN_pain_point`` for the new naming format, or
    ``HERO_P03_EN_pain_point`` for the legacy format when no slug is available):
        <FMT>_<persona>_<LANG>_<angle>[_A<NN>]_<aspect>.<ext>
    """
    full_path = ROOT / image_path_str
    meta_path = full_path.with_suffix(".json")
    legacy_meta_path = full_path.with_suffix(full_path.suffix + ".json")
    base = {"format": "UNKNOWN", "persona": "00", "lang": "EN", "concept_angle": "", "stem": "image"}
    hyp_label = ""

    if meta_path.exists() or legacy_meta_path.exists():
        try:
            if not meta_path.exists() and legacy_meta_path.exists():
                meta_path = legacy_meta_path
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            meta = {}
        fmt_value = str(meta.get("format") or meta.get("format_id") or "").strip().upper()
        persona_value = str(meta.get("persona") or meta.get("persona_id") or "").strip()
        lang_value = str(meta.get("language") or meta.get("lang") or meta.get("lang_id") or "").strip().upper()
        angle_value = str(meta.get("concept_angle") or "").strip()
        if fmt_value:
            base["format"] = fmt_value
        if persona_value:
            # Prefer the persona slug (new format, e.g. "stress_snacker" or
            # "Always Hungry"). Fall back to P<NN> only if the metadata only
            # carries the legacy numeric form.
            if re.fullmatch(r"P\d+", persona_value, flags=re.IGNORECASE):
                base["persona"] = persona_value.upper()
            else:
                base["persona"] = persona_value.lower()
        if lang_value:
            base["lang"] = lang_value
        if angle_value:
            base["concept_angle"] = angle_value
        prompt_file = str(meta.get("prompt_file_relative") or meta.get("prompt_file") or "").strip().replace("\\", "/")
        prompt_slug = _extract_persona_slug_from_prompt_filename(prompt_file)
        if prompt_slug:
            base["persona"] = prompt_slug
        parsed = parse_prompt_filename_full(prompt_file)
        if parsed:
            fmt, lang, persona_num, angle, _variant = parsed
            base["format"] = fmt
            # Only override persona with P<NN> if we don't already have a slug
            # from the prompt filename (slug is the canonical, human-readable
            # form; P<NN> is a legacy fallback).
            if not prompt_slug:
                base["persona"] = f"P{persona_num:02d}" if persona_num else "P00"
            base["lang"] = lang
            if angle:
                base["concept_angle"] = angle
        creative_total = int(meta.get("creative_total") or 1) if str(meta.get("creative_total") or "1").isdigit() else 1
        creative_index = int(meta.get("creative_index") or 1) if str(meta.get("creative_index") or "1").isdigit() else 1
        if creative_total > 1:
            base["creative_suffix"] = f"_A{creative_index:02d}"

        if not hyp_label:
            htype = str(meta.get("hypothesis_type") or "")
            hvar = str(meta.get("hypothesis_variant") or "")
            if htype and htype != "none":
                parts = [htype]
                if hvar:
                    parts.append(hvar)
                hyp_label = "_" + "_".join(parts)

    if base["format"] == "UNKNOWN" or base["persona"] in {"00", "P00"}:
        name = Path(image_path_str).stem.lower()
        match = re.search(r"(?:gemini|chatgpt)-(?P<fmt>[a-z0-9]+)-p(?P<num>\d+)-(?P<lang>[a-z0-9]+)(?:-a(?P<creative>\d+))?(?:-(?P<angle>[a-z_]+))?", name)
        if match:
            base["format"] = match.group("fmt").upper()
            base["persona"] = f"P{int(match.group('num')):02d}"
            base["lang"] = match.group("lang").upper()
            if match.group("creative"):
                base["creative_suffix"] = f"_A{int(match.group('creative')):02d}"
            if match.group("angle"):
                base["concept_angle"] = match.group("angle")

    # Try hypothesis
    if run_dir is not None:
        hyp_path = run_dir / "context" / "hypothesis_config.json"
        if hyp_path.exists():
            try:
                hyp_cfg = json.loads(hyp_path.read_text(encoding="utf-8"))
                htype = hyp_cfg.get("type", "")
                hvar = hyp_cfg.get("variant", "")
                if htype and htype != "none":
                    parts = [htype]
                    if hvar:
                        parts.append(hvar)
                    hyp_label = "_" + "_".join(parts)
            except Exception:
                pass

    ext = Path(image_path_str).suffix
    angle_part = f"_{base['concept_angle']}" if base.get("concept_angle") else ""
    aspect_part = ""
    if "/9_16/" in image_path_str.replace("\\", "/"):
        aspect_part = "_9_16"
    elif "/4_5/" in image_path_str.replace("\\", "/"):
        aspect_part = "_4_5"
    stem = f"{base['format']}_{base['persona']}_{base['lang']}{angle_part}{base.get('creative_suffix', '')}{hyp_label}{aspect_part}"
    base["stem"] = stem
    base["ext"] = ext
    base["aspect"] = aspect_part.lstrip("_") if aspect_part else ""
    return base


def _build_persona_name_map(run_dir: Path) -> dict[str, str]:
    """Map persona number (P01) to persona name from run's copy_batch.json."""
    copy_path = run_dir / "context" / "copy_batch.json"
    if not copy_path.exists():
        return {}
    try:
        data = json.loads(copy_path.read_text(encoding="utf-8"))
        ads = data.get("ads") if isinstance(data, dict) else []
        if not isinstance(ads, list):
            return {}
        mapping: dict[str, str] = {}
        for ad in ads:
            p = ad.get("persona") if isinstance(ad, dict) else {}
            if not isinstance(p, dict):
                continue
            num = p.get("number")
            name = p.get("name") or p.get("persona_name") or ""
            if isinstance(num, int) and name:
                mapping[f"P{num:02d}"] = str(name)
        return mapping
    except Exception:
        return {}


def _clean_metadata_for_download(meta: dict[str, Any], img_path: str, run_dir: Path | None) -> dict[str, Any]:
    """Strip excessive internal keys from image metadata and enrich with
    hypothesis info, persona name, and clean format labels for download ZIP."""
    clean = dict(meta)

    # Strip internal plumbing
    for key in ("generated_image_src", "saved_ext", "output_dir", "metadata_file", "type"):
        clean.pop(key, None)

    # Normalise key names
    if "format" not in clean and "format_id" in clean:
        clean["format"] = clean.pop("format_id")
    if clean.get("format_id"):
        clean.pop("format_id", None)
    if "persona" not in clean and "persona_id" in clean:
        clean["persona"] = clean.pop("persona_id")
    if clean.get("persona_id"):
        clean.pop("persona_id", None)
    if "language" not in clean and "lang_id" in clean:
        clean["language"] = clean.pop("lang_id")
    if clean.get("lang_id"):
        clean.pop("lang_id", None)

    # Ensure hypothesis keys are always present
    hyp_type = clean.get("hypothesis_type") or ""
    hyp_var = clean.get("hypothesis_variant") or ""
    clean["hypothesis_type"] = hyp_type
    clean["hypothesis_variant"] = hyp_var

    # Enrich with persona name if we have a run_dir
    if run_dir is not None:
        persona_val = clean.get("persona", "")
        if persona_val:
            mapping = _build_persona_name_map(run_dir)
            if persona_val in mapping:
                clean["persona_name"] = mapping[persona_val]

    return clean


def api_download_single_image(run_id: str, image_file: str):
    """Return a zip containing the image file and its metadata JSON."""
    run_dir = RUNS_ROOT / run_id
    full_path = ROOT / image_file
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="Image file not found")

    meta_path = full_path.with_suffix(".json")
    legacy_meta_path = full_path.with_suffix(full_path.suffix + ".json")

    import io, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(full_path, full_path.name)
        meta_content = {"source": image_file}
        if meta_path.exists() or legacy_meta_path.exists():
            try:
                if not meta_path.exists() and legacy_meta_path.exists():
                    meta_path = legacy_meta_path
                meta_content = json.loads(meta_path.read_text(encoding="utf-8"))
            except Exception:
                pass
        meta_content = _clean_metadata_for_download(meta_content, image_file, run_dir)
        meta_content["_download_name"] = full_path.stem
        zf.writestr(f"{full_path.stem}_metadata.json", json.dumps(meta_content, ensure_ascii=False, indent=2))

    buf.seek(0)
    return StreamingResponse(buf, media_type="application/zip",
                             headers={"Content-Disposition": f'attachment; filename="{full_path.stem}.zip"'})


def api_download_batch_images(run_id: str):
    """Return a zip of all images grouped by VN subfolders with metadata.
    Always scans the filesystem directly so newly generated images
    (e.g. 9:16 added after the manifest was saved) are included."""
    run_dir = RUNS_ROOT / run_id

    # Refresh cached thumbnail summary before scanning
    batch_label = run_id
    manifest_path = run_dir / "manifest.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        batch_label = manifest.get("batch", run_id)

    # Always scan the filesystem — manifest may be stale
    image_files = scan_image_files_for_batch(batch_label) if batch_label != run_id else []

    import io, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        vns: set[str] = set()
        for img_path in image_files:
            full_path = ROOT / img_path
            if not full_path.exists():
                continue

            vn = _extract_vn_from_image_path(img_path) or batch_label or "images"
            vns.add(vn)
            aspect = _extract_aspect_from_image_path(img_path)
            meta_path = full_path.with_suffix(".json")
            legacy_meta_path = full_path.with_suffix(full_path.suffix + ".json")

            folder = f"{vn}/{aspect}" if aspect else vn
            zf.write(full_path, f"{folder}/{full_path.name}")

            meta_content = {"source": img_path}
            if meta_path.exists() or legacy_meta_path.exists():
                try:
                    if not meta_path.exists() and legacy_meta_path.exists():
                        meta_path = legacy_meta_path
                    meta_content = json.loads(meta_path.read_text(encoding="utf-8"))
                except Exception:
                    pass
            meta_content = _clean_metadata_for_download(meta_content, img_path, run_dir)
            meta_content["_download_name"] = full_path.stem
            zf.writestr(f"{folder}/{full_path.stem}_metadata.json",
                        json.dumps(meta_content, ensure_ascii=False, indent=2))

        if not image_files:
            zf.writestr("README.txt",
                        "No generated images found for this run.\n"
                        "Run image generation first, then try again.")

    buf.seek(0)
    label = "_".join(sorted(vns)) if vns else (batch_label if batch_label != run_id else run_id)
    return StreamingResponse(buf, media_type="application/zip",
                             headers={"Content-Disposition": f'attachment; filename="batch_{label}.zip"'})


def api_download_batches(batch_names: list[str]):
    """Return a zip of all images for given batch names, grouped by VN folder."""
    image_files_by_vn: dict[str, list[str]] = {}
    for batch_name in batch_names:
        files = scan_image_files_for_batch(batch_name)
        if files:
            image_files_by_vn[batch_name] = files

    import io, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for vn, files in image_files_by_vn.items():
            for img_path in files:
                full_path = ROOT / img_path
                if not full_path.exists():
                    continue
                aspect = _extract_aspect_from_image_path(img_path)
                meta_path = full_path.with_suffix(".json")
                legacy_meta_path = full_path.with_suffix(full_path.suffix + ".json")
                folder = f"{vn}/{aspect}" if aspect else vn
                zf.write(full_path, f"{folder}/{full_path.name}")
                meta_content = {"source": img_path}
                if meta_path.exists() or legacy_meta_path.exists():
                    try:
                        if not meta_path.exists() and legacy_meta_path.exists():
                            meta_path = legacy_meta_path
                        meta_content = json.loads(meta_path.read_text(encoding="utf-8"))
                    except Exception:
                        pass
                meta_content = _clean_metadata_for_download(meta_content, img_path, None)
                meta_content["_download_name"] = full_path.stem
                zf.writestr(f"{folder}/{full_path.stem}_metadata.json",
                            json.dumps(meta_content, ensure_ascii=False, indent=2))

        if not image_files_by_vn:
            zf.writestr("README.txt",
                        "No generated images found for selected batch(es).\n"
                        "Run image generation first, then try again.")

    buf.seek(0)
    label = "_".join(batch_names) if batch_names else "batches"
    return StreamingResponse(buf, media_type="application/zip",
                             headers={"Content-Disposition": f'attachment; filename="{label}.zip"'})

def api_llm_traces(limit: int = 50, offset: int = 0, run_id_filter: str | None = None) -> dict[str, Any]:
    LLM_TRACES_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(LLM_TRACES_DIR.glob("*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
    all_traces: list[tuple[str, dict]] = []
    for f in files:
        try:
            trace = json.loads(f.read_text(encoding="utf-8"))
            if run_id_filter and trace.get("run_id") != run_id_filter:
                continue
            trace["_file"] = f.name
            all_traces.append(trace)
        except (json.JSONDecodeError, OSError):
            continue
    total = len(all_traces)
    page = all_traces[offset:offset + limit]
    return {"traces": page, "total": total, "offset": offset, "limit": limit}


def api_delete_llm_traces(run_id_filter: str | None = None, trace_files: list[str] | None = None) -> dict[str, Any]:
    LLM_TRACES_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(LLM_TRACES_DIR.glob("*.json"), key=lambda f: f.st_mtime)
    deleted = 0
    for f in files:
        if trace_files:
            if f.name not in trace_files:
                continue
        elif run_id_filter:
            try:
                trace = json.loads(f.read_text(encoding="utf-8"))
                if trace.get("run_id") != run_id_filter:
                    continue
            except (json.JSONDecodeError, OSError):
                continue
        f.unlink()
        deleted += 1
    return {"deleted": deleted, "filter": run_id_filter, "trace_files": trace_files}


def api_delete_llm_traces_by_files(files: list[str]) -> dict[str, Any]:
    LLM_TRACES_DIR.mkdir(parents=True, exist_ok=True)
    deleted = 0
    for fname in files:
        f = LLM_TRACES_DIR / fname
        if f.exists() and f.suffix == ".json":
            f.unlink()
            deleted += 1
    return {"deleted": deleted, "files": files}


# ── DB-backed prompt / image list endpoints ──────────────────────────────


def api_run_prompts(user_id: str, run_id: str, limit: int = 200, offset: int = 0) -> dict[str, Any]:
    try:
        from dashboard.backend.db.client import get_sync_db
        from dashboard.backend.db.collections import COLL_PROMPTS
        docs = list(
            get_sync_db()[COLL_PROMPTS]
            .find({"user_id": user_id, "run_id": run_id})
            .sort("created_at", 1)
            .skip(offset)
            .limit(limit)
        )
        for d in docs:
            d.pop("_id", None)
            d.pop("content", None)
        return {"prompts": docs, "total": len(docs), "run_id": run_id}
    except Exception:
        return {"prompts": [], "total": 0, "run_id": run_id}


def api_run_images(user_id: str, run_id: str, limit: int = 200, offset: int = 0) -> dict[str, Any]:
    try:
        from dashboard.backend.db.client import get_sync_db
        from dashboard.backend.db.collections import COLL_IMAGES
        docs = list(
            get_sync_db()[COLL_IMAGES]
            .find({"user_id": user_id, "run_id": run_id})
            .sort("created_at", 1)
            .skip(offset)
            .limit(limit)
        )
        for d in docs:
            d.pop("_id", None)
        return {"images": docs, "total": len(docs), "run_id": run_id}
    except Exception:
        return {"images": [], "total": 0, "run_id": run_id}


@app.get("/api/runs/{run_id}/prompts")
def get_run_prompts(run_id: str, request: Request, limit: int = 200, offset: int = 0) -> dict[str, Any]:
    user = _get_user_from_request(request)
    _check_ownership(run_id, user["user_id"])
    return api_run_prompts(user["user_id"], run_id, limit=limit, offset=offset)


@app.get("/api/runs/{run_id}/images")
def get_run_images(run_id: str, request: Request, limit: int = 200, offset: int = 0) -> dict[str, Any]:
    user = _get_user_from_request(request)
    _check_ownership(run_id, user["user_id"])
    return api_run_images(user["user_id"], run_id, limit=limit, offset=offset)


# ── Modular routes ───────────────────────────────────────────────────────────
from dashboard.backend.routes import defaults, progress, runs, generate, batch, export_import, execute, chrome

app.include_router(defaults.router)
app.include_router(progress.router)
app.include_router(runs.router)
app.include_router(generate.router)
app.include_router(batch.router)
app.include_router(export_import.router)
app.include_router(execute.router)
app.include_router(chrome.router)

# ── Cloud/multi-user routes ─────────────────────────────────────────────────
from dashboard.backend.auth.routes import router as auth_router
from dashboard.backend.services.provider_routes import router as provider_router
from dashboard.backend.services.blob_routes import router as blob_router
from dashboard.backend.agent.routes import router as agent_router

app.include_router(auth_router)
app.include_router(provider_router)
app.include_router(blob_router)
app.include_router(agent_router)

# ── Authenticated file download endpoints (production-safe) ────────────────
from fastapi.responses import FileResponse
from fastapi import Cookie


def _get_user_from_request(request: Request) -> dict[str, Any]:
    if app_settings.is_production:
        user = getattr(request.state, "user", None)
        if user is None:
            raise HTTPException(status_code=401, detail="Not authenticated")
        return user
    return {"user_id": "dev_user", "is_admin": True}


def _check_ownership(run_id: str, user_id: str) -> None:
    """Verify a user owns a run. In dev mode, allow if no owner recorded."""
    owner = _get_run_owner(run_id)
    if owner is None:
        if app_settings.is_production:
            raise HTTPException(status_code=403, detail="Run ownership unknown")
        return
    if owner != user_id:
        raise HTTPException(status_code=403, detail="Access denied: run belongs to another user")


def _register_download_routes():
    """Register download endpoints.

    New DB-backed endpoints are registered first so they take priority over
    the legacy path-based endpoints (which use {param:path} catch-alls).
    """
    from bson.objectid import ObjectId

    # ── New DB-backed endpoints (production-safe) ──────────────────────

    @app.get("/api/files/download/run/{run_id}/{file_id:path}")
    def download_run_file_new(
        run_id: str,
        file_id: str,
        request: Request,
    ):
        user = _get_user_from_request(request)
        _check_ownership(run_id, user["user_id"])
        known_files = {
            "manifest": "manifest.json",
            "run_context": "context/run_context.json",
            "copy_batch": "context/copy_batch.json",
            "hypothesis_config": "context/hypothesis_config.json",
            "visual_pattern_reuse": "context/visual_pattern_reuse.json",
            "background_reuse": "context/background_reuse.json",
        }
        relative = known_files.get(file_id, file_id)
        safe_path = resolve_safe_path(f"dashboard_storage/runs/{run_id}/{relative}")
        if not safe_path.exists():
            raise HTTPException(status_code=404, detail="File not found in run")
        target = safe_path.resolve()
        runs_root = RUNS_ROOT.resolve()
        run_dir_resolved = (RUNS_ROOT / run_id).resolve()
        if runs_root not in target.parents and target != runs_root:
            raise HTTPException(status_code=403, detail="Access denied")
        if run_dir_resolved not in target.parents and target != run_dir_resolved:
            raise HTTPException(status_code=403, detail="File outside run directory")
        if target.is_file():
            return FileResponse(target, filename=target.name)
        raise HTTPException(status_code=400, detail="Path is not a file")

    @app.get("/api/files/download/image/{image_id}")
    def download_image_by_id(
        image_id: str,
        request: Request,
    ):
        user = _get_user_from_request(request)
        try:
            from dashboard.backend.db.client import get_sync_db
            from dashboard.backend.db.collections import COLL_IMAGES
            doc = get_sync_db()[COLL_IMAGES].find_one({"image_id": image_id})
        except Exception:
            doc = None
        if not doc:
            raise HTTPException(status_code=404, detail="Image not found")
        if doc.get("user_id") != user["user_id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        # Cloudinary: redirect to secure_url
        secure_url = doc.get("secure_url") or ""
        if secure_url:
            from fastapi.responses import RedirectResponse
            return RedirectResponse(secure_url)
        # Legacy storage_url fallback
        storage_url = doc.get("storage_url") or ""
        if storage_url:
            from fastapi.responses import RedirectResponse
            return RedirectResponse(storage_url)
        # Local file fallback
        local_path = doc.get("local_path") or ""
        if local_path:
            img_path = ROOT / local_path
            if img_path.exists():
                return FileResponse(img_path, filename=doc.get("filename", "image.png"))
        raise HTTPException(status_code=404, detail="Image file not available")

    @app.get("/api/files/download/prompt/{prompt_id}")
    def download_prompt_by_id(
        prompt_id: str,
        request: Request,
    ):
        user = _get_user_from_request(request)
        try:
            from dashboard.backend.db.client import get_sync_db
            from dashboard.backend.db.collections import COLL_PROMPTS
            doc = get_sync_db()[COLL_PROMPTS].find_one({"_id": ObjectId(prompt_id)})
        except Exception:
            doc = None
        if not doc:
            raise HTTPException(status_code=404, detail="Prompt not found")
        if doc.get("user_id") != user["user_id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        content = doc.get("content", "")
        filename = doc.get("filename", "prompt.txt")
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse(content, headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    # ── Legacy path-based endpoints (dev-only in production) ───────────

    @app.get("/api/files/download/run/{run_id:path}")
    def download_run_file_legacy(
        run_id: str,
        request: Request,
    ):
        user = _get_user_from_request(request)
        _check_ownership(run_id, user["user_id"])
        if app_settings.is_production:
            raise HTTPException(status_code=403, detail="Use /api/files/download/run/{run_id}/{file_id} in production")
        safe_path = resolve_safe_path(f"dashboard_storage/runs/{run_id}")
        if not safe_path.exists():
            raise HTTPException(status_code=404, detail="Run not found")
        target = safe_path.resolve()
        runs_root = RUNS_ROOT.resolve()
        if runs_root not in target.parents and target != runs_root:
            raise HTTPException(status_code=403, detail="Access denied")
        if target.is_file():
            return FileResponse(target, filename=target.name)
        raise HTTPException(status_code=400, detail="Path is not a file")

    @app.get("/api/files/download/generated/{path:path}")
    def download_generated_file_legacy(
        path: str,
        request: Request,
    ):
        user = _get_user_from_request(request)
        if app_settings.is_production:
            raise HTTPException(status_code=403, detail="Use /api/files/download/image/{image_id} in production")
        run_id = _extract_run_id_from_generated_path(path)
        if run_id:
            _check_ownership(run_id, user["user_id"])
        elif app_settings.is_production:
            raise HTTPException(status_code=403, detail="Cannot determine run ownership from path")
        safe_path = resolve_safe_path(f"generated_images/{path}")
        if not safe_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
        target = safe_path.resolve()
        images_root = GENERATED_IMAGES_ROOT.resolve()
        if images_root not in target.parents:
            raise HTTPException(status_code=403, detail="Access denied")
        if target.is_file():
            return FileResponse(target, filename=target.name)
        raise HTTPException(status_code=400, detail="Path is not a file")

    @app.get("/api/files/download/output/{path:path}")
    def download_output_file_legacy(
        path: str,
        request: Request,
    ):
        user = _get_user_from_request(request)
        if app_settings.is_production:
            raise HTTPException(status_code=403, detail="Use /api/files/download/run/{run_id}/copy_batch in production")
        run_id = _extract_run_id_from_output_path(path)
        if run_id:
            _check_ownership(run_id, user["user_id"])
        elif app_settings.is_production:
            raise HTTPException(status_code=403, detail="Cannot determine run ownership from output path")
        safe_path = resolve_safe_path(f"output/{path}")
        if not safe_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
        target = safe_path.resolve()
        output_root = (ROOT / "output").resolve()
        if output_root not in target.parents:
            raise HTTPException(status_code=403, detail="Access denied")
        if target.is_file():
            return FileResponse(target, filename=target.name)
        raise HTTPException(status_code=400, detail="Path is not a file")


_register_download_routes()


INPUT_ROOT = ROOT / "input"

if app_settings.is_dev:
    # Dev: mount local folders directly for convenience
    app.mount("/storage", StaticFiles(directory=str(STORAGE_ROOT)), name="storage")
    app.mount("/output", StaticFiles(directory=str(ROOT / "output")), name="output")
    GENERATED_IMAGES_ROOT.mkdir(parents=True, exist_ok=True)
    app.mount("/generated_images", StaticFiles(directory=str(GENERATED_IMAGES_ROOT)), name="generated_images")
    print("[startup] Local dev mode: /storage, /output, /generated_images mounted publicly")
    # /input contains seed files stored in the repo
    app.mount("/input", StaticFiles(directory=str(INPUT_ROOT)), name="input")
else:
    # Production: do NOT mount user data folders publicly
    print("[startup] Production mode: user data folders not publicly mounted")


@app.get("/api/seeds")
def list_seed_files(request: Request) -> dict[str, list[dict[str, str]]]:
    """List available seed file names and paths (safe for production)."""
    files: list[dict[str, str]] = []
    if INPUT_ROOT.is_dir():
        for f in sorted(INPUT_ROOT.rglob("*")):
            if f.is_file() and not f.name.startswith("."):
                rel = f.relative_to(INPUT_ROOT)
                files.append({"name": f.name, "path": str(rel.as_posix())})
    return {"files": files}


@app.get("/api/seeds/download/{path:path}")
def download_seed_file(path: str, request: Request) -> FileResponse:
    user = _get_user_from_request(request)
    safe_path = resolve_safe_path(str(INPUT_ROOT / path))
    if not safe_path.exists():
        raise HTTPException(status_code=404, detail="Seed file not found")
    target = safe_path.resolve()
    if INPUT_ROOT.resolve() not in target.parents:
        raise HTTPException(status_code=403, detail="Access denied")
    if not target.is_file():
        raise HTTPException(status_code=400, detail="Path is not a file")
    return FileResponse(target, filename=target.name)


@app.get("/api/files/input/{path:path}")
def download_input_file(path: str, request: Request) -> FileResponse:
    user = _get_user_from_request(request)
    if app_settings.is_dev:
        safe_path = resolve_safe_path(str(INPUT_ROOT / path))
    else:
        safe_path = resolve_safe_path(str(INPUT_ROOT / path))
        if not safe_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
    target = safe_path.resolve()
    if INPUT_ROOT.resolve() not in target.parents:
        raise HTTPException(status_code=403, detail="Access denied")
    if not target.is_file():
        raise HTTPException(status_code=400, detail="Path is not a file")
    return FileResponse(target, filename=target.name)
app.mount("/", StaticFiles(directory=str(ROOT / "dashboard" / "frontend"), html=True), name="frontend")
